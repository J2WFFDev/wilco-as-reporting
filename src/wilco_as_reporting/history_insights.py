"""Coach-facing insights built from local Phase 6A history tables."""

from __future__ import annotations

import csv
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

IDENTITY_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "season_label",
    "issue_type",
    "issue_message",
    "affected_rows",
    "recommendation",
)

IMPROVEMENT_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "season_label",
    "matches_count",
    "scored_matches_count",
    "first_score",
    "latest_score",
    "best_score",
    "average_score",
    "median_score",
    "improvement_seconds",
    "improvement_percent",
    "confidence_level",
    "confidence_note",
    "insight_note",
)

REGRESSION_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "season_label",
    "matches_count",
    "first_score",
    "latest_score",
    "best_score",
    "median_score",
    "regression_seconds",
    "regression_percent",
    "confidence_level",
    "possible_reason",
    "coach_note",
)

ROSTER_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "season_label",
    "disciplines_count",
    "scored_entries_count",
    "best_discipline",
    "strongest_discipline_by_score",
    "most_recent_match",
    "most_recent_match_date",
    "has_2026_nationals_entry",
    "current_status_note",
)

DISCIPLINE_COLUMNS = (
    "season_label",
    "discipline",
    "athletes_count",
    "scored_entries_count",
    "team_best_score",
    "team_median_score",
    "team_average_score",
    "prior_season_median_score",
    "median_change_seconds",
    "trend_direction",
    "strength_note",
    "coach_priority",
)

BRIEF_COLUMNS = (
    "section",
    "insight",
    "supporting_metric",
    "confidence_level",
    "coach_action",
)

QUALITY_NOTE_COLUMNS = (
    "topic",
    "finding",
    "affected_rows",
    "coach_interpretation",
)

OUTPUTS = (
    ("wilco_identity_issues.csv", IDENTITY_COLUMNS),
    ("wilco_improvement_leaderboard.csv", IMPROVEMENT_COLUMNS),
    ("wilco_regression_watchlist.csv", REGRESSION_COLUMNS),
    ("wilco_current_season_roster.csv", ROSTER_COLUMNS),
    ("wilco_discipline_insights.csv", DISCIPLINE_COLUMNS),
    ("wilco_coach_brief.csv", BRIEF_COLUMNS),
    ("wilco_insight_data_quality_notes.csv", QUALITY_NOTE_COLUMNS),
)

PLACEHOLDER_ATHLETE_ID = "9999"


class HistoryInsightsError(RuntimeError):
    """Raised when Phase 6A history inputs cannot produce insights."""


@dataclass(frozen=True)
class HistoryInsightsResult:
    history_dir: Path
    workbook_path: Path | None
    season: str
    row_counts: dict[str, int]
    identity_issue_rows: int
    current_season_athletes: int
    top_improvements: tuple[dict[str, Any], ...]
    top_watchlist: tuple[dict[str, Any], ...]


def build_history_insights(
    *,
    output_root: Path | str,
    team_key: str,
    history_dir: Path | str | None = None,
    season: str | None = None,
    min_scored_matches: int = 2,
    current_season_only: bool = False,
    include_placeholders: bool = False,
    workbook: bool = True,
) -> HistoryInsightsResult:
    """Build local coach insights without calling the SASP API."""
    if team_key.casefold() != "wilco":
        raise HistoryInsightsError(
            "Phase 6B currently supports team_key 'wilco'."
        )
    if min_scored_matches < 2:
        raise HistoryInsightsError(
            "min_scored_matches must be at least 2."
        )
    root = Path(output_root)
    source_dir = Path(history_dir) if history_dir else root / "history"
    participation = _read_required(
        source_dir / "wilco_match_participation.csv"
    )
    athlete_history = _read_required(
        source_dir / "wilco_athlete_discipline_history.csv"
    )
    source_matches = _read_required(
        source_dir / "history_source_matches.csv"
    )
    quality = _read_required(
        source_dir / "wilco_data_quality_summary.csv"
    )

    found_seasons = {
        row.get("season_label", "")
        for row in participation
        if row.get("season_label")
        and row.get("season_label") != "Unknown Season"
    }
    selected_season = season or _latest_season(found_seasons)
    if selected_season not in found_seasons:
        raise HistoryInsightsError(
            f"Season {selected_season!r} was not found in Phase 6A outputs."
        )

    identity_issues = _identity_issues(participation)
    identity_affected_rows = sum(
        _is_placeholder(row) for row in participation
    )
    insight_history = [
        row
        for row in athlete_history
        if include_placeholders or not _is_placeholder(row)
    ]
    if season or current_season_only:
        insight_history = [
            row
            for row in insight_history
            if row.get("season_label") == selected_season
        ]
    valid_participation = [
        row
        for row in participation
        if include_placeholders or not _is_placeholder(row)
    ]

    improvements = _improvement_rows(
        insight_history,
        min_scored_matches,
    )
    regressions = _regression_rows(
        insight_history,
        min_scored_matches,
    )
    roster = _current_season_roster(
        valid_participation,
        selected_season,
    )
    discipline_insights = _discipline_insights(
        valid_participation,
        selected_season,
    )
    quality_notes = _quality_notes(
        identity_affected_rows,
        source_matches,
        quality,
        selected_season,
    )
    brief = _coach_brief(
        selected_season,
        roster,
        improvements,
        regressions,
        discipline_insights,
        identity_issues,
        identity_affected_rows,
        source_matches,
    )

    rows_by_file = {
        "wilco_identity_issues.csv": identity_issues,
        "wilco_improvement_leaderboard.csv": improvements,
        "wilco_regression_watchlist.csv": regressions,
        "wilco_current_season_roster.csv": roster,
        "wilco_discipline_insights.csv": discipline_insights,
        "wilco_coach_brief.csv": brief,
        "wilco_insight_data_quality_notes.csv": quality_notes,
    }
    for filename, columns in OUTPUTS:
        _write_csv(source_dir / filename, columns, rows_by_file[filename])
    workbook_path = (
        _build_workbook(source_dir, rows_by_file)
        if workbook
        else None
    )
    return HistoryInsightsResult(
        history_dir=source_dir,
        workbook_path=workbook_path,
        season=selected_season,
        row_counts={
            filename: len(rows_by_file[filename])
            for filename, _ in OUTPUTS
        },
        identity_issue_rows=identity_affected_rows,
        current_season_athletes=len(roster),
        top_improvements=tuple(
            row
            for row in improvements
            if row["confidence_level"] == "high"
        )[:5],
        top_watchlist=tuple(regressions[:5]),
    )


def _identity_issues(
    participation: list[dict[str, str]],
) -> list[dict[str, Any]]:
    grouped: dict[
        tuple[str, str, str, str, str],
        int,
    ] = defaultdict(int)
    for row in participation:
        common = (
            row.get("athlete_name", "").strip(),
            row.get("athlete_id", "").strip(),
            row.get("discipline", ""),
            row.get("season_label", ""),
        )
        if not common[0]:
            grouped[common + ("blank_athlete_name",)] += 1
        if common[1] == PLACEHOLDER_ATHLETE_ID:
            grouped[common + ("placeholder_athlete_id",)] += 1
    rows: list[dict[str, Any]] = []
    for key, affected_rows in sorted(grouped.items()):
        issue_type = key[4]
        if issue_type == "blank_athlete_name":
            message = "Athlete name is blank in historical participation."
            recommendation = (
                "Quarantine from coach insights until identity is resolved."
            )
        else:
            message = "Athlete ID 9999 is treated as a placeholder."
            recommendation = (
                "Confirm the athlete identity before using this row."
            )
        rows.append(
            {
                "athlete_name": key[0],
                "athlete_id": key[1],
                "discipline": key[2],
                "season_label": key[3],
                "issue_type": issue_type,
                "issue_message": message,
                "affected_rows": affected_rows,
                "recommendation": recommendation,
            }
        )
    return rows


def _improvement_rows(
    athlete_history: list[dict[str, str]],
    min_scored_matches: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in athlete_history:
        scored = _integer(source.get("scored_matches_count"))
        first = _number(source.get("first_score"))
        latest = _number(source.get("latest_score"))
        improvement = _number(source.get("improvement_seconds"))
        if (
            scored < min_scored_matches
            or first is None
            or latest is None
            or improvement is None
            or improvement <= 0
        ):
            continue
        percent = improvement / first * 100 if first else None
        confidence, confidence_note, concern = _confidence(
            scored,
            improvement,
            percent,
        )
        insight_note = (
            f"Improved {improvement:.3f} seconds from first to latest."
        )
        if concern:
            insight_note += f" {concern}"
        rows.append(
            {
                "athlete_name": source.get("athlete_name", ""),
                "athlete_id": source.get("athlete_id", ""),
                "discipline": source.get("discipline", ""),
                "season_label": source.get("season_label", ""),
                "matches_count": _integer(source.get("matches_count")),
                "scored_matches_count": scored,
                "first_score": _display(first),
                "latest_score": _display(latest),
                "best_score": source.get("best_score", ""),
                "average_score": source.get("average_score", ""),
                "median_score": source.get("median_score", ""),
                "improvement_seconds": _display(improvement),
                "improvement_percent": _display(percent),
                "confidence_level": confidence,
                "confidence_note": confidence_note,
                "insight_note": insight_note,
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            {"high": 0, "medium": 1, "low": 2}[row["confidence_level"]],
            -float(row["improvement_seconds"]),
            row["athlete_name"],
            row["discipline"],
        ),
    )


def _regression_rows(
    athlete_history: list[dict[str, str]],
    min_scored_matches: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in athlete_history:
        scored = _integer(source.get("scored_matches_count"))
        first = _number(source.get("first_score"))
        latest = _number(source.get("latest_score"))
        best = _number(source.get("best_score"))
        median = _number(source.get("median_score"))
        if (
            scored < min_scored_matches
            or latest is None
            or median is None
            or best is None
        ):
            continue
        regression = latest - median
        percent = regression / median * 100 if median else None
        if regression < max(1.0, median * 0.02):
            continue
        if scored == 2:
            confidence = "low"
            reason = "Only two scored matches; the latest swing may be noise."
            coach_note = "limited data"
        elif regression >= 10 or (percent is not None and percent >= 20):
            confidence = "medium"
            reason = "Latest score is far above the season median."
            coach_note = "likely outlier"
        elif scored == 3:
            confidence = "medium"
            reason = "Three scored matches show a possible upward-time trend."
            coach_note = "watch trend"
        else:
            confidence = "high"
            reason = "Latest score is meaningfully slower than the median."
            coach_note = "watch trend"
        rows.append(
            {
                "athlete_name": source.get("athlete_name", ""),
                "athlete_id": source.get("athlete_id", ""),
                "discipline": source.get("discipline", ""),
                "season_label": source.get("season_label", ""),
                "matches_count": _integer(source.get("matches_count")),
                "first_score": _display(first),
                "latest_score": _display(latest),
                "best_score": _display(best),
                "median_score": _display(median),
                "regression_seconds": _display(regression),
                "regression_percent": _display(percent),
                "confidence_level": confidence,
                "possible_reason": reason,
                "coach_note": coach_note,
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            {"high": 0, "medium": 1, "low": 2}[row["confidence_level"]],
            -float(row["regression_seconds"]),
            row["athlete_name"],
        ),
    )


def _current_season_roster(
    participation: list[dict[str, str]],
    season: str,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in participation:
        if row.get("season_label") == season:
            grouped[
                (
                    row.get("athlete_name", ""),
                    row.get("athlete_id", ""),
                )
            ].append(row)
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        scored = [
            row
            for row in rows
            if _number(row.get("score")) is not None
        ]
        by_discipline: dict[str, list[float]] = defaultdict(list)
        for row in scored:
            score = _number(row.get("score"))
            if score is not None:
                by_discipline[row.get("discipline", "")].append(score)
        best_discipline = ""
        strongest_discipline = ""
        if by_discipline:
            best_discipline = min(
                by_discipline,
                key=lambda value: min(by_discipline[value]),
            )
            strongest_discipline = min(
                by_discipline,
                key=lambda value: statistics.median(
                    by_discipline[value]
                ),
            )
        recent = max(
            rows,
            key=lambda row: (
                row.get("match_date", ""),
                _integer(row.get("match_id")),
            ),
        )
        nationals = any(row.get("match_id") == "671" for row in rows)
        if not scored:
            status = "Registered/participating; no scored results available."
        elif nationals:
            status = (
                "Current-season scores available; Nationals entry present."
            )
        else:
            status = "Current-season scored participation is available."
        results.append(
            {
                "athlete_name": key[0],
                "athlete_id": key[1],
                "season_label": season,
                "disciplines_count": len(
                    {
                        row.get("discipline", "")
                        for row in rows
                        if row.get("discipline")
                    }
                ),
                "scored_entries_count": len(scored),
                "best_discipline": best_discipline,
                "strongest_discipline_by_score": strongest_discipline,
                "most_recent_match": recent.get("match_name", ""),
                "most_recent_match_date": recent.get("match_date", ""),
                "has_2026_nationals_entry": _boolean(nationals),
                "current_status_note": status,
            }
        )
    return sorted(results, key=lambda row: row["athlete_name"])


def _discipline_insights(
    participation: list[dict[str, str]],
    current_season: str,
) -> list[dict[str, Any]]:
    seasons = sorted(
        {
            row.get("season_label", "")
            for row in participation
            if row.get("season_label")
            and row.get("season_label") != "Unknown Season"
        },
        key=_season_start,
    )
    current_index = seasons.index(current_season)
    prior_season = seasons[current_index - 1] if current_index else ""
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in participation:
        score = _number(row.get("score"))
        if score is not None:
            grouped[
                (
                    row.get("season_label", ""),
                    row.get("discipline", ""),
                )
            ].append(row)
    current_disciplines = sorted(
        discipline
        for season, discipline in grouped
        if season == current_season and discipline
    )
    results: list[dict[str, Any]] = []
    for discipline in current_disciplines:
        rows = grouped[(current_season, discipline)]
        scores = [
            score
            for score in (_number(row.get("score")) for row in rows)
            if score is not None
        ]
        prior_scores = [
            score
            for score in (
                _number(row.get("score"))
                for row in grouped.get((prior_season, discipline), [])
            )
            if score is not None
        ]
        median = statistics.median(scores)
        prior_median = (
            statistics.median(prior_scores) if prior_scores else None
        )
        change = (
            median - prior_median
            if prior_median is not None
            else None
        )
        if change is None:
            direction = "insufficient_data"
        elif change < -0.5:
            direction = "improving"
        elif change > 0.5:
            direction = "declining"
        else:
            direction = "flat"
        if direction == "improving":
            strength = "Median time improved versus the prior season."
            priority = "Maintain gains and reinforce repeatable execution."
        elif direction == "declining":
            strength = "Median time is slower than the prior season."
            priority = "Review fundamentals and recent match conditions."
        elif direction == "flat":
            strength = "Median time is broadly stable year over year."
            priority = "Target stage consistency and incremental gains."
        else:
            strength = "No comparable prior-season median is available."
            priority = "Establish a reliable current-season baseline."
        results.append(
            {
                "season_label": current_season,
                "discipline": discipline,
                "athletes_count": len(
                    {
                        (
                            row.get("athlete_name", ""),
                            row.get("athlete_id", ""),
                        )
                        for row in rows
                    }
                ),
                "scored_entries_count": len(scores),
                "team_best_score": _display(min(scores)),
                "team_median_score": _display(median),
                "team_average_score": _display(statistics.fmean(scores)),
                "prior_season_median_score": _display(prior_median),
                "median_change_seconds": _display(change),
                "trend_direction": direction,
                "strength_note": strength,
                "coach_priority": priority,
            }
        )
    return sorted(
        results,
        key=lambda row: (
            {
                "declining": 0,
                "flat": 1,
                "insufficient_data": 2,
                "improving": 3,
            }[row["trend_direction"]],
            row["discipline"],
        ),
    )


def _quality_notes(
    identity_affected_rows: int,
    sources: list[dict[str, str]],
    quality: list[dict[str, str]],
    season: str,
) -> list[dict[str, Any]]:
    partial = sum(row.get("data_status") == "partial" for row in sources)
    no_scores = sum(row.get("data_status") == "no_scores" for row in sources)
    broad_errors = sum(
        _integer(row.get("validation_error_count")) for row in quality
    )
    return [
        {
            "topic": "Placeholder identities",
            "finding": (
                "Blank names and athlete ID 9999 are quarantined by default."
            ),
            "affected_rows": identity_affected_rows,
            "coach_interpretation": (
                "Do not attribute these entries to an athlete until resolved."
            ),
        },
        {
            "topic": "Partial matches",
            "finding": "Partial matches remain historical context.",
            "affected_rows": partial,
            "coach_interpretation": (
                "Partial status is not automatically a data failure."
            ),
        },
        {
            "topic": "No-score matches",
            "finding": (
                "No-score matches remain participation context only."
            ),
            "affected_rows": no_scores,
            "coach_interpretation": (
                "They do not contribute to improvement or regression."
            ),
        },
        {
            "topic": "Historical validation",
            "finding": (
                f"{broad_errors} broad historical ERROR findings exist."
            ),
            "affected_rows": broad_errors,
            "coach_interpretation": (
                "Treat aggregate validation counts as data context, "
                "not coaching conclusions."
            ),
        },
        {
            "topic": "Selected season",
            "finding": f"Coach-facing roster is focused on {season}.",
            "affected_rows": 0,
            "coach_interpretation": (
                "Cross-season score comparisons remain discipline-specific."
            ),
        },
    ]


def _coach_brief(
    season: str,
    roster: list[dict[str, Any]],
    improvements: list[dict[str, Any]],
    regressions: list[dict[str, Any]],
    disciplines: list[dict[str, Any]],
    identity_issues: list[dict[str, Any]],
    identity_affected_rows: int,
    sources: list[dict[str, str]],
) -> list[dict[str, Any]]:
    brief: list[dict[str, Any]] = [
        {
            "section": "Current Season Snapshot",
            "insight": f"{len(roster)} athletes appear in {season}.",
            "supporting_metric": f"{len(roster)} roster athletes",
            "confidence_level": "high",
            "coach_action": "Confirm the active roster and discipline plans.",
        }
    ]
    high_improvements = [
        row for row in improvements if row["confidence_level"] == "high"
    ]
    for row in high_improvements[:3]:
        brief.append(
            {
                "section": "Biggest Reliable Improvements",
                "insight": (
                    f"{row['athlete_name']} improved "
                    f"{float(row['improvement_seconds']):.3f}s in "
                    f"{row['discipline']}."
                ),
                "supporting_metric": (
                    f"{row['scored_matches_count']} scored matches"
                ),
                "confidence_level": "high",
                "coach_action": (
                    "Identify repeatable technique and preparation factors."
                ),
            }
        )
    for row in regressions[:3]:
        brief.append(
            {
                "section": "Watchlist",
                "insight": (
                    f"{row['athlete_name']}'s latest {row['discipline']} "
                    f"score is {float(row['regression_seconds']):.3f}s "
                    "above the season median."
                ),
                "supporting_metric": (
                    f"{row['matches_count']} matches; "
                    f"{row['confidence_level']} confidence"
                ),
                "confidence_level": row["confidence_level"],
                "coach_action": row["coach_note"],
            }
        )
    for row in disciplines:
        section = (
            "Discipline Strengths"
            if row["trend_direction"] == "improving"
            else "Discipline Priorities"
        )
        brief.append(
            {
                "section": section,
                "insight": (
                    f"{row['discipline']} trend: "
                    f"{row['trend_direction']}."
                ),
                "supporting_metric": (
                    f"median change "
                    f"{row['median_change_seconds'] or 'n/a'} seconds"
                ),
                "confidence_level": (
                    "medium"
                    if row["trend_direction"] != "insufficient_data"
                    else "low"
                ),
                "coach_action": row["coach_priority"],
            }
        )
    brief.append(
        {
            "section": "Data Quality Notes",
            "insight": (
                f"{identity_affected_rows} placeholder-affected "
                "participation rows "
                "are excluded by default."
            ),
            "supporting_metric": (
                f"{len(identity_issues)} identity issue groups"
            ),
            "confidence_level": "high",
            "coach_action": "Resolve identities before athlete attribution.",
        }
    )
    nationals = next(
        (row for row in sources if row.get("match_id") == "671"),
        None,
    )
    if nationals:
        brief.append(
            {
                "section": "Nationals Context",
                "insight": (
                    "Match 671 is retained as participation context and "
                    "excluded from performance insights while no scores exist."
                ),
                "supporting_metric": (
                    f"data_status={nationals.get('data_status', '')}"
                ),
                "confidence_level": "high",
                "coach_action": (
                    "Refresh the historical layer after Nationals scores post."
                ),
            }
        )
    return brief


def _confidence(
    scored_matches: int,
    improvement: float,
    percent: float | None,
) -> tuple[str, str, str]:
    large = improvement >= 15 or (percent is not None and percent >= 30)
    moderate = improvement >= 8 or (percent is not None and percent >= 20)
    if scored_matches == 2 or large:
        concern = (
            "Large first-to-latest swing may reflect an outlier."
            if large
            else ""
        )
        return (
            "low",
            "Two scored matches or a large possible outlier.",
            concern,
        )
    if scored_matches == 3 or moderate:
        concern = (
            "Moderate outlier concern; review match context."
            if moderate
            else ""
        )
        return (
            "medium",
            "Three scored matches or moderate outlier concern.",
            concern,
        )
    return (
        "high",
        "Four or more scored matches with no obvious outlier signal.",
        "",
    )


def _build_workbook(
    history_dir: Path,
    rows_by_file: dict[str, list[dict[str, Any]]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise HistoryInsightsError(
            "openpyxl is required for --workbook."
        ) from exc
    sheets = (
        ("Coach Brief", "wilco_coach_brief.csv", BRIEF_COLUMNS),
        (
            "Current Season Roster",
            "wilco_current_season_roster.csv",
            ROSTER_COLUMNS,
        ),
        (
            "Improvement Leaderboard",
            "wilco_improvement_leaderboard.csv",
            IMPROVEMENT_COLUMNS,
        ),
        (
            "Regression Watchlist",
            "wilco_regression_watchlist.csv",
            REGRESSION_COLUMNS,
        ),
        (
            "Discipline Insights",
            "wilco_discipline_insights.csv",
            DISCIPLINE_COLUMNS,
        ),
        (
            "Identity Issues",
            "wilco_identity_issues.csv",
            IDENTITY_COLUMNS,
        ),
        (
            "Data Quality Notes",
            "wilco_insight_data_quality_notes.csv",
            QUALITY_NOTE_COLUMNS,
        ),
    )
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="E2F0D9")
    low_fill = PatternFill("solid", fgColor="FCE4D6")
    for sheet_number, (name, filename, columns) in enumerate(sheets, 1):
        ws = wb.create_sheet(name)
        ws.append([column.replace("_", " ").title() for column in columns])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
        for row in rows_by_file[filename]:
            ws.append([_workbook_value(row.get(column, "")) for column in columns])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        if rows_by_file[filename]:
            table = Table(
                displayName=f"HistoryInsightsTable{sheet_number}",
                ref=ws.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        for column_number, column in enumerate(columns, 1):
            letter = get_column_letter(column_number)
            values = (
                ws.cell(row=row_number, column=column_number).value
                for row_number in range(2, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = _column_width(
                column,
                values,
            )
            if column.endswith("percent"):
                for cell in ws[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0"
            if column.endswith("score") or column.endswith("seconds"):
                for cell in ws[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
            if column in {
                "insight",
                "coach_action",
                "confidence_note",
                "insight_note",
                "possible_reason",
                "coach_note",
                "strength_note",
                "coach_priority",
                "issue_message",
                "recommendation",
                "finding",
                "coach_interpretation",
                "current_status_note",
            }:
                for cell in ws[letter][1:]:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
        if "confidence_level" in columns and ws.max_row > 1:
            confidence_letter = get_column_letter(
                columns.index("confidence_level") + 1
            )
            target = f"A2:{get_column_letter(len(columns))}{ws.max_row}"
            ws.conditional_formatting.add(
                target,
                FormulaRule(
                    formula=[f'${confidence_letter}2="high"'],
                    fill=high_fill,
                ),
            )
            ws.conditional_formatting.add(
                target,
                FormulaRule(
                    formula=[f'${confidence_letter}2="low"'],
                    fill=low_fill,
                ),
            )
    path = history_dir / "wilco_history_insights.xlsx"
    try:
        wb.save(path)
    except OSError as exc:
        raise HistoryInsightsError(
            f"Could not save insights workbook {path}: {exc}"
        ) from exc
    return path


def _is_placeholder(row: dict[str, str]) -> bool:
    return (
        not row.get("athlete_name", "").strip()
        or row.get("athlete_id", "").strip() == PLACEHOLDER_ATHLETE_ID
    )


def _latest_season(seasons: Iterable[str]) -> str:
    values = list(seasons)
    if not values:
        raise HistoryInsightsError(
            "No shooting seasons were found in Phase 6A outputs."
        )
    return max(values, key=_season_start)


def _season_start(value: str) -> int:
    try:
        first = int(value[:2])
    except (TypeError, ValueError):
        return -1
    return 2000 + first if first < 70 else 1900 + first


def _read_required(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise HistoryInsightsError(
            f"Required Phase 6A input is missing: {path}"
        )
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except OSError as exc:
        raise HistoryInsightsError(f"Could not read {path}: {exc}") from exc


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise HistoryInsightsError(f"Could not write {path}: {exc}") from exc


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _display(value: float | None) -> float | str:
    return round(value, 3) if value is not None else ""


def _boolean(value: bool) -> str:
    return "true" if value else "false"


def _workbook_value(value: Any) -> Any:
    if value in ("true", "false"):
        return "Yes" if value == "true" else "No"
    if isinstance(value, str):
        number = _number(value)
        if number is not None and value.strip():
            return number
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    descriptive = {
        "insight",
        "coach_action",
        "confidence_note",
        "insight_note",
        "possible_reason",
        "coach_note",
        "strength_note",
        "coach_priority",
        "issue_message",
        "recommendation",
        "finding",
        "coach_interpretation",
        "current_status_note",
    }
    if column in descriptive:
        return 44
    max_length = max(
        [len(column.replace("_", " ").title())]
        + [len(str(value)) for value in values if value is not None]
    )
    return min(max(max_length + 2, 11), 28)
