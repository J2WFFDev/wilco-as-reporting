"""Wilco records and personal-record reports from local history tables."""

from __future__ import annotations

import csv
from difflib import SequenceMatcher
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from wilco_as_reporting.athlete_aliases import (
    AthleteAliasError,
    apply_athlete_aliases,
    load_athlete_aliases,
)

SUMMARY_COLUMNS = ("metric", "value", "notes")
ALL_TIME_COLUMNS = (
    "record_scope",
    "record_title",
    "display_group",
    "display_order",
    "discipline",
    "athlete_name",
    "athlete_id",
    "score",
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "division",
    "class",
    "gender",
    "tie_count",
    "confidence_level",
    "notes",
)
SEASON_COLUMNS = (
    "season_label",
    "discipline",
    "athlete_name",
    "athlete_id",
    "score",
    "match_id",
    "match_name",
    "match_date",
    "division",
    "class",
    "gender",
    "tie_count",
    "confidence_level",
    "notes",
)
PERSONAL_COLUMNS = (
    "athlete_name",
    "canonical_athlete_name",
    "athlete_id",
    "original_athlete_name_count",
    "identity_resolution_note",
    "discipline",
    "personal_record_score",
    "personal_record_match_id",
    "personal_record_match_name",
    "personal_record_match_date",
    "personal_record_season_label",
    "first_known_score",
    "latest_score",
    "latest_match_id",
    "latest_match_name",
    "latest_match_date",
    "seconds_from_pr",
    "percent_from_pr",
    "matches_count",
    "scored_matches_count",
    "confidence_level",
    "record_confidence",
    "display_eligible",
    "display_note",
    "notes",
)
PR_HISTORY_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "score",
    "was_new_pr",
    "previous_pr_score",
    "improvement_over_previous_pr",
    "notes",
)
NEW_PR_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "athlete_name",
    "athlete_id",
    "discipline",
    "pr_event_type",
    "new_pr_score",
    "previous_pr_score",
    "improvement_seconds",
    "improvement_percent",
    "notes",
)
HIGHLIGHT_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "athlete_name",
    "discipline",
    "pr_event_type",
    "new_pr_score",
    "previous_pr_score",
    "improvement_seconds",
    "improvement_percent",
    "confidence_level",
    "display_eligible",
    "display_note",
)
MATCH_BEST_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "discipline",
    "athlete_name",
    "athlete_id",
    "score",
    "division",
    "class",
    "gender",
    "match_best_scope",
    "tie_count",
    "notes",
)
STAGE_COLUMNS = (
    "benchmark_scope",
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "discipline",
    "stage_name",
    "athlete_name",
    "athlete_id",
    "stage_score",
    "division",
    "class",
    "confidence_level",
    "notes",
)
QUALITY_COLUMNS = (
    "issue_type",
    "record_area",
    "affected_rows",
    "severity",
    "notes",
)
IDENTITY_CANDIDATE_COLUMNS = (
    "candidate_type",
    "name_a",
    "athlete_id_a",
    "name_b",
    "athlete_id_b",
    "similarity_note",
    "recommended_action",
    "notes",
)

OUTPUTS = (
    ("records_summary.csv", SUMMARY_COLUMNS),
    ("wilco_all_time_records.csv", ALL_TIME_COLUMNS),
    ("wilco_team_season_records.csv", SEASON_COLUMNS),
    ("personal_records.csv", PERSONAL_COLUMNS),
    ("personal_record_history.csv", PR_HISTORY_COLUMNS),
    ("new_personal_records_by_match.csv", NEW_PR_COLUMNS),
    ("recent_pr_highlights.csv", HIGHLIGHT_COLUMNS),
    ("match_bests.csv", MATCH_BEST_COLUMNS),
    ("stage_benchmarks.csv", STAGE_COLUMNS),
    ("records_identity_candidates.csv", IDENTITY_CANDIDATE_COLUMNS),
    ("records_data_quality.csv", QUALITY_COLUMNS),
)

PLACEHOLDER_ID = "9999"
TOLERANCE = 0.0005


class RecordsBuildError(RuntimeError):
    """Raised when local history cannot produce records outputs."""


@dataclass(frozen=True)
class RecordsBuildResult:
    records_dir: Path
    workbook_path: Path | None
    row_counts: dict[str, int]
    placeholder_rows_excluded: int
    no_score_rows_excluded: int
    identity_candidates: int
    initial_pr_count: int
    improved_pr_count: int
    display_ineligible_count: int


def build_records(
    *,
    output_root: Path | str,
    team_key: str,
    records_dir: Path | str | None = None,
    history_dir: Path | str | None = None,
    match_id: int | None = None,
    match_ids: tuple[int, ...] = (),
    match_ids_file: Path | str | None = None,
    season: str | None = None,
    include_placeholders: bool = False,
    workbook: bool = True,
) -> RecordsBuildResult:
    """Build local Wilco records without API calls."""
    if team_key.casefold() != "wilco":
        raise RecordsBuildError("Records currently support team_key 'wilco'.")
    root = Path(output_root)
    history = Path(history_dir) if history_dir else root / "history"
    target = Path(records_dir) if records_dir else root / "records"
    participation_raw = _read_required(
        history / "wilco_match_participation.csv"
    )
    try:
        aliases = load_athlete_aliases()
    except AthleteAliasError as exc:
        raise RecordsBuildError(str(exc)) from exc
    participation = apply_athlete_aliases(participation_raw, aliases)
    source_matches = _read_required(
        history / "history_source_matches.csv"
    )
    stage_history = apply_athlete_aliases(
        _read_required(history / "wilco_stage_benchmark_history.csv"),
        aliases,
    )
    selected_ids = _selected_match_ids(
        match_id,
        match_ids,
        Path(match_ids_file) if match_ids_file else None,
    )
    scoped_participation = [
        row
        for row in participation
        if _in_scope(row, selected_ids, season)
    ]
    scoped_sources = [
        row
        for row in source_matches
        if _in_scope(row, selected_ids, season)
    ]
    placeholder_count = sum(
        _is_placeholder(row) for row in scoped_participation
    )
    no_score_count = sum(
        row.get("data_status") == "no_scores"
        or _number(row.get("score")) is None
        for row in scoped_participation
    )
    scores = [
        row
        for row in scoped_participation
        if (
            (include_placeholders or not _is_placeholder(row))
            and row.get("data_status") != "no_scores"
            and _number(row.get("score")) is not None
        )
    ]
    identity_candidates = _identity_candidates(
        scoped_participation,
        aliases,
    )
    all_time = _all_time_records(scores)
    season_records = _season_records(scores)
    personal, pr_history, new_prs = _personal_records(scores)
    highlights = _recent_pr_highlights(new_prs)
    for row in new_prs:
        for internal_column in (
            "_scored_entries",
            "_data_status",
            "_score_median",
        ):
            row.pop(internal_column, None)
    match_bests = _match_bests(scores)
    stage_benchmarks, missing_stage_rows = _stage_benchmarks(
        stage_history,
        scores,
        selected_ids,
        season,
        include_placeholders,
    )
    quality = _quality_rows(
        placeholder_count if not include_placeholders else 0,
        no_score_count,
        scoped_sources,
        personal,
        all_time,
        season_records,
        match_bests,
        missing_stage_rows,
        identity_candidates,
        highlights,
    )
    summary = _summary_rows(
        scoped_sources,
        scores,
        all_time,
        season_records,
        personal,
        new_prs,
        match_bests,
        stage_benchmarks,
        placeholder_count if not include_placeholders else 0,
        no_score_count,
    )
    rows_by_file = {
        "records_summary.csv": summary,
        "wilco_all_time_records.csv": all_time,
        "wilco_team_season_records.csv": season_records,
        "personal_records.csv": personal,
        "personal_record_history.csv": pr_history,
        "new_personal_records_by_match.csv": new_prs,
        "recent_pr_highlights.csv": highlights,
        "match_bests.csv": match_bests,
        "stage_benchmarks.csv": stage_benchmarks,
        "records_identity_candidates.csv": identity_candidates,
        "records_data_quality.csv": quality,
    }
    for filename, columns in OUTPUTS:
        _write_csv(target / filename, columns, rows_by_file[filename])
    workbook_path = (
        _build_workbook(target, rows_by_file)
        if workbook
        else None
    )
    return RecordsBuildResult(
        records_dir=target,
        workbook_path=workbook_path,
        row_counts={
            filename: len(rows_by_file[filename])
            for filename, _ in OUTPUTS
        },
        placeholder_rows_excluded=(
            placeholder_count if not include_placeholders else 0
        ),
        no_score_rows_excluded=no_score_count,
        identity_candidates=len(identity_candidates),
        initial_pr_count=sum(
            row["pr_event_type"] == "initial_pr" for row in new_prs
        ),
        improved_pr_count=sum(
            row["pr_event_type"] == "improved_pr" for row in new_prs
        ),
        display_ineligible_count=sum(
            row["display_eligible"] == "false" for row in highlights
        ),
    )


def _all_time_records(
    scores: list[dict[str, str]],
) -> list[dict[str, Any]]:
    specifications = (
        (
            "wilco_all_time_discipline",
            lambda row: (row.get("discipline", ""),),
        ),
        (
            "wilco_all_time_discipline_class",
            lambda row: (
                row.get("discipline", ""),
                row.get("class", ""),
            ),
        ),
        (
            "wilco_all_time_discipline_division",
            lambda row: (
                row.get("discipline", ""),
                row.get("division", ""),
            ),
        ),
    )
    results: list[dict[str, Any]] = []
    display = {
        "wilco_all_time_discipline": (
            "Wilco All-Time Discipline Record",
            "Discipline Records",
            1,
        ),
        "wilco_all_time_discipline_class": (
            "Wilco All-Time Class Record",
            "Class Records",
            2,
        ),
        "wilco_all_time_discipline_division": (
            "Wilco All-Time Division Record",
            "Division Records",
            3,
        ),
    }
    for scope, key_function in specifications:
        grouped = _group_nonblank(scores, key_function)
        for rows in grouped.values():
            for row, tie_count in _minimum_rows(rows, "score"):
                results.append(
                    {
                        "record_scope": scope,
                        "record_title": display[scope][0],
                        "display_group": display[scope][1],
                        "display_order": display[scope][2],
                        "discipline": row.get("discipline", ""),
                        "athlete_name": row.get("athlete_name", ""),
                        "athlete_id": row.get("athlete_id", ""),
                        "score": _display(_number(row.get("score"))),
                        "match_id": row.get("match_id", ""),
                        "match_name": row.get("match_name", ""),
                        "match_date": row.get("match_date", ""),
                        "season_label": row.get("season_label", ""),
                        "division": row.get("division", ""),
                        "class": row.get("class", ""),
                        "gender": row.get("gender", ""),
                        "tie_count": tie_count,
                        "confidence_level": _record_confidence(rows),
                        "notes": (
                            "Internal Wilco record candidate; not an "
                            "official SASP record."
                        ),
                    }
                )
    return _sort_records(results, "record_scope")


def _season_records(
    scores: list[dict[str, str]],
) -> list[dict[str, Any]]:
    grouped = _group_nonblank(
        scores,
        lambda row: (
            row.get("season_label", ""),
            row.get("discipline", ""),
        ),
    )
    results: list[dict[str, Any]] = []
    for rows in grouped.values():
        for row, tie_count in _minimum_rows(rows, "score"):
            results.append(
                {
                    "season_label": row.get("season_label", ""),
                    "discipline": row.get("discipline", ""),
                    "athlete_name": row.get("athlete_name", ""),
                    "athlete_id": row.get("athlete_id", ""),
                    "score": _display(_number(row.get("score"))),
                    "match_id": row.get("match_id", ""),
                    "match_name": row.get("match_name", ""),
                    "match_date": row.get("match_date", ""),
                    "division": row.get("division", ""),
                    "class": row.get("class", ""),
                    "gender": row.get("gender", ""),
                    "tie_count": tie_count,
                    "confidence_level": _record_confidence(rows),
                    "notes": (
                        "Best known Wilco season performance; not an "
                        "official SASP record."
                    ),
                }
            )
    return sorted(
        results,
        key=lambda row: (
            row["season_label"],
            row["discipline"],
            float(row["score"]),
            row["athlete_name"],
        ),
    )


def _personal_records(
    scores: list[dict[str, str]],
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    grouped = _group_nonblank(
        scores,
        lambda row: (
            row.get("athlete_name", ""),
            row.get("athlete_id", ""),
            row.get("discipline", ""),
        ),
    )
    personal: list[dict[str, Any]] = []
    history: list[dict[str, Any]] = []
    new_prs: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        ordered = sorted(rows, key=_chronology_key)
        match_count = len({row.get("match_id", "") for row in ordered})
        score_values = [
            value
            for value in (_number(row.get("score")) for row in ordered)
            if value is not None
        ]
        original_names = {
            row.get("original_athlete_name", "")
            for row in ordered
            if row.get("original_athlete_name")
        }
        resolution_notes = {
            row.get("identity_resolution_note", "")
            for row in ordered
            if row.get("identity_resolution_note")
            and row.get("identity_resolution_note")
            != "No alias mapping applied."
        }
        current_pr: float | None = None
        for row in ordered:
            score = _number(row.get("score"))
            if score is None:
                continue
            previous = current_pr
            is_new = previous is None or score < previous - TOLERANCE
            improvement = (
                previous - score
                if is_new and previous is not None
                else None
            )
            history.append(
                {
                    "athlete_name": key[0],
                    "athlete_id": key[1],
                    "discipline": key[2],
                    "match_id": row.get("match_id", ""),
                    "match_name": row.get("match_name", ""),
                    "match_date": row.get("match_date", ""),
                    "season_label": row.get("season_label", ""),
                    "score": _display(score),
                    "was_new_pr": _boolean(is_new),
                    "previous_pr_score": _display(previous),
                    "improvement_over_previous_pr": _display(improvement),
                    "notes": (
                        "Initial known PR."
                        if previous is None
                        else (
                            "New internal personal record."
                            if is_new
                            else "Did not improve the prior PR."
                        )
                    ),
                }
            )
            if is_new:
                event_type = (
                    "initial_pr" if previous is None else "improved_pr"
                )
                percent = (
                    improvement / previous * 100
                    if improvement is not None and previous
                    else None
                )
                new_prs.append(
                    {
                        "match_id": row.get("match_id", ""),
                        "match_name": row.get("match_name", ""),
                        "match_date": row.get("match_date", ""),
                        "season_label": row.get("season_label", ""),
                        "athlete_name": key[0],
                        "athlete_id": key[1],
                        "discipline": key[2],
                        "pr_event_type": event_type,
                        "new_pr_score": _display(score),
                        "previous_pr_score": _display(previous),
                        "improvement_seconds": _display(improvement),
                        "improvement_percent": _display(percent),
                        "notes": (
                            "Initial known personal record; no prior PR "
                            "was beaten."
                            if event_type == "initial_pr"
                            else "Improved internal personal record."
                        ),
                        "_scored_entries": match_count,
                        "_data_status": row.get("data_status", ""),
                        "_score_median": _median(score_values),
                    }
                )
                current_pr = score
        best_score = min(_number(row["score"]) for row in ordered)
        best_rows = [
            row
            for row in ordered
            if abs((_number(row["score"]) or 0) - best_score) <= TOLERANCE
        ]
        best = best_rows[0]
        latest = ordered[-1]
        latest_score = _number(latest.get("score"))
        distance = (
            latest_score - best_score
            if latest_score is not None
            else None
        )
        count = match_count
        note = (
            "Current PR."
            if distance is not None and abs(distance) <= TOLERANCE
            else f"Latest score is {distance:.3f}s from the PR."
        )
        if count == 1:
            note += " Limited history: one scored match."
        confidence, display_eligible, display_note = _personal_confidence(
            count,
            best_score,
            _median(score_values),
            any(row.get("data_status") == "partial" for row in ordered),
        )
        personal.append(
            {
                "athlete_name": key[0],
                "canonical_athlete_name": key[0],
                "athlete_id": key[1],
                "original_athlete_name_count": len(original_names),
                "identity_resolution_note": (
                    " ".join(sorted(resolution_notes))
                    if resolution_notes
                    else "No alias mapping applied."
                ),
                "discipline": key[2],
                "personal_record_score": _display(best_score),
                "personal_record_match_id": best.get("match_id", ""),
                "personal_record_match_name": best.get("match_name", ""),
                "personal_record_match_date": best.get("match_date", ""),
                "personal_record_season_label": best.get(
                    "season_label",
                    "",
                ),
                "first_known_score": ordered[0].get("score", ""),
                "latest_score": _display(latest_score),
                "latest_match_id": latest.get("match_id", ""),
                "latest_match_name": latest.get("match_name", ""),
                "latest_match_date": latest.get("match_date", ""),
                "seconds_from_pr": _display(distance),
                "percent_from_pr": _display(
                    distance / best_score * 100
                    if distance is not None and best_score
                    else None
                ),
                "matches_count": count,
                "scored_matches_count": count,
                "confidence_level": confidence,
                "record_confidence": confidence,
                "display_eligible": _boolean(display_eligible),
                "display_note": display_note,
                "notes": note,
            }
        )
    personal.sort(
        key=lambda row: (
            row["athlete_name"],
            row["discipline"],
        )
    )
    history.sort(
        key=lambda row: (
            row["athlete_name"],
            row["discipline"],
            row["match_date"],
            _integer(row["match_id"]),
        )
    )
    new_prs.sort(
        key=lambda row: (
            row["match_date"],
            _integer(row["match_id"]),
            row["athlete_name"],
            row["discipline"],
        )
    )
    return personal, history, new_prs


def _recent_pr_highlights(
    new_prs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    highlights: list[dict[str, Any]] = []
    for row in new_prs:
        event_type = row["pr_event_type"]
        score = _number(row.get("new_pr_score")) or 0
        median = _number(row.get("_score_median"))
        count = _integer(row.get("_scored_entries"))
        partial = row.get("_data_status") == "partial"
        outlier = score > 150 or (
            median is not None and median > 0 and score > median * 2
        )
        if event_type == "initial_pr":
            confidence = "low"
            eligible = False
            display_note = (
                "Initial baseline PR; do not describe as beating a prior PR."
            )
        elif partial or outlier:
            confidence = "low"
            eligible = False
            display_note = (
                "Valid PR event, but partial/noisy or outlier context "
                "requires review."
            )
        elif count >= 3:
            confidence = "high"
            eligible = True
            display_note = "Celebration-safe improved PR candidate."
        elif count == 2:
            confidence = "medium"
            eligible = True
            display_note = "Improved PR with two scored entries."
        else:
            confidence = "low"
            eligible = False
            display_note = "Limited history."
        highlights.append(
            {
                "match_id": row["match_id"],
                "match_name": row["match_name"],
                "match_date": row["match_date"],
                "season_label": row["season_label"],
                "athlete_name": row["athlete_name"],
                "discipline": row["discipline"],
                "pr_event_type": event_type,
                "new_pr_score": row["new_pr_score"],
                "previous_pr_score": row["previous_pr_score"],
                "improvement_seconds": row["improvement_seconds"],
                "improvement_percent": row["improvement_percent"],
                "confidence_level": confidence,
                "display_eligible": _boolean(eligible),
                "display_note": display_note,
            }
        )
    return sorted(
        highlights,
        key=lambda row: (
            row["match_date"],
            _integer(row["match_id"]),
            row["athlete_name"],
            row["discipline"],
        ),
        reverse=True,
    )


def _match_bests(
    scores: list[dict[str, str]],
) -> list[dict[str, Any]]:
    specifications = (
        (
            "match_best_wilco_discipline",
            lambda row: (
                row.get("match_id", ""),
                row.get("discipline", ""),
            ),
        ),
        (
            "match_best_class",
            lambda row: (
                row.get("match_id", ""),
                row.get("discipline", ""),
                row.get("class", ""),
            ),
        ),
        (
            "match_best_division",
            lambda row: (
                row.get("match_id", ""),
                row.get("discipline", ""),
                row.get("division", ""),
            ),
        ),
    )
    results: list[dict[str, Any]] = []
    for scope, key_function in specifications:
        for rows in _group_nonblank(scores, key_function).values():
            for row, tie_count in _minimum_rows(rows, "score"):
                results.append(
                    {
                        "match_id": row.get("match_id", ""),
                        "match_name": row.get("match_name", ""),
                        "match_date": row.get("match_date", ""),
                        "season_label": row.get("season_label", ""),
                        "discipline": row.get("discipline", ""),
                        "athlete_name": row.get("athlete_name", ""),
                        "athlete_id": row.get("athlete_id", ""),
                        "score": row.get("score", ""),
                        "division": row.get("division", ""),
                        "class": row.get("class", ""),
                        "gender": row.get("gender", ""),
                        "match_best_scope": scope,
                        "tie_count": tie_count,
                        "notes": (
                            "Best among Wilco entries in this match scope."
                        ),
                    }
                )
    return sorted(
        results,
        key=lambda row: (
            row["match_date"],
            _integer(row["match_id"]),
            row["discipline"],
            row["match_best_scope"],
            float(row["score"]),
        ),
    )


def _stage_benchmarks(
    stages: list[dict[str, str]],
    scores: list[dict[str, str]],
    selected_ids: set[int] | None,
    season: str | None,
    include_placeholders: bool,
) -> tuple[list[dict[str, Any]], int]:
    identity = {
        (
            row.get("match_id", ""),
            row.get("athlete_name", ""),
            row.get("discipline", ""),
        ): row
        for row in scores
    }
    valid: list[dict[str, str]] = []
    missing = 0
    for stage in stages:
        if not _in_scope(stage, selected_ids, season):
            continue
        context = identity.get(
            (
                stage.get("match_id", ""),
                stage.get("athlete_name", ""),
                stage.get("discipline", ""),
            ),
            {},
        )
        score = _number(stage.get("stage_score"))
        if not stage.get("stage_name", "").strip() or score is None:
            missing += 1
            continue
        enriched = dict(stage)
        enriched.update(
            {
                "athlete_id": context.get("athlete_id", ""),
                "division": context.get("division", ""),
                "class": context.get("class", ""),
                "data_status": context.get("data_status", ""),
            }
        )
        if (
            enriched.get("data_status") == "no_scores"
            or (
                not include_placeholders
                and _is_placeholder(enriched)
            )
        ):
            continue
        valid.append(enriched)
    specifications = (
        (
            "all_time_stage_benchmark",
            lambda row: (
                row.get("discipline", ""),
                row.get("stage_name", ""),
            ),
        ),
        (
            "wilco_stage_benchmark",
            lambda row: (
                row.get("season_label", ""),
                row.get("discipline", ""),
                row.get("stage_name", ""),
            ),
        ),
        (
            "match_stage_benchmark",
            lambda row: (
                row.get("match_id", ""),
                row.get("discipline", ""),
                row.get("stage_name", ""),
            ),
        ),
    )
    results: list[dict[str, Any]] = []
    for scope, key_function in specifications:
        for rows in _group_nonblank(valid, key_function).values():
            for row, _ in _minimum_rows(rows, "stage_score"):
                results.append(
                    {
                        "benchmark_scope": scope,
                        "match_id": row.get("match_id", ""),
                        "match_name": row.get("match_name", ""),
                        "match_date": row.get("match_date", ""),
                        "season_label": row.get("season_label", ""),
                        "discipline": row.get("discipline", ""),
                        "stage_name": row.get("stage_name", ""),
                        "athlete_name": row.get("athlete_name", ""),
                        "athlete_id": row.get("athlete_id", ""),
                        "stage_score": row.get("stage_score", ""),
                        "division": row.get("division", ""),
                        "class": row.get("class", ""),
                        "confidence_level": _record_confidence(rows),
                        "notes": (
                            "Coach benchmark only; not a formal stage record."
                        ),
                    }
                )
    return (
        sorted(
            results,
            key=lambda row: (
                row["benchmark_scope"],
                row["discipline"],
                row["stage_name"],
                row["match_date"],
            ),
        ),
        missing,
    )


def _identity_candidates(
    rows: list[dict[str, Any]],
    aliases: tuple[Any, ...],
) -> list[dict[str, Any]]:
    profiles: dict[tuple[str, str], dict[str, set[str]]] = {}
    for row in rows:
        name = row.get("original_athlete_name", "").strip()
        athlete_id = row.get("athlete_id", "").strip()
        if not name or athlete_id == PLACEHOLDER_ID:
            continue
        profile = profiles.setdefault(
            (name, athlete_id),
            {"disciplines": set(), "seasons": set()},
        )
        if row.get("discipline"):
            profile["disciplines"].add(row["discipline"])
        if row.get("season_label"):
            profile["seasons"].add(row["season_label"])
    resolved_pairs = {
        (
            alias.canonical_name.casefold(),
            alias.alias_name.casefold(),
        )
        for alias in aliases
    }
    candidates: list[dict[str, Any]] = []
    identities = sorted(profiles)
    for index, identity_a in enumerate(identities):
        for identity_b in identities[index + 1 :]:
            name_a, id_a = identity_a
            name_b, id_b = identity_b
            similarity = SequenceMatcher(
                None,
                name_a.casefold(),
                name_b.casefold(),
            ).ratio()
            discipline_overlap = (
                profiles[identity_a]["disciplines"]
                & profiles[identity_b]["disciplines"]
            )
            season_overlap = (
                profiles[identity_a]["seasons"]
                & profiles[identity_b]["seasons"]
            )
            alias_pair = (
                (name_a.casefold(), name_b.casefold()) in resolved_pairs
                or (name_b.casefold(), name_a.casefold()) in resolved_pairs
            )
            same_id = bool(id_a and id_b and id_a == id_b)
            likely = same_id or (
                similarity >= 0.88
                and bool(discipline_overlap)
                and bool(season_overlap)
            )
            if not likely:
                continue
            candidates.append(
                {
                    "candidate_type": (
                        "resolved_alias"
                        if alias_pair
                        else (
                            "same_athlete_id"
                            if same_id
                            else "similar_name_overlap"
                        )
                    ),
                    "name_a": name_a,
                    "athlete_id_a": id_a,
                    "name_b": name_b,
                    "athlete_id_b": id_b,
                    "similarity_note": (
                        f"Name similarity {similarity:.1%}; "
                        f"{len(discipline_overlap)} overlapping discipline(s); "
                        f"{len(season_overlap)} overlapping season(s)."
                    ),
                    "recommended_action": (
                        "Alias mapping applied; periodically confirm identity."
                        if alias_pair
                        else "Review and add an alias only if identity matches."
                    ),
                    "notes": (
                        "Matching athlete ID supports the candidate."
                        if same_id
                        else "Candidate is based on name and activity overlap."
                    ),
                }
            )
    return candidates


def _quality_rows(
    placeholder_count: int,
    no_score_count: int,
    sources: list[dict[str, str]],
    personal: list[dict[str, Any]],
    all_time: list[dict[str, Any]],
    season_records: list[dict[str, Any]],
    match_bests: list[dict[str, Any]],
    missing_stage_rows: int,
    identity_candidates: list[dict[str, Any]],
    highlights: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    tied = sum(
        _integer(row.get("tie_count")) > 1
        for row in all_time + season_records + match_bests
    )
    return [
        {
            "issue_type": "identity_candidates",
            "record_area": "athlete identity",
            "affected_rows": len(identity_candidates),
            "severity": "REVIEW",
            "notes": "Review resolved and unresolved spelling variants.",
        },
        {
            "issue_type": "display_ineligible_pr_events",
            "record_area": "recent PR highlights",
            "affected_rows": sum(
                row["display_eligible"] == "false" for row in highlights
            ),
            "severity": "INFO",
            "notes": (
                "Valid PR events retained but withheld from celebration view."
            ),
        },
        {
            "issue_type": "placeholder_athlete_rows_excluded",
            "record_area": "all individual records",
            "affected_rows": placeholder_count,
            "severity": "WARNING",
            "notes": "Blank names and athlete ID 9999 are excluded.",
        },
        {
            "issue_type": "no_score_rows_excluded",
            "record_area": "all score records",
            "affected_rows": no_score_count,
            "severity": "INFO",
            "notes": "No-score and blank-score entries cannot set records.",
        },
        {
            "issue_type": "partial_matches_included",
            "record_area": "record candidates",
            "affected_rows": sum(
                row.get("data_status") == "partial"
                and row.get("has_wilco_entries") == "true"
                for row in sources
            ),
            "severity": "REVIEW",
            "notes": "Partial status is retained and reflected in confidence.",
        },
        {
            "issue_type": "missing_stage_data",
            "record_area": "stage benchmarks",
            "affected_rows": missing_stage_rows,
            "severity": "INFO",
            "notes": "Blank stage names or scores are not benchmarked.",
        },
        {
            "issue_type": "limited_personal_record_history",
            "record_area": "personal records",
            "affected_rows": sum(
                row.get("confidence_level") == "low" for row in personal
            ),
            "severity": "REVIEW",
            "notes": "One scored match creates a low-confidence initial PR.",
        },
        {
            "issue_type": "tie_handling",
            "record_area": "record and match best tables",
            "affected_rows": tied,
            "severity": "INFO",
            "notes": "All tied minimum-score rows are retained.",
        },
    ]


def _summary_rows(
    sources: list[dict[str, str]],
    scores: list[dict[str, str]],
    all_time: list[dict[str, Any]],
    season_records: list[dict[str, Any]],
    personal: list[dict[str, Any]],
    new_prs: list[dict[str, Any]],
    match_bests: list[dict[str, Any]],
    stages: list[dict[str, Any]],
    placeholder_count: int,
    no_score_count: int,
) -> list[dict[str, Any]]:
    metrics = (
        ("source_matches_checked", len(sources), ""),
        (
            "wilco_matches_checked",
            sum(row.get("has_wilco_entries") == "true" for row in sources),
            "",
        ),
        (
            "athletes_checked",
            len(
                {
                    (row.get("athlete_name"), row.get("athlete_id"))
                    for row in scores
                }
            ),
            "",
        ),
        (
            "disciplines_checked",
            len({row.get("discipline") for row in scores}),
            "",
        ),
        ("all_time_records_generated", len(all_time), ""),
        ("team_season_records_generated", len(season_records), ""),
        ("personal_records_generated", len(personal), ""),
        ("new_pr_rows_generated", len(new_prs), ""),
        ("match_best_rows_generated", len(match_bests), ""),
        ("stage_benchmark_rows_generated", len(stages), ""),
        ("placeholder_rows_excluded", placeholder_count, ""),
        ("no_score_rows_excluded", no_score_count, ""),
        (
            "generated_at",
            datetime.now().astimezone().isoformat(timespec="seconds"),
            "Local generation time.",
        ),
    )
    return [
        {"metric": metric, "value": value, "notes": notes}
        for metric, value, notes in metrics
    ]


def _build_workbook(
    records_dir: Path,
    rows_by_file: dict[str, list[dict[str, Any]]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RecordsBuildError("openpyxl is required for --workbook.") from exc
    sheets = (
        ("Summary", "records_summary.csv", SUMMARY_COLUMNS),
        ("Wilco All-Time Records", "wilco_all_time_records.csv", ALL_TIME_COLUMNS),
        ("Team Season Records", "wilco_team_season_records.csv", SEASON_COLUMNS),
        ("Personal Records", "personal_records.csv", PERSONAL_COLUMNS),
        ("PR History", "personal_record_history.csv", PR_HISTORY_COLUMNS),
        ("New PRs by Match", "new_personal_records_by_match.csv", NEW_PR_COLUMNS),
        ("Recent PR Highlights", "recent_pr_highlights.csv", HIGHLIGHT_COLUMNS),
        ("Match Bests", "match_bests.csv", MATCH_BEST_COLUMNS),
        ("Stage Benchmarks", "stage_benchmarks.csv", STAGE_COLUMNS),
        (
            "Identity Candidates",
            "records_identity_candidates.csv",
            IDENTITY_CANDIDATE_COLUMNS,
        ),
        ("Data Quality", "records_data_quality.csv", QUALITY_COLUMNS),
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for number, (name, filename, columns) in enumerate(sheets, 1):
        sheet = workbook.create_sheet(name)
        sheet.append([column.replace("_", " ").title() for column in columns])
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True)
        for row in rows_by_file[filename]:
            sheet.append(
                [_workbook_value(row.get(column, "")) for column in columns]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        if rows_by_file[filename]:
            table = Table(
                displayName=f"RecordsTable{number}",
                ref=sheet.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_number, column in enumerate(columns, 1):
            letter = get_column_letter(column_number)
            values = (
                sheet.cell(row=row, column=column_number).value
                for row in range(2, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = _column_width(
                column,
                values,
            )
            if (
                "score" in column
                or "seconds" in column
                or column == "improvement_over_previous_pr"
            ):
                for cell in sheet[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
            if "percent" in column:
                for cell in sheet[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0"
            if column == "notes":
                for cell in sheet[letter][1:]:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
    path = records_dir / "wilco_records_report.xlsx"
    try:
        workbook.save(path)
    except OSError as exc:
        raise RecordsBuildError(f"Could not save workbook {path}: {exc}") from exc
    return path


def _selected_match_ids(
    match_id: int | None,
    match_ids: tuple[int, ...],
    match_ids_file: Path | None,
) -> set[int] | None:
    selected = set(match_ids)
    if match_id is not None:
        selected.add(match_id)
    if match_ids_file:
        try:
            selected.update(
                int(line.strip())
                for line in match_ids_file.read_text(
                    encoding="utf-8-sig"
                ).splitlines()
                if line.strip()
            )
        except (OSError, ValueError) as exc:
            raise RecordsBuildError(
                f"Could not read match ID file {match_ids_file}: {exc}"
            ) from exc
    return selected or None


def _in_scope(
    row: dict[str, str],
    selected_ids: set[int] | None,
    season: str | None,
) -> bool:
    if selected_ids is not None:
        if _integer(row.get("match_id")) not in selected_ids:
            return False
    return season is None or row.get("season_label") == season


def _group_nonblank(
    rows: list[dict[str, str]],
    key_function: Callable[[dict[str, str]], tuple[str, ...]],
) -> dict[tuple[str, ...], list[dict[str, str]]]:
    grouped: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = key_function(row)
        if all(str(value).strip() for value in key):
            grouped[key].append(row)
    return grouped


def _minimum_rows(
    rows: list[dict[str, str]],
    score_column: str,
) -> list[tuple[dict[str, str], int]]:
    minimum = min(
        value
        for value in (_number(row.get(score_column)) for row in rows)
        if value is not None
    )
    tied = [
        row
        for row in rows
        if abs((_number(row.get(score_column)) or 0) - minimum) <= TOLERANCE
    ]
    return [(row, len(tied)) for row in tied]


def _record_confidence(rows: list[dict[str, str]]) -> str:
    match_count = len({row.get("match_id", "") for row in rows})
    if any(row.get("data_status") == "partial" for row in rows):
        return "medium" if match_count >= 2 else "low"
    return _count_confidence(match_count)


def _count_confidence(count: int) -> str:
    if count >= 4:
        return "high"
    if count >= 2:
        return "medium"
    return "low"


def _personal_confidence(
    count: int,
    score: float,
    median: float | None,
    partial: bool,
) -> tuple[str, bool, str]:
    outlier = score > 150 or (
        median is not None and median > 0 and score > median * 2
    )
    if count == 1:
        return (
            "low",
            False,
            "Valid initial PR, but one score is not celebration-safe.",
        )
    if partial or outlier:
        return (
            "low",
            False,
            "Record is retained but requires partial/outlier review.",
        )
    if count >= 3:
        return (
            "high",
            True,
            "Established across at least three scored matches.",
        )
    return (
        "medium",
        True,
        "Reasonable record candidate based on two scored matches.",
    )


def _median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    middle = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[middle]
    return (ordered[middle - 1] + ordered[middle]) / 2


def _chronology_key(row: dict[str, str]) -> tuple[str, int]:
    return (
        row.get("match_date") or "9999-12-31",
        _integer(row.get("match_id")),
    )


def _sort_records(
    rows: list[dict[str, Any]],
    scope_column: str,
) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row[scope_column],
            row["discipline"],
            float(row["score"]),
            row["athlete_name"],
        ),
    )


def _is_placeholder(row: dict[str, str]) -> bool:
    return (
        not row.get("athlete_name", "").strip()
        or row.get("athlete_id", "").strip() == PLACEHOLDER_ID
    )


def _read_required(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise RecordsBuildError(f"Required history input is missing: {path}")
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except OSError as exc:
        raise RecordsBuildError(f"Could not read {path}: {exc}") from exc


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise RecordsBuildError(f"Could not write {path}: {exc}") from exc


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _display(value: float | None) -> float | str:
    return round(value, 3) if value is not None else ""


def _boolean(value: bool) -> str:
    return "true" if value else "false"


def _workbook_value(value: Any) -> Any:
    if value in ("true", "false"):
        return "Yes" if value == "true" else "No"
    if isinstance(value, str):
        number = _number(value)
        if number is not None and value.strip():
            return number
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    if column in {
        "notes",
        "identity_resolution_note",
        "display_note",
        "similarity_note",
        "recommended_action",
    }:
        return 60
    maximum = max(
        [len(column.replace("_", " ").title())]
        + [len(str(value)) for value in values if value is not None]
    )
    return min(max(maximum + 2, 11), 30)
