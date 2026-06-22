"""Coach-facing Nationals readiness brief from local Wilco outputs."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from wilco_as_reporting.athlete_aliases import (
    AthleteAliasError,
    apply_athlete_aliases,
    load_athlete_aliases,
)
from wilco_as_reporting.team_profiles import TeamProfile

SUMMARY_COLUMNS = ("metric", "value", "notes")
ROSTER_COLUMNS = (
    "match_id",
    "athlete_name",
    "athlete_id",
    "disciplines_entered",
    "discipline_count",
    "division",
    "class",
    "gender",
    "has_historical_data",
    "current_season_entries",
    "notes",
)
READINESS_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "division",
    "class",
    "gender",
    "entered_for_nationals",
    "has_history",
    "scored_matches_count",
    "current_season_scored_count",
    "personal_record_score",
    "latest_score",
    "seconds_from_pr",
    "latest_match_name",
    "latest_match_date",
    "recent_pr_flag",
    "recent_pr_match",
    "improvement_trend",
    "watchlist_flag",
    "readiness_level",
    "coach_focus",
    "coach_note",
)
ATHLETE_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "disciplines_entered",
    "strongest_discipline",
    "best_pr_context",
    "highest_priority_focus",
    "readiness_summary",
    "coach_note",
)
DISCIPLINE_COLUMNS = (
    "discipline",
    "athletes_entered",
    "athletes_with_history",
    "athletes_near_pr",
    "athletes_with_recent_pr",
    "watchlist_count",
    "team_record_score",
    "current_season_best",
    "current_season_median",
    "discipline_trend_direction",
    "coach_priority",
    "notes",
)
PR_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "personal_record_score",
    "latest_score",
    "seconds_from_pr",
    "recent_pr_flag",
    "pr_opportunity_type",
    "coach_note",
)
WATCH_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "watch_reason",
    "supporting_metric",
    "confidence_level",
    "coach_action",
    "private_note",
)
ACTION_COLUMNS = (
    "priority",
    "area",
    "action",
    "athletes_or_disciplines",
    "rationale",
    "timing",
)
QUALITY_COLUMNS = ("issue_type", "affected_rows", "severity", "notes")

OUTPUTS = (
    ("nationals_readiness_summary.csv", SUMMARY_COLUMNS),
    ("nationals_roster.csv", ROSTER_COLUMNS),
    ("athlete_discipline_readiness.csv", READINESS_COLUMNS),
    ("athlete_readiness_summary.csv", ATHLETE_COLUMNS),
    ("discipline_readiness.csv", DISCIPLINE_COLUMNS),
    ("pr_opportunities.csv", PR_COLUMNS),
    ("watchlist.csv", WATCH_COLUMNS),
    ("coach_action_plan.csv", ACTION_COLUMNS),
    ("data_quality_notes.csv", QUALITY_COLUMNS),
)

PLACEHOLDER_ID = "9999"


class NationalsReadinessError(RuntimeError):
    """Raised when local readiness inputs cannot be processed."""


@dataclass(frozen=True)
class NationalsReadinessResult:
    output_dir: Path
    workbook_path: Path | None
    row_counts: dict[str, int]
    athletes_entered: int
    athlete_discipline_entries: int
    readiness_counts: dict[str, int]
    opportunity_counts: dict[str, int]
    watchlist_count: int


def build_nationals_readiness(
    *,
    output_root: Path | str,
    profile: TeamProfile,
    match_id: int = 671,
    history_dir: Path | str | None = None,
    records_dir: Path | str | None = None,
    season: str | None = None,
    workbook: bool = True,
    include_limited_history: bool = True,
) -> NationalsReadinessResult:
    """Build a local, private coach readiness brief."""
    root = Path(output_root)
    history = Path(history_dir) if history_dir else root / "history"
    records = Path(records_dir) if records_dir else root / "records"
    match_scores = _read_required(
        root / str(match_id) / "tables" / "match_scores.csv"
    )
    current_roster = _read_required(
        history / "wilco_current_season_roster.csv"
    )
    improvements = _read_required(
        history / "wilco_improvement_leaderboard.csv"
    )
    regressions = _read_required(
        history / "wilco_regression_watchlist.csv"
    )
    discipline_insights = _read_required(
        history / "wilco_discipline_insights.csv"
    )
    personal_records = _read_required(records / "personal_records.csv")
    pr_history = _read_required(records / "personal_record_history.csv")
    highlights = _read_required(records / "recent_pr_highlights.csv")
    all_time = _read_required(records / "wilco_all_time_records.csv")
    season_records = _read_required(
        records / "wilco_team_season_records.csv"
    )
    quality_history = _read_optional(
        history / "wilco_data_quality_summary.csv"
    )
    try:
        aliases = load_athlete_aliases()
    except AthleteAliasError as exc:
        raise NationalsReadinessError(str(exc)) from exc
    team_entries_raw = [
        row
        for row in match_scores
        if profile.matches_name(row.get("team_name"))
    ]
    placeholder_count = sum(_is_placeholder(row) for row in team_entries_raw)
    team_entries = apply_athlete_aliases(
        [
            row
            for row in team_entries_raw
            if not _is_placeholder(row)
        ],
        aliases,
    )
    if not team_entries:
        raise NationalsReadinessError(
            f"No {profile.team_name} entries were found for match {match_id}."
        )
    match_name = team_entries[0].get("match_name") or f"Match {match_id}"
    selected_season = season or _latest_season(
        row.get("season_label", "") for row in current_roster
    )
    roster_index = _identity_index(current_roster)
    personal_index = _discipline_index(personal_records)
    improvement_index = _selected_season_index(
        improvements,
        selected_season,
    )
    current_counts = Counter(
        (row.get("athlete_id", ""), row.get("discipline", ""))
        for row in pr_history
        if row.get("season_label") == selected_season
    )
    regression_index = _selected_season_index(
        regressions,
        selected_season,
    )
    highlight_index = _recent_highlight_index(
        highlights,
        personal_records,
        selected_season,
    )
    readiness = _readiness_rows(
        team_entries,
        personal_index,
        improvement_index,
        current_counts,
        regression_index,
        highlight_index,
        include_limited_history,
    )
    roster = _roster_rows(
        match_id,
        team_entries,
        roster_index,
        personal_index,
    )
    athlete_summary = _athlete_summary(readiness)
    discipline_rows = _discipline_rows(
        readiness,
        discipline_insights,
        all_time,
        season_records,
        selected_season,
    )
    opportunities = _pr_opportunities(readiness)
    watchlist = _watchlist_rows(readiness)
    action_plan = _action_plan(
        readiness,
        discipline_rows,
        watchlist,
    )
    alias_applied = sum(
        row.get("original_athlete_name", "") != row.get("athlete_name", "")
        for row in team_entries
    )
    quality = _quality_rows(
        team_entries,
        readiness,
        quality_history,
        alias_applied,
        placeholder_count,
    )
    summary = _summary_rows(
        match_id,
        match_name,
        roster,
        readiness,
        opportunities,
        watchlist,
    )
    rows_by_file = {
        "nationals_readiness_summary.csv": summary,
        "nationals_roster.csv": roster,
        "athlete_discipline_readiness.csv": readiness,
        "athlete_readiness_summary.csv": athlete_summary,
        "discipline_readiness.csv": discipline_rows,
        "pr_opportunities.csv": opportunities,
        "watchlist.csv": watchlist,
        "coach_action_plan.csv": action_plan,
        "data_quality_notes.csv": quality,
    }
    target = root / "nationals_readiness"
    for filename, columns in OUTPUTS:
        _write_csv(target / filename, columns, rows_by_file[filename])
    workbook_path = (
        _build_workbook(target, match_id, rows_by_file)
        if workbook
        else None
    )
    readiness_counts = Counter(
        row["readiness_level"] for row in readiness
    )
    opportunity_counts = Counter(
        row["pr_opportunity_type"] for row in opportunities
    )
    return NationalsReadinessResult(
        output_dir=target,
        workbook_path=workbook_path,
        row_counts={
            filename: len(rows_by_file[filename])
            for filename, _ in OUTPUTS
        },
        athletes_entered=len(roster),
        athlete_discipline_entries=len(readiness),
        readiness_counts=dict(readiness_counts),
        opportunity_counts=dict(opportunity_counts),
        watchlist_count=len(watchlist),
    )


def _readiness_rows(
    entries: list[dict[str, Any]],
    personal: dict[tuple[str, str], dict[str, str]],
    improvements: dict[tuple[str, str], dict[str, str]],
    current_counts: Counter[tuple[str, str]],
    regressions: dict[tuple[str, str], dict[str, str]],
    highlights: dict[tuple[str, str], dict[str, str]],
    include_limited_history: bool,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for entry in entries:
        key = (entry.get("athlete_id", ""), entry.get("discipline", ""))
        pr = personal.get(key, {})
        improvement = improvements.get(key, {})
        regression = regressions.get(key, {})
        highlight = highlights.get(key, {})
        has_history = bool(pr)
        scored_count = _integer(pr.get("scored_matches_count"))
        current_count = current_counts.get(key, 0)
        pr_score = _number(pr.get("personal_record_score"))
        latest = _number(pr.get("latest_score"))
        distance = (
            latest - pr_score
            if latest is not None and pr_score is not None
            else None
        )
        recent_pr = (
            highlight.get("pr_event_type") == "improved_pr"
            and highlight.get("display_eligible") == "true"
        )
        watch = bool(regression)
        latest_date = pr.get("latest_match_date", "")
        if not has_history:
            level = "new_or_unscored"
            focus = "confirm discipline plan"
            note = "No scored history is available; observe first."
        elif scored_count <= 1:
            level = "limited_history"
            focus = "limited data, observe first"
            note = "Use the first strings to establish confidence and rhythm."
        elif watch:
            level = "watch_trend"
            focus = "stabilize recent trend"
            note = "Use a recent-form check and reinforce repeatable process."
        elif recent_pr:
            level = "strong_pr_momentum"
            focus = "recent PR momentum"
            note = "Maintain and reinforce the routine behind the recent PR."
        elif distance is not None and distance <= 2:
            level = "near_pr"
            focus = "confidence under pressure"
            note = "Latest result is near the internal PR."
        elif latest_date and latest_date < "2025-09-01":
            level = "no_recent_data"
            focus = "needs recent-form check"
            note = "Historical context exists, but current form needs checking."
        else:
            level = "stable"
            focus = "maintain and reinforce routine"
            note = "Use steady execution and familiar pre-stage routines."
        if level == "limited_history" and not include_limited_history:
            continue
        results.append(
            {
                "athlete_name": entry.get("athlete_name", ""),
                "athlete_id": entry.get("athlete_id", ""),
                "discipline": entry.get("discipline", ""),
                "division": _division(entry.get("class", "")),
                "class": entry.get("class", ""),
                "gender": entry.get("gender", ""),
                "entered_for_nationals": "true",
                "has_history": _boolean(has_history),
                "scored_matches_count": scored_count,
                "current_season_scored_count": current_count,
                "personal_record_score": _display(pr_score),
                "latest_score": _display(latest),
                "seconds_from_pr": _display(distance),
                "latest_match_name": pr.get("latest_match_name", ""),
                "latest_match_date": latest_date,
                "recent_pr_flag": _boolean(recent_pr),
                "recent_pr_match": (
                    highlight.get("match_name", "") if recent_pr else ""
                ),
                "improvement_trend": (
                    "improving"
                    if improvement
                    and _number(improvement.get("improvement_seconds", 0)) > 0
                    else "insufficient_data"
                ),
                "watchlist_flag": _boolean(watch),
                "readiness_level": level,
                "coach_focus": focus,
                "coach_note": note,
            }
        )
    return sorted(
        results,
        key=lambda row: (
            row["athlete_name"],
            row["discipline"],
        ),
    )


def _roster_rows(
    match_id: int,
    entries: list[dict[str, Any]],
    roster_index: dict[str, dict[str, str]],
    personal_index: dict[tuple[str, str], dict[str, str]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in entries:
        grouped[(row.get("athlete_name", ""), row.get("athlete_id", ""))].append(
            row
        )
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        disciplines = sorted({row.get("discipline", "") for row in rows})
        roster = roster_index.get(key[1], {})
        has_history = any(
            (key[1], discipline) in personal_index
            for discipline in disciplines
        )
        results.append(
            {
                "match_id": match_id,
                "athlete_name": key[0],
                "athlete_id": key[1],
                "disciplines_entered": ", ".join(disciplines),
                "discipline_count": len(disciplines),
                "division": _division(rows[0].get("class", "")),
                "class": rows[0].get("class", ""),
                "gender": rows[0].get("gender", ""),
                "has_historical_data": _boolean(has_history),
                "current_season_entries": roster.get(
                    "scored_entries_count",
                    0,
                ),
                "notes": (
                    "Nationals entry context only; Match 671 has no scores."
                ),
            }
        )
    return sorted(results, key=lambda row: row["athlete_name"])


def _athlete_summary(
    readiness: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in readiness:
        grouped[(row["athlete_name"], row["athlete_id"])].append(row)
    priority = {
        "watch_trend": 0,
        "new_or_unscored": 1,
        "limited_history": 2,
        "no_recent_data": 3,
        "near_pr": 4,
        "stable": 5,
        "strong_pr_momentum": 6,
    }
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        strongest = min(
            (row for row in rows if _number(row.get("seconds_from_pr")) is not None),
            key=lambda row: _number(row["seconds_from_pr"]) or 0,
            default=rows[0],
        )
        focus_row = min(rows, key=lambda row: priority[row["readiness_level"]])
        counts = Counter(row["readiness_level"] for row in rows)
        results.append(
            {
                "athlete_name": key[0],
                "athlete_id": key[1],
                "disciplines_entered": ", ".join(
                    sorted(row["discipline"] for row in rows)
                ),
                "strongest_discipline": strongest["discipline"],
                "best_pr_context": (
                    f"{strongest['discipline']}: "
                    f"{strongest['seconds_from_pr']}s from PR"
                    if strongest.get("seconds_from_pr") != ""
                    else "Limited scored history"
                ),
                "highest_priority_focus": focus_row["coach_focus"],
                "readiness_summary": "; ".join(
                    f"{count} {level}"
                    for level, count in sorted(counts.items())
                ),
                "coach_note": (
                    "Coordinate discipline-specific support with constructive "
                    "language and private context."
                ),
            }
        )
    return sorted(results, key=lambda row: row["athlete_name"])


def _discipline_rows(
    readiness: list[dict[str, Any]],
    insights: list[dict[str, str]],
    all_time: list[dict[str, str]],
    season_records: list[dict[str, str]],
    season: str,
) -> list[dict[str, Any]]:
    insight_index = {
        row.get("discipline", ""): row
        for row in insights
        if row.get("season_label") == season
    }
    team_record_index = {
        row.get("discipline", ""): row
        for row in all_time
        if row.get("record_scope") == "wilco_all_time_discipline"
    }
    season_record_index = {
        row.get("discipline", ""): row
        for row in season_records
        if row.get("season_label") == season
    }
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in readiness:
        grouped[row["discipline"]].append(row)
    results: list[dict[str, Any]] = []
    for discipline, rows in grouped.items():
        insight = insight_index.get(discipline, {})
        results.append(
            {
                "discipline": discipline,
                "athletes_entered": len(rows),
                "athletes_with_history": sum(
                    row["has_history"] == "true" for row in rows
                ),
                "athletes_near_pr": sum(
                    row["readiness_level"] in {
                        "near_pr",
                        "strong_pr_momentum",
                    }
                    for row in rows
                ),
                "athletes_with_recent_pr": sum(
                    row["recent_pr_flag"] == "true" for row in rows
                ),
                "watchlist_count": sum(
                    row["watchlist_flag"] == "true" for row in rows
                ),
                "team_record_score": team_record_index.get(
                    discipline,
                    {},
                ).get("score", ""),
                "current_season_best": season_record_index.get(
                    discipline,
                    {},
                ).get("score", ""),
                "current_season_median": insight.get(
                    "team_median_score",
                    "",
                ),
                "discipline_trend_direction": insight.get(
                    "trend_direction",
                    "insufficient_data",
                ),
                "coach_priority": insight.get(
                    "coach_priority",
                    "Confirm discipline plan and reinforce routine.",
                ),
                "notes": "Match 671 scores are not included.",
            }
        )
    return sorted(results, key=lambda row: row["discipline"])


def _pr_opportunities(
    readiness: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in readiness:
        distance = _number(row.get("seconds_from_pr"))
        if not row.get("personal_record_score"):
            opportunity = "limited_history"
            note = "Use Nationals to establish constructive baseline context."
        elif row["recent_pr_flag"] == "true":
            opportunity = "recent_improver"
            note = "Reinforce the routine behind recent PR momentum."
        elif distance is not None and abs(distance) <= 0.001:
            opportunity = "current_pr_holder"
            note = "Latest score equals the internal PR."
        elif distance is not None and distance <= 2:
            opportunity = "within_2_seconds"
            note = "A PR is close; focus on confidence and process."
        elif distance is not None and distance <= 5:
            opportunity = "within_5_seconds"
            note = "A PR is plausible with steady execution."
        elif row["scored_matches_count"] <= 1:
            opportunity = "limited_history"
            note = "Use limited history carefully; observe first."
        else:
            opportunity = "not_near_pr"
            note = "Prioritize stable execution over chasing a number."
        results.append(
            {
                "athlete_name": row["athlete_name"],
                "athlete_id": row["athlete_id"],
                "discipline": row["discipline"],
                "personal_record_score": row["personal_record_score"],
                "latest_score": row["latest_score"],
                "seconds_from_pr": row["seconds_from_pr"],
                "recent_pr_flag": row["recent_pr_flag"],
                "pr_opportunity_type": opportunity,
                "coach_note": note,
            }
        )
    return results


def _watchlist_rows(
    readiness: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in readiness:
        reason = ""
        metric = ""
        confidence = "medium"
        action = ""
        if row["watchlist_flag"] == "true":
            reason = "watch trend"
            metric = f"{row['seconds_from_pr']}s from PR"
            confidence = "medium"
            action = "stabilize recent trend"
        elif row["readiness_level"] == "limited_history":
            reason = "limited history"
            metric = f"{row['scored_matches_count']} scored match(es)"
            confidence = "low"
            action = "limited data, observe first"
        elif row["readiness_level"] == "new_or_unscored":
            reason = "new discipline context"
            metric = "no scored history"
            confidence = "low"
            action = "confirm discipline plan"
        elif row["readiness_level"] == "no_recent_data":
            reason = "needs recent-form check"
            metric = row["latest_match_date"]
            confidence = "medium"
            action = "validate gear/process"
        elif (
            _number(row.get("seconds_from_pr")) is not None
            and (_number(row["seconds_from_pr"]) or 0) > 10
        ):
            reason = "confidence focus"
            metric = f"{row['seconds_from_pr']}s from PR"
            confidence = "medium"
            action = "confidence under pressure"
        if not reason:
            continue
        results.append(
            {
                "athlete_name": row["athlete_name"],
                "athlete_id": row["athlete_id"],
                "discipline": row["discipline"],
                "watch_reason": reason,
                "supporting_metric": metric,
                "confidence_level": confidence,
                "coach_action": action,
                "private_note": (
                    "Internal coaching context; use constructive language."
                ),
            }
        )
    return results


def _action_plan(
    readiness: list[dict[str, Any]],
    disciplines: list[dict[str, Any]],
    watchlist: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    recent = sorted(
        {
            row["athlete_name"]
            for row in readiness
            if row["recent_pr_flag"] == "true"
        }
    )
    limited = sorted(
        {
            row["athlete_name"]
            for row in readiness
            if row["readiness_level"] in {
                "limited_history",
                "new_or_unscored",
            }
        }
    )
    priorities = sorted(
        {
            row["discipline"]
            for row in disciplines
            if row["watchlist_count"]
            or row["discipline_trend_direction"] == "declining"
        }
    )
    return [
        {
            "priority": 1,
            "area": "Entry and equipment confirmation",
            "action": "Confirm discipline plan, equipment, and travel checklist.",
            "athletes_or_disciplines": "All Nationals entries",
            "rationale": "Match 671 currently has entries but no scores.",
            "timing": "before_travel",
        },
        {
            "priority": 2,
            "area": "PR momentum",
            "action": "Reinforce routines associated with recent PRs.",
            "athletes_or_disciplines": ", ".join(recent) or "None flagged",
            "rationale": "Protect confidence without chasing record numbers.",
            "timing": "practice_day",
        },
        {
            "priority": 3,
            "area": "Stability focus",
            "action": "Use short recent-form and equipment process checks.",
            "athletes_or_disciplines": ", ".join(priorities) or "As needed",
            "rationale": f"{len(watchlist)} private watchlist item(s).",
            "timing": "practice_day",
        },
        {
            "priority": 4,
            "area": "Limited history",
            "action": "Observe first and give simple process cues.",
            "athletes_or_disciplines": ", ".join(limited) or "None flagged",
            "rationale": "Avoid overinterpreting sparse historical data.",
            "timing": "match_day",
        },
        {
            "priority": 5,
            "area": "Post-match learning",
            "action": "Refresh history, records, and readiness after scores post.",
            "athletes_or_disciplines": "All Nationals entries",
            "rationale": "Match 671 performance is excluded until scores exist.",
            "timing": "post_match",
        },
    ]


def _quality_rows(
    entries: list[dict[str, Any]],
    readiness: list[dict[str, Any]],
    quality_history: list[dict[str, str]],
    alias_applied: int,
    placeholder_count: int,
) -> list[dict[str, Any]]:
    return [
        {
            "issue_type": "missing_671_score_data",
            "affected_rows": len(entries),
            "severity": "INFO",
            "notes": (
                "Match 671 entries are context only; no scores are used."
            ),
        },
        {
            "issue_type": "limited_athlete_history",
            "affected_rows": sum(
                row["readiness_level"] in {
                    "limited_history",
                    "new_or_unscored",
                }
                for row in readiness
            ),
            "severity": "REVIEW",
            "notes": "Use limited data constructively and observe first.",
        },
        {
            "issue_type": "alias_applied_names",
            "affected_rows": alias_applied,
            "severity": "INFO",
            "notes": "Active athlete aliases were applied to Nationals entries.",
        },
        {
            "issue_type": "placeholder_exclusions",
            "affected_rows": placeholder_count,
            "severity": "WARNING",
            "notes": "Blank names and athlete ID 9999 are excluded.",
        },
        {
            "issue_type": "partial_historical_matches",
            "affected_rows": sum(
                row.get("data_status") == "partial"
                and row.get("has_wilco_entries") == "true"
                for row in quality_history
            ),
            "severity": "INFO",
            "notes": "Partial history is context, not a coaching conclusion.",
        },
        {
            "issue_type": "no_recent_data",
            "affected_rows": sum(
                row["readiness_level"] == "no_recent_data"
                for row in readiness
            ),
            "severity": "REVIEW",
            "notes": "Recent form should be checked before competition.",
        },
    ]


def _summary_rows(
    match_id: int,
    match_name: str,
    roster: list[dict[str, Any]],
    readiness: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
    watchlist: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    metrics = (
        ("match_id", match_id, ""),
        ("match_name", match_name, ""),
        ("athletes_entered", len(roster), ""),
        (
            "disciplines_entered",
            len({row["discipline"] for row in readiness}),
            "",
        ),
        ("athlete_discipline_entries", len(readiness), ""),
        (
            "athletes_with_history",
            len(
                {
                    row["athlete_id"]
                    for row in readiness
                    if row["has_history"] == "true"
                }
            ),
            "",
        ),
        (
            "athletes_with_limited_history",
            len(
                {
                    row["athlete_id"]
                    for row in readiness
                    if row["readiness_level"] in {
                        "limited_history",
                        "new_or_unscored",
                    }
                }
            ),
            "",
        ),
        (
            "athletes_with_recent_pr",
            len(
                {
                    row["athlete_id"]
                    for row in readiness
                    if row["recent_pr_flag"] == "true"
                }
            ),
            "",
        ),
        (
            "athletes_near_pr",
            len(
                {
                    row["athlete_id"]
                    for row in readiness
                    if row["readiness_level"] in {
                        "near_pr",
                        "strong_pr_momentum",
                    }
                }
            ),
            "",
        ),
        ("watchlist_items", len(watchlist), ""),
        (
            "generated_at",
            datetime.now().astimezone().isoformat(timespec="seconds"),
            "Local generation time.",
        ),
    )
    return [
        {"metric": metric, "value": value, "notes": notes}
        for metric, value, notes in metrics
    ]


def _build_workbook(
    output_dir: Path,
    match_id: int,
    rows_by_file: dict[str, list[dict[str, Any]]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise NationalsReadinessError(
            "openpyxl is required for --workbook."
        ) from exc
    sheets = (
        ("Summary", "nationals_readiness_summary.csv", SUMMARY_COLUMNS),
        ("Nationals Roster", "nationals_roster.csv", ROSTER_COLUMNS),
        (
            "Athlete Discipline Readiness",
            "athlete_discipline_readiness.csv",
            READINESS_COLUMNS,
        ),
        ("Athlete Summary", "athlete_readiness_summary.csv", ATHLETE_COLUMNS),
        (
            "Discipline Readiness",
            "discipline_readiness.csv",
            DISCIPLINE_COLUMNS,
        ),
        ("PR Opportunities", "pr_opportunities.csv", PR_COLUMNS),
        ("Watchlist", "watchlist.csv", WATCH_COLUMNS),
        ("Coach Action Plan", "coach_action_plan.csv", ACTION_COLUMNS),
        ("Data Quality", "data_quality_notes.csv", QUALITY_COLUMNS),
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for number, (name, filename, columns) in enumerate(sheets, 1):
        sheet = workbook.create_sheet(name)
        sheet.append([column.replace("_", " ").title() for column in columns])
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True)
        for row in rows_by_file[filename]:
            sheet.append(
                [_workbook_value(row.get(column, "")) for column in columns]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        if rows_by_file[filename]:
            table = Table(
                displayName=f"ReadinessTable{number}",
                ref=sheet.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_number, column in enumerate(columns, 1):
            letter = get_column_letter(column_number)
            values = (
                sheet.cell(row=row, column=column_number).value
                for row in range(2, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = _column_width(
                column,
                values,
            )
            if "score" in column or "seconds" in column:
                for cell in sheet[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
            if column in {
                "notes",
                "coach_note",
                "coach_focus",
                "readiness_summary",
                "coach_priority",
                "private_note",
                "action",
                "rationale",
                "disciplines_entered",
            }:
                for cell in sheet[letter][1:]:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
    path = output_dir / f"wilco_{match_id}_nationals_readiness.xlsx"
    try:
        workbook.save(path)
    except OSError as exc:
        raise NationalsReadinessError(
            f"Could not save readiness workbook {path}: {exc}"
        ) from exc
    return path


def _identity_index(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {
        row.get("athlete_id", ""): row
        for row in rows
        if row.get("athlete_id")
    }


def _discipline_index(
    rows: list[dict[str, str]],
) -> dict[tuple[str, str], dict[str, str]]:
    return {
        (row.get("athlete_id", ""), row.get("discipline", "")): row
        for row in rows
    }


def _selected_season_index(
    rows: list[dict[str, str]],
    season: str,
) -> dict[tuple[str, str], dict[str, str]]:
    return {
        (row.get("athlete_id", ""), row.get("discipline", "")): row
        for row in rows
        if row.get("season_label") == season
    }


def _recent_highlight_index(
    rows: list[dict[str, str]],
    personal_records: list[dict[str, str]],
    season: str,
) -> dict[tuple[str, str], dict[str, str]]:
    name_to_id = {
        row.get("athlete_name", ""): row.get("athlete_id", "")
        for row in personal_records
    }
    result: dict[tuple[str, str], dict[str, str]] = {}
    for row in sorted(
        (row for row in rows if row.get("season_label") == season),
        key=lambda row: (
            row.get("match_date", ""),
            _integer(row.get("match_id")),
        ),
    ):
        athlete_name = row.get("athlete_name", "")
        athlete_id = name_to_id.get(athlete_name, "")
        if athlete_id:
            result[(athlete_id, row.get("discipline", ""))] = row
    return result


def _latest_season(values: Iterable[str]) -> str:
    seasons = [value for value in values if value and value != "Unknown Season"]
    if not seasons:
        raise NationalsReadinessError("No season was found in history outputs.")
    return max(seasons, key=lambda value: int(value[:2]))


def _read_required(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise NationalsReadinessError(f"Required local input is missing: {path}")
    return _read_csv(path)


def _read_optional(path: Path) -> list[dict[str, str]]:
    return _read_csv(path) if path.exists() else []


def _read_csv(path: Path) -> list[dict[str, str]]:
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except OSError as exc:
        raise NationalsReadinessError(f"Could not read {path}: {exc}") from exc


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise NationalsReadinessError(f"Could not write {path}: {exc}") from exc


def _is_placeholder(row: dict[str, Any]) -> bool:
    return (
        not str(row.get("athlete_name", "")).strip()
        or str(row.get("athlete_id", "")).strip() == PLACEHOLDER_ID
    )


def _division(class_value: str) -> str:
    return class_value.split("/", 1)[0].strip() if class_value else ""


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _display(value: float | None) -> float | str:
    return round(value, 3) if value is not None else ""


def _boolean(value: bool) -> str:
    return "true" if value else "false"


def _workbook_value(value: Any) -> Any:
    if value in ("true", "false"):
        return "Yes" if value == "true" else "No"
    if isinstance(value, str):
        number = _number(value)
        if number is not None and value.strip():
            return number
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    if column in {
        "notes",
        "coach_note",
        "coach_focus",
        "readiness_summary",
        "coach_priority",
        "private_note",
        "action",
        "rationale",
        "disciplines_entered",
    }:
        return 48
    maximum = max(
        [len(column.replace("_", " ").title())]
        + [len(str(value)) for value in values if value is not None]
    )
    return min(max(maximum + 2, 11), 30)
