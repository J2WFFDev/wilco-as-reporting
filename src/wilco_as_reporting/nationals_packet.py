"""Concise private coach packet built from local Nationals readiness data."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

BRIEF_COLUMNS = (
    "section",
    "message",
    "supporting_metric",
    "coach_action",
    "visibility",
)
PRIORITY_COLUMNS = (
    "priority_rank",
    "athlete_name",
    "discipline",
    "priority_type",
    "reason",
    "supporting_metric",
    "coach_action",
    "timing",
    "confidence_level",
    "private_note",
)
ATHLETE_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "disciplines_entered",
    "readiness_summary",
    "top_strength",
    "top_pr_opportunity",
    "top_coach_focus",
    "match_day_note",
    "private_coach_note",
    "public_safe_note",
)
SHOUTOUT_COLUMNS = (
    "athlete_name",
    "discipline",
    "shoutout_type",
    "supporting_metric",
    "match_name",
    "public_safe_message",
    "display_eligible",
    "notes",
)
PRACTICE_COLUMNS = (
    "focus_area",
    "athletes",
    "disciplines",
    "reason",
    "suggested_action",
    "notes",
)
WATCH_COLUMNS = (
    "athlete_name",
    "discipline",
    "watch_point",
    "positive_reframe",
    "coach_action",
    "private_note",
)
DISCIPLINE_COLUMNS = (
    "discipline",
    "athletes_entered",
    "readiness_note",
    "pr_opportunity_note",
    "watchlist_note",
    "coach_priority",
    "suggested_focus",
)
QUALITY_COLUMNS = ("issue_type", "severity", "affected_rows", "notes")

OUTPUTS = (
    ("coach_meeting_brief.csv", BRIEF_COLUMNS),
    ("top_coach_priorities.csv", PRIORITY_COLUMNS),
    ("athlete_cards.csv", ATHLETE_COLUMNS),
    ("pr_shoutout_candidates.csv", SHOUTOUT_COLUMNS),
    ("practice_day_focus.csv", PRACTICE_COLUMNS),
    ("match_day_watch_points.csv", WATCH_COLUMNS),
    ("discipline_coach_plan.csv", DISCIPLINE_COLUMNS),
    ("packet_data_quality.csv", QUALITY_COLUMNS),
)


class NationalsPacketError(RuntimeError):
    """Raised when the local readiness packet cannot be generated."""


@dataclass(frozen=True)
class NationalsPacketResult:
    output_dir: Path
    workbook_path: Path | None
    row_counts: dict[str, int]


def build_nationals_packet(
    *,
    output_root: Path | str,
    team_key: str,
    match_id: int = 671,
    readiness_dir: Path | str | None = None,
    records_dir: Path | str | None = None,
    history_dir: Path | str | None = None,
    top_priorities: int = 10,
    workbook: bool = True,
) -> NationalsPacketResult:
    """Build a private coach meeting packet without API calls."""
    if team_key.casefold() != "wilco":
        raise NationalsPacketError("Packet currently supports team_key 'wilco'.")
    if top_priorities < 1:
        raise NationalsPacketError("top_priorities must be at least 1.")
    root = Path(output_root)
    readiness = (
        Path(readiness_dir)
        if readiness_dir
        else root / "nationals_readiness"
    )
    records = Path(records_dir) if records_dir else root / "records"
    history = Path(history_dir) if history_dir else root / "history"
    summary = _read_required(readiness / "nationals_readiness_summary.csv")
    athlete_readiness = _read_required(
        readiness / "athlete_discipline_readiness.csv"
    )
    athlete_summary = _read_required(
        readiness / "athlete_readiness_summary.csv"
    )
    discipline_readiness = _read_required(
        readiness / "discipline_readiness.csv"
    )
    pr_opportunities = _read_required(readiness / "pr_opportunities.csv")
    watchlist = _read_required(readiness / "watchlist.csv")
    action_plan = _read_required(readiness / "coach_action_plan.csv")
    highlights = _read_required(records / "recent_pr_highlights.csv")
    personal_records = _read_required(records / "personal_records.csv")
    discipline_insights = _read_required(
        history / "wilco_discipline_insights.csv"
    )

    priorities = _top_priorities(
        athlete_readiness,
        discipline_readiness,
        watchlist,
        top_priorities,
    )
    shoutouts = _shoutout_candidates(
        match_id,
        athlete_readiness,
        highlights,
        pr_opportunities,
    )
    cards = _athlete_cards(
        athlete_summary,
        athlete_readiness,
        priorities,
        pr_opportunities,
    )
    practice = _practice_focus(athlete_readiness, priorities)
    watch_points = _match_day_watch_points(priorities)
    discipline_plan = _discipline_plan(
        discipline_readiness,
        discipline_insights,
    )
    brief = _meeting_brief(
        summary,
        priorities,
        shoutouts,
        discipline_plan,
        action_plan,
    )
    quality = _quality_rows(
        watchlist,
        priorities,
        shoutouts,
        highlights,
        personal_records,
    )
    rows_by_file = {
        "coach_meeting_brief.csv": brief,
        "top_coach_priorities.csv": priorities,
        "athlete_cards.csv": cards,
        "pr_shoutout_candidates.csv": shoutouts,
        "practice_day_focus.csv": practice,
        "match_day_watch_points.csv": watch_points,
        "discipline_coach_plan.csv": discipline_plan,
        "packet_data_quality.csv": quality,
    }
    target = root / "nationals_packet"
    for filename, columns in OUTPUTS:
        _write_csv(target / filename, columns, rows_by_file[filename])
    workbook_path = (
        _build_workbook(target, match_id, rows_by_file)
        if workbook
        else None
    )
    return NationalsPacketResult(
        output_dir=target,
        workbook_path=workbook_path,
        row_counts={
            filename: len(rows_by_file[filename])
            for filename, _ in OUTPUTS
        },
    )


def _top_priorities(
    readiness: list[dict[str, str]],
    disciplines: list[dict[str, str]],
    watchlist: list[dict[str, str]],
    limit: int,
) -> list[dict[str, Any]]:
    candidates: list[tuple[int, dict[str, Any]]] = []
    priority_score = {
        "new discipline context": 100,
        "limited history": 90,
        "watch trend": 80,
        "confidence focus": 70,
    }
    priority_type = {
        "new discipline context": "new_discipline",
        "limited history": "limited_history",
        "watch trend": "watch_trend",
        "confidence focus": "confidence_focus",
    }
    timing = {
        "new discipline context": "before_travel",
        "limited history": "practice_day",
        "watch trend": "practice_day",
        "confidence focus": "match_day",
    }
    for row in watchlist:
        reason = row.get("watch_reason", "")
        score = priority_score.get(reason, 50)
        candidates.append(
            (
                score,
                {
                    "athlete_name": row.get("athlete_name", ""),
                    "discipline": row.get("discipline", ""),
                    "priority_type": priority_type.get(
                        reason,
                        "watch_trend",
                    ),
                    "reason": reason,
                    "supporting_metric": row.get("supporting_metric", ""),
                    "coach_action": row.get("coach_action", ""),
                    "timing": timing.get(reason, "practice_day"),
                    "confidence_level": row.get(
                        "confidence_level",
                        "medium",
                    ),
                    "private_note": row.get("private_note", ""),
                },
            )
        )
    for row in readiness:
        if row.get("readiness_level") == "near_pr":
            candidates.append(
                (
                    60,
                    {
                        "athlete_name": row.get("athlete_name", ""),
                        "discipline": row.get("discipline", ""),
                        "priority_type": "near_pr_opportunity",
                        "reason": "near personal record",
                        "supporting_metric": (
                            f"{row.get('seconds_from_pr', '')}s from PR"
                        ),
                        "coach_action": "confidence under pressure",
                        "timing": "match_day",
                        "confidence_level": "high",
                        "private_note": (
                            "Positive opportunity; avoid number chasing."
                        ),
                    },
                )
            )
        if row.get("readiness_level") == "strong_pr_momentum":
            candidates.append(
                (
                    50,
                    {
                        "athlete_name": row.get("athlete_name", ""),
                        "discipline": row.get("discipline", ""),
                        "priority_type": "recent_pr_momentum",
                        "reason": "recent PR momentum",
                        "supporting_metric": row.get("recent_pr_match", ""),
                        "coach_action": "maintain and reinforce routine",
                        "timing": "practice_day",
                        "confidence_level": "high",
                        "private_note": "Positive reinforcement opportunity.",
                    },
                )
            )
    for row in disciplines:
        if (
            row.get("discipline_trend_direction") == "declining"
            or _integer(row.get("watchlist_count")) > 0
        ):
            candidates.append(
                (
                    65,
                    {
                        "athlete_name": "",
                        "discipline": row.get("discipline", ""),
                        "priority_type": "discipline_priority",
                        "reason": (
                            "discipline trend and watchlist concentration"
                        ),
                        "supporting_metric": (
                            f"{row.get('watchlist_count', 0)} watch item(s)"
                        ),
                        "coach_action": row.get("coach_priority", ""),
                        "timing": "practice_day",
                        "confidence_level": "medium",
                        "private_note": "Team-level coaching focus.",
                    },
                )
            )
    candidates.sort(
        key=lambda item: (
            -item[0],
            item[1]["athlete_name"],
            item[1]["discipline"],
        )
    )
    selected: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    athlete_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    for _, row in candidates:
        key = (
            row["athlete_name"],
            row["discipline"],
            row["priority_type"],
        )
        if key in seen:
            continue
        athlete_name = row["athlete_name"]
        priority_type_value = row["priority_type"]
        if athlete_name and athlete_counts[athlete_name] >= 2:
            continue
        if type_counts[priority_type_value] >= 4:
            continue
        seen.add(key)
        selected.append(row)
        if athlete_name:
            athlete_counts[athlete_name] += 1
        type_counts[priority_type_value] += 1
        if len(selected) == limit:
            break
    return [
        {"priority_rank": rank, **row}
        for rank, row in enumerate(selected, 1)
    ]


def _shoutout_candidates(
    match_id: int,
    readiness: list[dict[str, str]],
    highlights: list[dict[str, str]],
    opportunities: list[dict[str, str]],
) -> list[dict[str, Any]]:
    candidate_keys = {
        (
            row.get("athlete_name", ""),
            row.get("discipline", ""),
            row.get("match_name", ""),
        ): row
        for row in highlights
        if (
            row.get("display_eligible") == "true"
            and row.get("pr_event_type") == "improved_pr"
        )
    }
    opportunity_index = {
        (row.get("athlete_name", ""), row.get("discipline", "")): row
        for row in opportunities
    }
    results: list[dict[str, Any]] = []
    for row in readiness:
        key = (row.get("athlete_name", ""), row.get("discipline", ""))
        highlight = candidate_keys.get(
            key + (row.get("recent_pr_match", ""),)
        )
        opportunity = opportunity_index.get(key, {})
        if (
            row.get("recent_pr_flag") == "true"
            and highlight
            and _integer(highlight.get("match_id")) <= match_id
        ):
            results.append(
                {
                    "athlete_name": key[0],
                    "discipline": key[1],
                    "shoutout_type": "recent_improved_pr",
                    "supporting_metric": (
                        f"Improved by "
                        f"{highlight.get('improvement_seconds', '')}s"
                    ),
                    "match_name": highlight.get("match_name", ""),
                    "public_safe_message": (
                        f"Celebrating {key[0]}'s recent progress in "
                        f"{key[1]}—great momentum heading into Nationals!"
                    ),
                    "display_eligible": "true",
                    "notes": "Built only from display-eligible improved PR data.",
                }
            )
        elif opportunity.get("pr_opportunity_type") == "current_pr_holder":
            results.append(
                {
                    "athlete_name": key[0],
                    "discipline": key[1],
                    "shoutout_type": "current_pr_holder",
                    "supporting_metric": (
                        f"Current PR {row.get('personal_record_score', '')}s"
                    ),
                    "match_name": row.get("latest_match_name", ""),
                    "public_safe_message": (
                        f"{key[0]} enters Nationals at their current Wilco "
                        f"personal best in {key[1]}."
                    ),
                    "display_eligible": "true",
                    "notes": "Positive current-PR context.",
                }
            )
        elif opportunity.get("pr_opportunity_type") in {
            "within_2_seconds",
            "within_5_seconds",
        }:
            results.append(
                {
                    "athlete_name": key[0],
                    "discipline": key[1],
                    "shoutout_type": "near_pr_opportunity",
                    "supporting_metric": (
                        f"{row.get('seconds_from_pr', '')}s from PR"
                    ),
                    "match_name": row.get("latest_match_name", ""),
                    "public_safe_message": (
                        f"{key[0]} is building strong momentum in "
                        f"{key[1]} heading into Nationals."
                    ),
                    "display_eligible": "true",
                    "notes": "Positive near-PR context.",
                }
            )
    unique: dict[tuple[str, str], dict[str, Any]] = {}
    for row in results:
        unique[(row["athlete_name"], row["discipline"])] = row
    return sorted(
        unique.values(),
        key=lambda row: (row["athlete_name"], row["discipline"]),
    )


def _athlete_cards(
    summaries: list[dict[str, str]],
    readiness: list[dict[str, str]],
    priorities: list[dict[str, Any]],
    opportunities: list[dict[str, str]],
) -> list[dict[str, Any]]:
    readiness_by_athlete: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in readiness:
        readiness_by_athlete[row.get("athlete_id", "")].append(row)
    priority_by_athlete: dict[str, list[dict[str, Any]]] = defaultdict(list)
    name_to_id = {
        row.get("athlete_name", ""): row.get("athlete_id", "")
        for row in summaries
    }
    for row in priorities:
        athlete_id = name_to_id.get(row.get("athlete_name", ""), "")
        if athlete_id:
            priority_by_athlete[athlete_id].append(row)
    opportunity_by_athlete: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in opportunities:
        opportunity_by_athlete[row.get("athlete_id", "")].append(row)
    results: list[dict[str, Any]] = []
    for summary in summaries:
        athlete_id = summary.get("athlete_id", "")
        rows = readiness_by_athlete[athlete_id]
        strength = next(
            (
                row
                for row in rows
                if row.get("readiness_level") == "strong_pr_momentum"
            ),
            next(
                (
                    row
                    for row in rows
                    if row.get("readiness_level") == "near_pr"
                ),
                rows[0],
            ),
        )
        opportunity = min(
            opportunity_by_athlete[athlete_id],
            key=lambda row: (
                _number(row.get("seconds_from_pr"))
                if _number(row.get("seconds_from_pr")) is not None
                else 10**9
            ),
            default={},
        )
        private_priorities = priority_by_athlete[athlete_id]
        private_note = (
            "; ".join(
                f"{row['discipline']}: {row['reason']}"
                for row in private_priorities[:3]
            )
            or "No top-ten private priority assigned."
        )
        results.append(
            {
                "athlete_name": summary.get("athlete_name", ""),
                "athlete_id": athlete_id,
                "disciplines_entered": summary.get(
                    "disciplines_entered",
                    "",
                ),
                "readiness_summary": summary.get("readiness_summary", ""),
                "top_strength": (
                    f"{strength.get('discipline', '')}: "
                    f"{strength.get('readiness_level', '')}"
                ),
                "top_pr_opportunity": (
                    f"{opportunity.get('discipline', '')}: "
                    f"{opportunity.get('pr_opportunity_type', '')}"
                    if opportunity
                    else "Use process goals"
                ),
                "top_coach_focus": summary.get(
                    "highest_priority_focus",
                    "",
                ),
                "match_day_note": (
                    "Keep cues simple, familiar, and discipline-specific."
                ),
                "private_coach_note": private_note,
                "public_safe_note": (
                    f"{summary.get('athlete_name', '')} brings experience "
                    "and positive preparation into Nationals."
                ),
            }
        )
    return results


def _practice_focus(
    readiness: list[dict[str, str]],
    priorities: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    focus_map = (
        (
            "confirm new discipline plan",
            {"new_or_unscored"},
            "Confirm entry plan, equipment, and simple stage cues.",
        ),
        (
            "confidence under pressure",
            {"near_pr"},
            "Use process goals and low-pressure match simulations.",
        ),
        (
            "maintain PR routine",
            {"strong_pr_momentum"},
            "Repeat the routine associated with recent PR momentum.",
        ),
        (
            "transition discipline support",
            {"limited_history"},
            "Observe first and reinforce one or two stable cues.",
        ),
        (
            "gear/process check",
            {"watch_trend", "no_recent_data"},
            "Run a brief equipment and recent-form check.",
        ),
    )
    results: list[dict[str, Any]] = []
    for focus, levels, action in focus_map:
        rows = [row for row in readiness if row.get("readiness_level") in levels]
        if not rows:
            continue
        results.append(
            {
                "focus_area": focus,
                "athletes": ", ".join(
                    sorted({row.get("athlete_name", "") for row in rows})
                ),
                "disciplines": ", ".join(
                    sorted({row.get("discipline", "") for row in rows})
                ),
                "reason": (
                    f"{len(rows)} athlete-discipline readiness item(s)."
                ),
                "suggested_action": action,
                "notes": (
                    f"{sum(row['timing'] == 'practice_day' for row in priorities)} "
                    "top priorities are assigned to practice day."
                ),
            }
        )
    return results


def _match_day_watch_points(
    priorities: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = [
        row
        for row in priorities
        if row.get("athlete_name")
        and row.get("timing") in {"match_day", "practice_day"}
    ]
    return [
        {
            "athlete_name": row["athlete_name"],
            "discipline": row["discipline"],
            "watch_point": row["reason"],
            "positive_reframe": (
                "Coach attention can support confidence and repeatable process."
            ),
            "coach_action": row["coach_action"],
            "private_note": row["private_note"],
        }
        for row in rows
    ]


def _discipline_plan(
    readiness: list[dict[str, str]],
    insights: list[dict[str, str]],
) -> list[dict[str, Any]]:
    insight_index = {
        row.get("discipline", ""): row
        for row in insights
    }
    results: list[dict[str, Any]] = []
    for row in readiness:
        insight = insight_index.get(row.get("discipline", ""), {})
        results.append(
            {
                "discipline": row.get("discipline", ""),
                "athletes_entered": row.get("athletes_entered", ""),
                "readiness_note": (
                    f"{row.get('athletes_near_pr', 0)} near/at PR; "
                    f"{row.get('athletes_with_recent_pr', 0)} recent PR."
                ),
                "pr_opportunity_note": (
                    f"{row.get('athletes_near_pr', 0)} positive PR "
                    "opportunity athlete(s)."
                ),
                "watchlist_note": (
                    f"{row.get('watchlist_count', 0)} private watch item(s)."
                ),
                "coach_priority": row.get("coach_priority", ""),
                "suggested_focus": insight.get(
                    "strength_note",
                    "Reinforce repeatable execution.",
                ),
            }
        )
    return results


def _meeting_brief(
    summary: list[dict[str, str]],
    priorities: list[dict[str, Any]],
    shoutouts: list[dict[str, Any]],
    disciplines: list[dict[str, Any]],
    action_plan: list[dict[str, str]],
) -> list[dict[str, Any]]:
    metrics = {row.get("metric", ""): row.get("value", "") for row in summary}
    recent_pr_count = sum(
        row["shoutout_type"] == "recent_improved_pr"
        for row in shoutouts
    )
    near_pr_count = sum(
        row["shoutout_type"] == "near_pr_opportunity"
        for row in shoutouts
    )
    discipline_watch_count = sum(
        _integer(row["watchlist_note"].split()[0])
        for row in disciplines
    )
    brief = [
        {
            "section": "Team Snapshot",
            "message": (
                f"{metrics.get('athletes_entered', 0)} athletes have "
                f"{metrics.get('athlete_discipline_entries', 0)} Nationals "
                "entries."
            ),
            "supporting_metric": (
                f"{metrics.get('disciplines_entered', 0)} disciplines"
            ),
            "coach_action": "Confirm entry, travel, and equipment details.",
            "visibility": "team_general",
        },
        {
            "section": "PR Momentum",
            "message": (
                f"{recent_pr_count} "
                "display-safe recent improved PR opportunities are available."
            ),
            "supporting_metric": f"{len(shoutouts)} public-safe candidates",
            "coach_action": "Reinforce routines, not record chasing.",
            "visibility": "coach_public_safe",
        },
        {
            "section": "Near-PR Opportunities",
            "message": (
                f"{near_pr_count} "
                "positive near-PR messages are available."
            ),
            "supporting_metric": (
                f"{metrics.get('athletes_near_pr', 0)} athletes near PR"
            ),
            "coach_action": "Use process goals and confidence cues.",
            "visibility": "coach_public_safe",
        },
        {
            "section": "Discipline Priorities",
            "message": (
                f"{len(disciplines)} disciplines have concise coaching plans."
            ),
            "supporting_metric": (
                f"{discipline_watch_count} "
                "private watch items"
            ),
            "coach_action": "Assign discipline leads and practice focus.",
            "visibility": "coach_private",
        },
        {
            "section": "Top Coach Attention Items",
            "message": (
                f"The readiness watchlist is compressed to "
                f"{len(priorities)} ranked priorities."
            ),
            "supporting_metric": "Private internal list",
            "coach_action": "Review privately before practice day.",
            "visibility": "coach_private",
        },
    ]
    for section, timing, visibility in (
        ("Practice-Day Focus", "practice_day", "coach_private"),
        ("Match-Day Focus", "match_day", "coach_private"),
    ):
        actions = [
            row.get("action", "")
            for row in action_plan
            if row.get("timing") == timing
        ]
        brief.append(
            {
                "section": section,
                "message": " ".join(actions) or "Use established routines.",
                "supporting_metric": f"{len(actions)} action item(s)",
                "coach_action": "Keep cues simple and constructive.",
                "visibility": visibility,
            }
        )
    brief.append(
        {
            "section": "Data Quality Notes",
            "message": (
                "Match 671 has entries but no scores; performance context "
                "comes from prior local history only."
            ),
            "supporting_metric": "No Match 671 scores used",
            "coach_action": "Refresh the packet after Nationals scores post.",
            "visibility": "coach_private",
        }
    )
    return brief


def _quality_rows(
    watchlist: list[dict[str, str]],
    priorities: list[dict[str, Any]],
    shoutouts: list[dict[str, Any]],
    highlights: list[dict[str, str]],
    personal_records: list[dict[str, str]],
) -> list[dict[str, Any]]:
    return [
        {
            "issue_type": "watchlist_compression",
            "severity": "INFO",
            "affected_rows": len(watchlist) - len(priorities),
            "notes": (
                f"{len(watchlist)} readiness watch rows were reduced to "
                f"{len(priorities)} ranked priorities."
            ),
        },
        {
            "issue_type": "public_safe_filter",
            "severity": "INFO",
            "affected_rows": sum(
                row.get("display_eligible") != "true" for row in highlights
            ),
            "notes": "Display-ineligible PR rows are excluded from shoutouts.",
        },
        {
            "issue_type": "private_public_separation",
            "severity": "INFO",
            "affected_rows": len(priorities),
            "notes": "Private watch context is not copied into shoutout rows.",
        },
        {
            "issue_type": "limited_personal_history",
            "severity": "REVIEW",
            "affected_rows": sum(
                row.get("record_confidence") == "low"
                for row in personal_records
            ),
            "notes": "Low-confidence PRs remain excluded from celebration.",
        },
        {
            "issue_type": "shoutout_candidates",
            "severity": "INFO",
            "affected_rows": len(shoutouts),
            "notes": "All candidates use positive, public-safe context.",
        },
    ]


def _build_workbook(
    output_dir: Path,
    match_id: int,
    rows_by_file: dict[str, list[dict[str, Any]]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise NationalsPacketError(
            "openpyxl is required for --workbook."
        ) from exc
    sheets = (
        ("Coach Meeting Brief", "coach_meeting_brief.csv", BRIEF_COLUMNS),
        ("Top Coach Priorities", "top_coach_priorities.csv", PRIORITY_COLUMNS),
        ("Athlete Cards", "athlete_cards.csv", ATHLETE_COLUMNS),
        (
            "PR Shoutout Candidates",
            "pr_shoutout_candidates.csv",
            SHOUTOUT_COLUMNS,
        ),
        ("Practice-Day Focus", "practice_day_focus.csv", PRACTICE_COLUMNS),
        (
            "Match-Day Watch Points",
            "match_day_watch_points.csv",
            WATCH_COLUMNS,
        ),
        (
            "Discipline Coach Plan",
            "discipline_coach_plan.csv",
            DISCIPLINE_COLUMNS,
        ),
        ("Data Quality", "packet_data_quality.csv", QUALITY_COLUMNS),
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for number, (name, filename, columns) in enumerate(sheets, 1):
        sheet = workbook.create_sheet(name)
        sheet.append([column.replace("_", " ").title() for column in columns])
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True)
        for row in rows_by_file[filename]:
            sheet.append(
                [_workbook_value(row.get(column, "")) for column in columns]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        if rows_by_file[filename]:
            table = Table(
                displayName=f"NationalsPacketTable{number}",
                ref=sheet.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_number, column in enumerate(columns, 1):
            letter = get_column_letter(column_number)
            values = (
                sheet.cell(row=row, column=column_number).value
                for row in range(2, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = _column_width(
                column,
                values,
            )
            if column in {
                "message",
                "coach_action",
                "private_note",
                "private_coach_note",
                "public_safe_note",
                "public_safe_message",
                "notes",
                "reason",
                "suggested_action",
                "readiness_note",
                "pr_opportunity_note",
                "watchlist_note",
                "coach_priority",
                "suggested_focus",
                "disciplines_entered",
                "readiness_summary",
                "athletes",
                "disciplines",
                "watch_point",
                "positive_reframe",
            }:
                for cell in sheet[letter][1:]:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
    path = output_dir / f"wilco_{match_id}_nationals_coach_packet.xlsx"
    try:
        workbook.save(path)
    except OSError as exc:
        raise NationalsPacketError(
            f"Could not save coach packet workbook {path}: {exc}"
        ) from exc
    return path


def _read_required(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise NationalsPacketError(f"Required local input is missing: {path}")
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except OSError as exc:
        raise NationalsPacketError(f"Could not read {path}: {exc}") from exc


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise NationalsPacketError(f"Could not write {path}: {exc}") from exc


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _workbook_value(value: Any) -> Any:
    if value in ("true", "false"):
        return "Yes" if value == "true" else "No"
    if isinstance(value, str):
        number = _number(value)
        if number is not None and value.strip():
            return number
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    if column in {
        "message",
        "coach_action",
        "private_note",
        "private_coach_note",
        "public_safe_note",
        "public_safe_message",
        "notes",
        "reason",
        "suggested_action",
        "readiness_note",
        "pr_opportunity_note",
        "watchlist_note",
        "coach_priority",
        "suggested_focus",
        "disciplines_entered",
        "readiness_summary",
        "athletes",
        "disciplines",
        "watch_point",
        "positive_reframe",
    }:
        return 48
    maximum = max(
        [len(column.replace("_", " ").title())]
        + [len(str(value)) for value in values if value is not None]
    )
    return min(max(maximum + 2, 11), 30)
