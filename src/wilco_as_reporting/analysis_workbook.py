"""Build the curated Wilco analysis workbook from local report outputs."""

from __future__ import annotations

import csv
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from wilco_as_reporting.athlete_aliases import (
    AthleteAliasError,
    apply_athlete_aliases,
    load_athlete_aliases,
)
from wilco_as_reporting.team_profiles import TeamProfile

PLACEHOLDER_ID = "9999"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LABEL_FILL = PatternFill("solid", fgColor="D9EAF7")

EXCEL_MAX_ROWS = 1_048_576
VIEW_SETS = ("historical-prep", "wilco-match", "staff-match", "all")

HISTORICAL_PREP_SHEETS = (
    (
        "Wilco Historical Score History",
        "wilco_historical_score_history.csv",
        (
            "match_name", "match_date", "season_label", "athlete_name",
            "athlete_id", "division", "class", "gender", "discipline",
            "overall_place", "gender_place", "division_place",
            "class_place", "match_score",
        ),
    ),
    (
        "Athlete Perf by Discipline",
        "athlete_performance_by_discipline.csv",
        (
            "athlete_name", "athlete_id", "discipline",
            "number_of_matches", "best_score", "match_with_best_score",
            "most_recent_score", "most_recent_match", "avg_score",
            "median_score", "personal_record_score", "seconds_from_pr",
            "trend_note",
        ),
    ),
    (
        "Athlete Capability Matrix",
        "athlete_capability_matrix.csv",
        (),
    ),
    (
        "Wilco vs Field by Discipline",
        "wilco_vs_field_by_discipline.csv",
        (
            "discipline", "Wilco Entries", "Field Entries", "Wilco Avg",
            "Field Avg", "Wilco Best", "Field Best", "Top Team Avg",
            "Avg Gap To Field",
        ),
    ),
    (
        "Competitive Analysis",
        "competitive_analysis.csv",
        (
            "Athlete Name", "Athlete ID", "Team", "Gender", "Division",
            "Class", "Discipline", "Number Of Matches", "Best Score",
            "Match With Best Score", "Best Score Date", "Most Recent Score",
            "Most Recent Match", "Most Recent Date", "Avg Score",
            "Median Score", "Data Match Method", "Competitive Note",
        ),
    ),
    (
        "Records and PRs",
        "records_and_prs.csv",
        (
            "record_type", "athlete_name", "discipline", "score",
            "match_name", "match_date", "previous_score",
            "improvement_seconds", "display_eligible",
            "confidence_level", "notes",
        ),
    ),
    (
        "Data Quality Notes",
        "historical_prep_data_quality_notes.csv",
        ("area", "issue_type", "affected_rows", "severity", "notes"),
    ),
)

WILCO_MATCH_SHEETS = (
    (
        "Wilco Match Results",
        "wilco_match_results.csv",
        (
            "athlete_name", "athlete_id", "discipline", "division",
            "class", "gender", "match_score", "current_rank",
            "rank_scope", "field_size", "notes",
        ),
    ),
    (
        "Wilco Stage Review",
        "wilco_stage_review.csv",
        (
            "athlete_name", "athlete_id", "discipline", "stage_name",
            "stage_score", "fastest_string", "avg_string", "match_total",
            "current_rank", "notes",
        ),
    ),
    (
        "Match Records and PRs",
        "match_records_and_prs.csv",
        (
            "athlete_name", "discipline", "highlight_type", "score",
            "previous_record_or_pr", "improvement_seconds",
            "display_eligible", "notes",
        ),
    ),
    (
        "Coach Review Queue",
        "coach_review_queue.csv",
        (
            "severity", "athlete_name", "discipline", "stage_name",
            "finding_type", "finding_message", "recommended_review", "notes",
        ),
    ),
    (
        "Data Quality Notes",
        "wilco_match_data_quality_notes.csv",
        ("area", "issue_type", "affected_rows", "severity", "notes"),
    ),
)

STAFF_MATCH_SHEETS = (
    (
        "All Competitor Match Results",
        "all_competitor_match_results.csv",
        (
            "athlete_name", "team_name", "discipline", "division",
            "class", "gender", "match_score", "place", "rank_scope",
            "field_size", "notes",
        ),
    ),
    (
        "Published Ranking Detail",
        "published_ranking_detail.csv",
        (
            "discipline", "athlete_name", "team_name", "score",
            "rank_type", "rank_scope", "place", "field_size", "percentile",
            "margin_to_leader", "margin_to_previous_place",
            "margin_to_podium_cutoff", "athlete_class", "gender",
            "source_file",
        ),
    ),
    (
        "Validation Findings",
        "validation_findings.csv",
        (
            "severity", "finding_type", "athlete_name", "team_name",
            "discipline", "stage_name", "score", "expected_score",
            "difference", "finding_message", "recommended_review", "notes",
        ),
    ),
    (
        "Field by Discipline",
        "field_by_discipline.csv",
        (
            "discipline", "teams_count", "athletes_count", "entries_count",
            "field_avg", "field_best", "top_team_avg", "Wilco Avg",
            "Wilco Best", "Wilco Gap To Field",
        ),
    ),
    (
        "Records Match Bests",
        "records_match_bests.csv",
        (
            "highlight_type", "athlete_name", "team_name", "discipline",
            "score", "record_scope", "notes",
        ),
    ),
    (
        "Data Quality Notes",
        "staff_match_data_quality_notes.csv",
        ("area", "issue_type", "affected_rows", "severity", "notes"),
    ),
)

LONG_CSVS = {
    "athlete_discipline_stage_values.csv": (
        "athlete_name", "athlete_id", "discipline", "stage_name",
        "season_label", "match_id", "match_name", "avg_string",
        "fastest_string", "avg_stage", "fastest_stage",
        "match_score", "notes",
    ),
    "all_historical_match_scores_full.csv": (
        "match_id", "match_name", "discipline", "athlete_id",
        "athlete_name", "team_name", "class", "gender",
        "match_score_seconds", "dnf_flag", "dq_flag",
    ),
}

VIEW_CONFIG = {
    "historical-prep": {
        "directory": "historical_prep",
        "filename": "wilco_historical_records_prep.xlsx",
        "workbook_type": "Wilco Historical Records and Match Prep",
        "sheets": HISTORICAL_PREP_SHEETS,
    },
    "wilco-match": {
        "directory": "wilco_match",
        "filename": "wilco_match_management_{match_id}.xlsx",
        "workbook_type": "Wilco Match Management and Results with Records",
        "sheets": WILCO_MATCH_SHEETS,
    },
    "staff-match": {
        "directory": "staff_match",
        "filename": "staff_match_review_{match_id}.xlsx",
        "workbook_type": "Nationals Staff Match Review, Results with Records",
        "sheets": STAFF_MATCH_SHEETS,
    },
}

ALL_SHEETS = (
    *HISTORICAL_PREP_SHEETS,
    *WILCO_MATCH_SHEETS,
    *STAFF_MATCH_SHEETS,
)

SECONDS_COLUMNS = {
    "match_score", "score", "best_score", "most_recent_score", "avg_score",
    "median_score", "personal_record_score", "seconds_from_pr",
    "avg_string", "fastest_string", "avg_stage", "fastest_stage",
    "wilco_avg", "field_avg", "wilco_best", "field_best", "top_team_avg",
    "avg_gap_to_field", "best_gap_to_field", "stage_score", "match_total",
    "expected_score", "difference", "previous_score",
    "improvement_seconds", "previous_record_or_pr",
    "Wilco Avg", "Field Avg", "Wilco Best", "Field Best", "Top Team Avg",
    "Avg Gap To Field", "Wilco Gap To Field",
    "Best Score", "Most Recent Score", "Avg Score", "Median Score",
}


class AnalysisWorkbookError(RuntimeError):
    """Raised when local analysis inputs cannot be composed."""


@dataclass(frozen=True)
class AnalysisWorkbookResult:
    workbook_path: Path | None
    workbook_paths: dict[str, Path]
    tables_dir: Path
    selected_match_id: int
    selected_match_name: str
    sheet_names: dict[str, tuple[str, ...]]
    row_counts: dict[str, dict[str, int]]
    csv_row_counts: dict[str, int]
    chart_included: dict[str, bool]
    no_score_selected_match: bool
    excel_row_limit_notes: tuple[str, ...]
    competitor_source: str
    competitors_checked: int
    matched_competitors: int
    unmatched_competitors: int
    historical_score_rows_scanned: int


def build_analysis_workbook(
    *,
    output_root: Path | str,
    profile: TeamProfile,
    match_id: int | None = None,
    history_dir: Path | str | None = None,
    records_dir: Path | str | None = None,
    nationals_readiness_dir: Path | str | None = None,
    past_seasons: int = 2,
    workbook_name: str | None = None,
    include_all_teams: bool = True,
    include_validation: bool = True,
    view_set: str = "historical-prep",
    competitor_list_file: Path | str | None = None,
) -> AnalysisWorkbookResult:
    """Build the analysis tables and workbook without making API calls."""
    if view_set not in VIEW_SETS:
        raise AnalysisWorkbookError(
            f"--view-set must be one of: {', '.join(VIEW_SETS)}"
        )
    root = Path(output_root)
    history = Path(history_dir) if history_dir else root / "history"
    records = Path(records_dir) if records_dir else root / "records"
    readiness = (
        Path(nationals_readiness_dir)
        if nationals_readiness_dir
        else root / "nationals_readiness"
    )
    source_rows = _read_required(history / "history_source_matches.csv")
    participation = _read_required(history / "wilco_match_participation.csv")
    try:
        aliases = load_athlete_aliases()
    except AthleteAliasError as exc:
        raise AnalysisWorkbookError(str(exc)) from exc

    participation = apply_athlete_aliases(
        [row for row in participation if not _is_placeholder(row)],
        aliases,
    )
    selected_id = match_id or _latest_scored_match(participation)
    source_by_id = {row["match_id"]: row for row in source_rows}
    selected_source = source_by_id.get(str(selected_id), {})
    selected_scores_path = root / str(selected_id) / "tables" / "match_scores.csv"
    selected_scores = _read_required(selected_scores_path)
    selected_scores = apply_athlete_aliases(selected_scores, aliases)
    selected_name = (
        selected_source.get("match_name")
        or _first(selected_scores, "match_name")
        or f"Match {selected_id}"
    )
    selected_date = selected_source.get("match_date", "")

    cutoff_date = selected_date or datetime.now(timezone.utc).date().isoformat()
    seasons = _selected_seasons(source_rows, past_seasons, cutoff_date)
    scoped_sources = [
        row for row in source_rows
        if row.get("season_label") in seasons
        and (
            not row.get("match_date")
            or row.get("match_date", "") <= cutoff_date
        )
    ]
    scoped_ids = {row["match_id"] for row in scoped_sources}
    scoped_participation = [
        row for row in participation if row.get("match_id") in scoped_ids
    ]
    source_lookup = {row["match_id"]: row for row in scoped_sources}

    rankings = _load_rankings(root, scoped_ids, source_lookup, aliases)
    ranking_index = _ranking_index(rankings)
    squad_index = _load_squad_places(root, scoped_ids)
    match_history = _match_history_rows(
        scoped_participation, ranking_index, squad_index
    )
    athlete_performance = _athlete_performance_rows(
        scoped_participation,
        _read_optional(records / "personal_records.csv"),
    )
    all_scores = _load_all_scores(root, scoped_ids, aliases)
    all_local_scores = _load_all_local_match_scores(root, source_by_id, aliases)
    stage_values = _stage_value_rows(
        root, scoped_ids, source_lookup, profile, aliases,
        scoped_participation,
    )
    capability_columns, capability = _capability_rows(stage_values)
    field_comparison = _field_comparison_rows(all_scores, profile)
    competitor_source, competitors = _competitor_set(
        selected_scores=selected_scores,
        competitor_list_file=Path(competitor_list_file)
        if competitor_list_file else None,
    )
    competitive_analysis = _competitive_analysis_rows(
        competitors, all_local_scores
    )
    competitive_stats = _competitive_stats(
        competitors, competitive_analysis, all_local_scores
    )
    match_results = _selected_match_rows(
        selected_id, selected_name, selected_date, selected_scores,
        profile, ranking_index,
    )
    all_match_results = _all_selected_match_rows(
        selected_id, selected_name, selected_date, selected_scores,
        ranking_index,
    )
    stage_review = _selected_stage_rows(
        root, selected_id, selected_name, selected_scores, profile,
        ranking_index, aliases,
    )
    validation = _validation_rows(
        root, selected_id, selected_name, include_validation
    )
    records_rows = _records_rows(records, seasons)
    highlights = _highlight_rows(
        records, selected_id, selected_name, selected_date
    )
    coach_queue = [
        row for row in validation
        if row.get("severity") in {"ERROR", "WARNING", "REVIEW"}
        and profile.matches_name(row.get("team_name"))
    ]
    field_by_discipline = _field_by_discipline_rows(
        selected_scores, profile
    )
    selected_match_bests = _selected_match_best_rows(
        highlights, selected_scores
    )
    no_score = not any(_number(row.get("match_score_seconds"))
                       for row in selected_scores
                       if profile.matches_name(row.get("team_name")))
    quality = _quality_rows(
        source_rows=scoped_sources,
        selected_match_id=selected_id,
        selected_no_score=no_score,
        selected_scores=selected_scores,
        profile=profile,
        participation=participation,
        aliases_applied=sum(
            row.get("original_athlete_name", "") != row.get("athlete_name", "")
            for row in participation
        ),
        validation_rows=validation,
        readiness_dir=readiness,
    )
    historical_quality = _with_quality_context(
        quality,
        [
            ("csv_full_table", "athlete_discipline_stage_values.csv",
             len(stage_values),
             "Long-form stage values are CSV-only capability detail."),
            ("csv_full_table", "all_historical_match_scores_full.csv",
             len(all_scores),
             "Full all-team historical match scores are written to CSV."),
            ("competitive_analysis", "competitors_checked",
             competitive_stats["competitors_checked"],
             "Competitive Analysis competitors checked."),
            ("competitive_analysis", "matched_competitors",
             competitive_stats["matched_competitors"],
             "Competitors matched to at least one local historical score."),
            ("competitive_analysis", "unmatched_competitors",
             competitive_stats["unmatched_competitors"],
             "Competitors with no matching local historical scores."),
            ("competitive_analysis", "local_historical_score_rows_scanned",
             competitive_stats["historical_score_rows_scanned"],
             "Competitive Analysis scanned all local processed match score rows."),
            ("competitive_analysis", "competitor_source_used",
             competitive_stats["competitors_checked"],
             competitor_source),
            ("competitive_analysis", "selected_match_used",
             selected_id,
             f"Selected match {selected_id}: {selected_name}."),
        ],
    )
    wilco_match_quality = _with_quality_context(
        quality,
        [("selected_match", "wilco_scope", len(match_results),
          "Workbook is scoped to Wilco Shooting Sports entries.")]
    )
    staff_match_quality = _with_quality_context(
        quality,
        [("selected_match", "all_competitors", len(all_match_results),
          "Workbook includes all available competitors for the selected match.")]
    )
    rows_by_filename: dict[str, list[dict[str, Any]]] = {
        "wilco_historical_score_history.csv": _project_rows(
            match_history,
            _columns_for("wilco_historical_score_history.csv"),
        ),
        "athlete_performance_by_discipline.csv": athlete_performance,
        "athlete_capability_matrix.csv": capability,
        "athlete_discipline_stage_values.csv": stage_values,
        "all_historical_match_scores_full.csv": all_local_scores,
        "wilco_vs_field_by_discipline.csv": _historical_field_rows(
            field_comparison
        ),
        "competitive_analysis.csv": competitive_analysis,
        "wilco_match_results.csv": _project_rows(
            match_results, _columns_for("wilco_match_results.csv"),
        ),
        "wilco_stage_review.csv": _project_rows(
            stage_review, _columns_for("wilco_stage_review.csv"),
        ),
        "match_records_and_prs.csv": _project_rows(
            highlights, _columns_for("match_records_and_prs.csv"),
        ),
        "coach_review_queue.csv": _project_rows(
            coach_queue, _columns_for("coach_review_queue.csv"),
        ),
        "all_competitor_match_results.csv": all_match_results,
        "published_ranking_detail.csv": _selected_ranking_detail(
            rankings, selected_id
        ),
        "validation_findings.csv": _project_rows(
            validation, _columns_for("validation_findings.csv"),
        ),
        "field_by_discipline.csv": field_by_discipline,
        "records_match_bests.csv": selected_match_bests,
        "records_and_prs.csv": records_rows,
        "historical_prep_data_quality_notes.csv": historical_quality,
        "wilco_match_data_quality_notes.csv": wilco_match_quality,
        "staff_match_data_quality_notes.csv": staff_match_quality,
    }
    columns_by_filename = {
        filename: tuple(columns)
        for _, filename, columns in ALL_SHEETS
    }
    columns_by_filename.update(LONG_CSVS)
    columns_by_filename["athlete_capability_matrix.csv"] = capability_columns

    analysis_dir = root / "analysis"
    tables_dir = analysis_dir / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)
    for filename, rows in rows_by_filename.items():
        _write_csv(tables_dir / filename, columns_by_filename[filename], rows)

    views = (
        ("historical-prep", "wilco-match", "staff-match")
        if view_set == "all" else (view_set,)
    )
    workbook_paths: dict[str, Path] = {}
    sheet_names: dict[str, tuple[str, ...]] = {}
    row_counts: dict[str, dict[str, int]] = {}
    chart_included: dict[str, bool] = {}
    row_limit_notes: list[str] = []
    source_counts = {
        "source_matches_count": len(scoped_sources),
        "wilco_score_history_rows": len(match_history),
        "all_historical_score_rows": len(all_scores),
        "all_local_score_rows_scanned": len(all_local_scores),
        "competitive_analysis_rows": len(competitive_analysis),
        "selected_match_score_rows": len(selected_scores),
        "selected_match_validation_rows": len(validation),
        "long_stage_value_rows": len(stage_values),
    }
    for view in views:
        config = VIEW_CONFIG[view]
        filename = (
            workbook_name if workbook_name and len(views) == 1
            else config["filename"].format(match_id=selected_id)
        )
        workbook_path = analysis_dir / config["directory"] / filename
        if workbook_path.suffix.casefold() != ".xlsx":
            workbook_path = workbook_path.with_suffix(".xlsx")
        sheets = config["sheets"]
        cover = _cover_values(
            workbook_type=config["workbook_type"],
            workbook_path=workbook_path,
            profile=profile,
            selected_id=selected_id,
            selected_name=selected_name,
            selected_date=selected_date,
            seasons=seasons,
            scoped_sources=scoped_sources,
            scoped_participation=scoped_participation,
            source_counts=source_counts,
        )
        workbook_row_limit_notes = _row_limit_notes(
            sheets, rows_by_filename
        )
        row_limit_notes.extend(workbook_row_limit_notes)
        if workbook_row_limit_notes:
            cover["excel_row_limit_notes"] = " | ".join(
                workbook_row_limit_notes
            )
        chart_included[view] = _write_workbook(
            workbook_path, cover, rows_by_filename, columns_by_filename,
            sheets,
        )
        workbook_paths[view] = workbook_path
        sheet_names[view] = ("Cover", *(sheet for sheet, _, _ in sheets))
        row_counts[view] = {
            filename: len(rows_by_filename[filename])
            for _, filename, _ in sheets
        }
    return AnalysisWorkbookResult(
        workbook_path=next(iter(workbook_paths.values()), None),
        workbook_paths=workbook_paths,
        tables_dir=tables_dir,
        selected_match_id=selected_id,
        selected_match_name=selected_name,
        sheet_names=sheet_names,
        row_counts=row_counts,
        csv_row_counts={
            filename: len(rows) for filename, rows in rows_by_filename.items()
        },
        chart_included=chart_included,
        no_score_selected_match=no_score,
        excel_row_limit_notes=tuple(dict.fromkeys(row_limit_notes)),
        competitor_source=competitor_source,
        competitors_checked=competitive_stats["competitors_checked"],
        matched_competitors=competitive_stats["matched_competitors"],
        unmatched_competitors=competitive_stats["unmatched_competitors"],
        historical_score_rows_scanned=competitive_stats[
            "historical_score_rows_scanned"
        ],
    )


def _columns_for(filename: str) -> tuple[str, ...]:
    for _, candidate, columns in ALL_SHEETS:
        if candidate == filename:
            return tuple(columns)
    if filename in LONG_CSVS:
        return LONG_CSVS[filename]
    raise AnalysisWorkbookError(f"No column definition for {filename}")


def _project_rows(
    rows: list[dict[str, Any]], columns: tuple[str, ...]
) -> list[dict[str, Any]]:
    return [{column: row.get(column, "") for column in columns} for row in rows]


def _with_quality_context(
    rows: list[dict[str, Any]],
    additions: list[tuple[str, str, int, str]],
) -> list[dict[str, Any]]:
    result = list(rows)
    for area, issue_type, count, notes in additions:
        result.append({
            "area": area,
            "issue_type": issue_type,
            "affected_rows": count,
            "severity": "INFO",
            "notes": notes,
        })
    return result


def _historical_field_rows(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {
            "discipline": row.get("discipline", ""),
            "Wilco Entries": row.get("wilco_entries", ""),
            "Field Entries": row.get("field_entries", ""),
            "Wilco Avg": row.get("wilco_avg", ""),
            "Field Avg": row.get("field_avg", ""),
            "Wilco Best": row.get("wilco_best", ""),
            "Field Best": row.get("field_best", ""),
            "Top Team Avg": row.get("top_team_avg", ""),
            "Avg Gap To Field": row.get("avg_gap_to_field", ""),
        }
        for row in rows
    ]


def _cover_values(
    *,
    workbook_type: str,
    workbook_path: Path,
    profile: TeamProfile,
    selected_id: int,
    selected_name: str,
    selected_date: str,
    seasons: tuple[str, ...],
    scoped_sources: list[dict[str, str]],
    scoped_participation: list[dict[str, Any]],
    source_counts: dict[str, int],
) -> dict[str, Any]:
    return {
        "workbook_type": workbook_type,
        "workbook_name": workbook_path.stem.replace("_", " ").title(),
        "generated_at": datetime.now(timezone.utc)
        .replace(microsecond=0).isoformat(),
        "team_key": profile.team_key,
        "team_name": profile.team_name,
        "team_number": profile.team_number,
        "selected_match_id": selected_id,
        "selected_match_name": selected_name,
        "selected_match_date": selected_date,
        "seasons_included": ", ".join(seasons),
        "wilco_matches_included": len(scoped_sources),
        "wilco_athletes_included": len({
            (row.get("athlete_name"), row.get("athlete_id"))
            for row in scoped_participation
        }),
        "source_row_counts": "; ".join(
            f"{key}={value}" for key, value in source_counts.items()
        ),
        "excel_row_limit_notes": "No workbook tabs exceed Excel row limits.",
        "notes": (
            "Lower score/time is better. Blank athlete names and athlete ID "
            "9999 are excluded by default. Stage/string values are capability "
            "benchmarks only, not formal records."
        ),
    }


def _row_limit_notes(
    sheets: tuple[tuple[str, str, tuple[str, ...]], ...],
    rows_by_filename: dict[str, list[dict[str, Any]]],
) -> list[str]:
    notes = []
    for sheet_name, filename, _ in sheets:
        row_count = len(rows_by_filename.get(filename, []))
        if row_count + 1 > EXCEL_MAX_ROWS:
            notes.append(
                f"{sheet_name}: {row_count} source rows exceed Excel's "
                f"{EXCEL_MAX_ROWS:,}-row sheet limit; full data is in "
                f"output/analysis/tables/{filename}."
            )
    return notes


def _competitor_set(
    *,
    selected_scores: list[dict[str, Any]],
    competitor_list_file: Path | None,
) -> tuple[str, list[dict[str, Any]]]:
    if competitor_list_file:
        rows = _read_required(competitor_list_file)
        competitors = []
        for row in rows:
            athlete_name = (
                row.get("athlete_name") or row.get("name") or ""
            ).strip()
            athlete_id = (
                row.get("athlete_id") or row.get("ath_id") or ""
            ).strip()
            if not athlete_name or athlete_id == PLACEHOLDER_ID:
                continue
            competitors.append({
                "athlete_name": athlete_name,
                "athlete_id": athlete_id,
                "team_name": (row.get("team_name") or row.get("team") or "")
                .strip(),
                "gender": row.get("gender", "").strip(),
                "class": row.get("class", "").strip(),
                "division": row.get("division", "").strip(),
                "discipline": row.get("discipline", "").strip(),
            })
        return (
            f"competitor-list-file: {competitor_list_file}",
            _dedupe_competitors(competitors),
        )

    competitors = []
    for row in selected_scores:
        if _is_placeholder(row):
            continue
        competitors.append({
            "athlete_name": row.get("athlete_name", "").strip(),
            "athlete_id": row.get("athlete_id", "").strip(),
            "team_name": row.get("team_name", "").strip(),
            "gender": row.get("gender", "").strip(),
            "class": row.get("class", "").strip(),
            "division": _division(row.get("class", "")),
            "discipline": row.get("discipline", "").strip(),
        })
    return ("selected match entries", _dedupe_competitors(competitors))


def _dedupe_competitors(
    competitors: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    seen = set()
    results = []
    for row in competitors:
        identity = row.get("athlete_id") or _normalized_name(
            row.get("athlete_name", "")
        )
        key = (identity, row.get("discipline", ""))
        if key in seen:
            continue
        seen.add(key)
        results.append(row)
    return results


def _load_all_local_match_scores(
    root: Path,
    sources: dict[str, dict[str, str]],
    aliases: tuple[Any, ...],
) -> list[dict[str, Any]]:
    results = []
    for path in sorted(root.rglob("tables/match_scores.csv")):
        match_id = path.parent.parent.name
        if not match_id.isdigit():
            continue
        source = sources.get(match_id, {})
        rows = apply_athlete_aliases(_read_optional(path), aliases)
        for row in rows:
            if _is_placeholder(row):
                continue
            score = _number(row.get("match_score_seconds"))
            if score is None:
                continue
            results.append({
                "match_id": row.get("match_id") or match_id,
                "match_name": row.get("match_name")
                or source.get("match_name", ""),
                "match_date": source.get("match_date", ""),
                "discipline": row.get("discipline", "").strip(),
                "athlete_id": row.get("athlete_id", "").strip(),
                "athlete_name": row.get("athlete_name", "").strip(),
                "team_name": row.get("team_name", "").strip(),
                "class": row.get("class", "").strip(),
                "gender": row.get("gender", "").strip(),
                "match_score_seconds": _display(score),
                "dnf_flag": row.get("dnf_flag", ""),
                "dq_flag": row.get("dq_flag", ""),
                "_score": score,
                "_name_key": _normalized_name(row.get("athlete_name", "")),
            })
    return results


def _competitive_analysis_rows(
    competitors: list[dict[str, Any]],
    scores: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scores:
        if row.get("athlete_id"):
            by_id[row["athlete_id"]].append(row)
        if row.get("_name_key"):
            by_name[row["_name_key"]].append(row)

    results = []
    for competitor in competitors:
        athlete_id = competitor.get("athlete_id", "")
        name_key = _normalized_name(competitor.get("athlete_name", ""))
        if athlete_id and athlete_id in by_id:
            matched = by_id[athlete_id]
            method = "athlete_id"
        elif name_key and name_key in by_name:
            matched = by_name[name_key]
            method = "normalized_name"
        else:
            matched = []
            method = "unmatched"

        requested_discipline = competitor.get("discipline", "")
        disciplines = (
            [requested_discipline] if requested_discipline
            else sorted({row.get("discipline", "") for row in matched
                         if row.get("discipline")})
        )
        if not disciplines:
            disciplines = [requested_discipline]

        for discipline in disciplines:
            discipline_rows = [
                row for row in matched
                if not discipline or row.get("discipline") == discipline
            ]
            results.append(_competitive_row(
                competitor, discipline, discipline_rows, method, bool(matched)
            ))
    return sorted(results, key=lambda row: (
        row["Discipline"], row["Athlete Name"], row["Team"]
    ))


def _competitive_row(
    competitor: dict[str, Any],
    discipline: str,
    rows: list[dict[str, Any]],
    method: str,
    matched_any: bool,
) -> dict[str, Any]:
    numeric_rows = [
        row for row in rows if _number(row.get("match_score_seconds")) is not None
    ]
    values = [_number(row["match_score_seconds"]) for row in numeric_rows]
    values = [value for value in values if value is not None]
    best = min(
        numeric_rows,
        key=lambda row: _number(row.get("match_score_seconds")) or 999999,
        default={},
    )
    most_recent = max(
        numeric_rows,
        key=lambda row: (
            row.get("match_date", ""), _integer(row.get("match_id"))
        ),
        default={},
    )
    if values:
        note = "Historical scores found."
    elif matched_any and discipline:
        note = "No historical score found for this discipline"
    else:
        note = "No matching historical scores found"
    return {
        "Athlete Name": competitor.get("athlete_name", ""),
        "Athlete ID": competitor.get("athlete_id", ""),
        "Team": competitor.get("team_name", ""),
        "Gender": competitor.get("gender", ""),
        "Division": competitor.get("division", "")
        or _division(competitor.get("class", "")),
        "Class": competitor.get("class", ""),
        "Discipline": discipline,
        "Number Of Matches": len({
            row.get("match_id", "") for row in numeric_rows
        }),
        "Best Score": _display(min(values) if values else None),
        "Match With Best Score": best.get("match_name", ""),
        "Best Score Date": best.get("match_date", ""),
        "Most Recent Score": _display(
            _number(most_recent.get("match_score_seconds"))
        ),
        "Most Recent Match": most_recent.get("match_name", ""),
        "Most Recent Date": most_recent.get("match_date", ""),
        "Avg Score": _display(statistics.mean(values) if values else None),
        "Median Score": _display(statistics.median(values) if values else None),
        "Data Match Method": method if values or matched_any else "unmatched",
        "Competitive Note": note,
    }


def _competitive_stats(
    competitors: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    scores: list[dict[str, Any]],
) -> dict[str, int]:
    matched_keys = {
        (
            row.get("Athlete ID")
            or _normalized_name(row.get("Athlete Name", ""))
        )
        for row in rows
        if row.get("Data Match Method") != "unmatched"
    }
    checked_keys = {
        (
            row.get("athlete_id")
            or _normalized_name(row.get("athlete_name", ""))
        )
        for row in competitors
    }
    return {
        "competitors_checked": len(checked_keys),
        "matched_competitors": len(matched_keys),
        "unmatched_competitors": len(checked_keys - matched_keys),
        "historical_score_rows_scanned": len(scores),
    }


def _match_history_rows(
    rows: list[dict[str, Any]],
    rankings: dict[tuple[str, str, str], list[dict[str, Any]]],
    squads: dict[tuple[str, str, str], str],
) -> list[dict[str, Any]]:
    results = []
    for row in rows:
        key = (row["match_id"], row["athlete_name"], row["discipline"])
        rank_rows = rankings.get(key, [])
        preferred = _preferred_ranking(rank_rows)
        results.append({
            "match_id": row["match_id"],
            "match_name": row["match_name"],
            "match_date": row["match_date"],
            "season_label": row["season_label"],
            "athlete_name": row["athlete_name"],
            "athlete_id": row["athlete_id"],
            "division": row.get("division", ""),
            "class": row.get("class", ""),
            "gender": row.get("gender", ""),
            "discipline": row["discipline"],
            "match_score": row.get("score", ""),
            "overall_place": _place_for(rank_rows, "all"),
            "gender_place": _place_for(rank_rows, "gender"),
            "division_place": _place_for(rank_rows, "division"),
            "class_place": _place_for(rank_rows, "class"),
            "squad_place": squads.get(key, ""),
            "rank_scope": preferred.get("rank_scope", ""),
            "award_scope": preferred.get("award_scope", ""),
            "data_status": row.get("data_status", ""),
            "notes": row.get("notes", ""),
        })
    return sorted(results, key=_history_sort)


def _athlete_performance_rows(
    participation: list[dict[str, Any]],
    personal_records: list[dict[str, str]],
) -> list[dict[str, Any]]:
    pr_index = {
        (row.get("athlete_id", ""), row.get("discipline", "")): row
        for row in personal_records
    }
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in participation:
        grouped[(row["athlete_name"], row["athlete_id"],
                 row["discipline"])].append(row)
    results = []
    for key, rows in grouped.items():
        scored = [row for row in rows if _number(row.get("score")) is not None]
        ordered = sorted(rows, key=lambda row: (
            row.get("match_date", ""), _integer(row.get("match_id"))
        ))
        latest = ordered[-1]
        scores = [_number(row["score"]) for row in scored]
        numeric = [score for score in scores if score is not None]
        best = min(scored, key=lambda row: _number(row["score"]) or 999999) \
            if scored else {}
        pr = pr_index.get((key[1], key[2]), {})
        latest_score = _number(latest.get("score"))
        pr_score = _number(pr.get("personal_record_score"))
        results.append({
            "athlete_name": key[0], "athlete_id": key[1],
            "discipline": key[2],
            "number_of_matches": len({row["match_id"] for row in rows}),
            "scored_matches_count": len(scored),
            "best_score": _display(min(numeric) if numeric else None),
            "match_with_best_score": best.get("match_name", ""),
            "best_score_date": best.get("match_date", ""),
            "most_recent_score": _display(latest_score),
            "most_recent_match": latest.get("match_name", ""),
            "most_recent_date": latest.get("match_date", ""),
            "avg_score": _display(statistics.mean(numeric) if numeric else None),
            "median_score": _display(
                statistics.median(numeric) if numeric else None
            ),
            "personal_record_score": _display(pr_score),
            "seconds_from_pr": _display(
                latest_score - pr_score
                if latest_score is not None and pr_score is not None else None
            ),
            "trend_note": _trend_note(scored),
            "confidence_level": _confidence(len(scored)),
        })
    return sorted(results, key=lambda row: (
        row["athlete_name"], row["discipline"]
    ))


def _all_team_rows(
    scores: list[dict[str, Any]],
    sources: dict[str, dict[str, str]],
    profile: TeamProfile,
    include_all: bool,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in scores:
        if not include_all and not profile.matches_name(row.get("team_name")):
            continue
        grouped[(row["match_id"], row["discipline"],
                 row.get("team_name", ""))].append(row)
    results = []
    for key, rows in grouped.items():
        values = sorted(
            value for value in (_number(row.get("match_score_seconds"))
                                for row in rows) if value is not None
        )
        if not values:
            continue
        source = sources.get(key[0], {})
        results.append({
            "match_id": key[0],
            "match_name": source.get("match_name")
            or rows[0].get("match_name", ""),
            "match_date": source.get("match_date", ""),
            "discipline": key[1], "team_name": key[2], "team_id": "",
            "athlete_count": len({
                (row.get("athlete_id"), row.get("athlete_name"))
                for row in rows if row.get("athlete_name")
            }),
            "entry_count": len(values),
            "avg_score": _display(statistics.mean(values)),
            "median_score": _display(statistics.median(values)),
            "best_score": _display(min(values)),
            "top_4_average": _display(statistics.mean(values[:4])),
            "squad_count": "",
            "is_wilco": _boolean(profile.matches_name(key[2])),
            "notes": "Team ID and squad count are unavailable in match scores.",
        })
    return sorted(results, key=lambda row: (
        row["match_date"], row["match_id"], row["discipline"],
        row["team_name"]
    ))


def _stage_value_rows(
    root: Path,
    match_ids: set[str],
    sources: dict[str, dict[str, str]],
    profile: TeamProfile,
    aliases: tuple[Any, ...],
    participation: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    score_index = {
        (row["match_id"], row["athlete_id"], row["discipline"]):
        row.get("score", "") for row in participation
    }
    results = []
    for match_id in sorted(match_ids, key=_integer):
        path = root / match_id / "tables" / "stage_scores.csv"
        rows = apply_athlete_aliases(_read_optional(path), aliases)
        source = sources.get(match_id, {})
        for row in rows:
            if not profile.matches_name(row.get("team_name")) \
                    or _is_placeholder(row):
                continue
            stage_score = _number(row.get("stage_score_seconds"))
            avg_string = _number(row.get("scored_avg_string_seconds"))
            fastest = _number(row.get("fastest_string_seconds"))
            if stage_score is None and avg_string is None and fastest is None:
                continue
            results.append({
                "athlete_name": row.get("athlete_name", ""),
                "athlete_id": row.get("athlete_id", ""),
                "discipline": row.get("discipline", ""),
                "stage_name": row.get("stage_name", ""),
                "season_label": source.get("season_label", ""),
                "match_id": match_id,
                "match_name": source.get("match_name")
                or row.get("match_name", ""),
                "avg_string": _display(avg_string),
                "fastest_string": _display(fastest),
                "avg_stage": _display(stage_score),
                "fastest_stage": _display(stage_score),
                "match_score": score_index.get((
                    match_id, row.get("athlete_id", ""),
                    row.get("discipline", "")
                ), ""),
                "notes": (
                    "Stage and string values are capability benchmarks only."
                ),
            })
    return sorted(results, key=lambda row: (
        row["athlete_name"], row["discipline"], row["stage_name"],
        row["season_label"], row["match_id"]
    ))


def _capability_rows(
    values: list[dict[str, Any]],
) -> tuple[tuple[str, ...], list[dict[str, Any]]]:
    stages = sorted({row["stage_name"] for row in values if row["stage_name"]})
    columns = ["athlete_name", "athlete_id", "discipline"]
    for stage in stages:
        columns.extend((
            f"{stage} Avg String", f"{stage} Fastest String",
            f"{stage} Avg Stage", f"{stage} Fastest Stage",
        ))
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in values:
        grouped[(row["athlete_name"], row["athlete_id"],
                 row["discipline"])].append(row)
    results = []
    for key, rows in grouped.items():
        result: dict[str, Any] = {
            "athlete_name": key[0], "athlete_id": key[1],
            "discipline": key[2],
        }
        for stage in stages:
            stage_rows = [row for row in rows if row["stage_name"] == stage]
            strings = [_number(row["avg_string"]) for row in stage_rows]
            fast_strings = [_number(row["fastest_string"]) for row in stage_rows]
            stage_scores = [_number(row["avg_stage"]) for row in stage_rows]
            strings = [value for value in strings if value is not None]
            fast_strings = [value for value in fast_strings if value is not None]
            stage_scores = [value for value in stage_scores if value is not None]
            result[f"{stage} Avg String"] = _display(
                statistics.mean(strings) if strings else None
            )
            result[f"{stage} Fastest String"] = _display(
                min(fast_strings) if fast_strings else None
            )
            result[f"{stage} Avg Stage"] = _display(
                statistics.mean(stage_scores) if stage_scores else None
            )
            result[f"{stage} Fastest Stage"] = _display(
                min(stage_scores) if stage_scores else None
            )
        results.append(result)
    return tuple(columns), sorted(results, key=lambda row: (
        row["athlete_name"], row["discipline"]
    ))


def _field_comparison_rows(
    scores: list[dict[str, Any]], profile: TeamProfile
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scores:
        if _number(row.get("match_score_seconds")) is not None:
            grouped[row["discipline"]].append(row)
    results = []
    for discipline, rows in grouped.items():
        wilco = [_number(row["match_score_seconds"]) for row in rows
                 if profile.matches_name(row.get("team_name"))]
        field = [_number(row["match_score_seconds"]) for row in rows
                 if not profile.matches_name(row.get("team_name"))]
        wilco_values = [value for value in wilco if value is not None]
        field_values = [value for value in field if value is not None]
        if not wilco_values:
            continue
        team_values: dict[str, list[float]] = defaultdict(list)
        for row in rows:
            value = _number(row["match_score_seconds"])
            if value is not None:
                team_values[row.get("team_name", "")].append(value)
        top_team = min(
            (statistics.mean(sorted(values)[:4])
             for values in team_values.values() if values),
            default=None,
        )
        wilco_avg = statistics.mean(wilco_values)
        field_avg = statistics.mean(field_values) if field_values else None
        wilco_best = min(wilco_values)
        field_best = min(field_values) if field_values else None
        avg_gap = wilco_avg - field_avg if field_avg is not None else None
        best_gap = wilco_best - field_best if field_best is not None else None
        results.append({
            "discipline": discipline, "wilco_entries": len(wilco_values),
            "field_entries": len(field_values),
            "wilco_avg": _display(wilco_avg),
            "field_avg": _display(field_avg),
            "wilco_best": _display(wilco_best),
            "field_best": _display(field_best),
            "top_team_avg": _display(top_team),
            "avg_gap_to_field": _display(avg_gap),
            "best_gap_to_field": _display(best_gap),
            "trend_note": (
                "Negative gaps mean Wilco is faster/better than the field. "
                + ("Wilco average is faster." if avg_gap is not None
                   and avg_gap < 0 else "Wilco average trails the field."
                   if avg_gap is not None else "Field comparison unavailable.")
            ),
        })
    return sorted(results, key=lambda row: row["discipline"])


def _selected_match_rows(
    match_id: int, match_name: str, match_date: str,
    scores: list[dict[str, Any]], profile: TeamProfile,
    rankings: dict[tuple[str, str, str], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    results = []
    for row in scores:
        if not profile.matches_name(row.get("team_name")) \
                or _is_placeholder(row):
            continue
        key = (str(match_id), row["athlete_name"], row["discipline"])
        preferred = _preferred_ranking(rankings.get(key, []))
        score = _number(row.get("match_score_seconds"))
        results.append({
            "match_id": match_id, "match_name": match_name,
            "match_date": match_date, "athlete_name": row["athlete_name"],
            "athlete_id": row["athlete_id"],
            "discipline": row["discipline"],
            "division": _division(row.get("class", "")),
            "class": row.get("class", ""), "gender": row.get("gender", ""),
            "match_score": _display(score),
            "current_rank": preferred.get("place", ""),
            "rank_scope": preferred.get("rank_scope", ""),
            "award_scope": preferred.get("award_scope", ""),
            "field_size": preferred.get("field_size", ""),
            "notes": (
                "No scored result is currently available."
                if score is None else ""
            ),
        })
    return sorted(results, key=lambda row: (
        row["athlete_name"], row["discipline"]
    ))


def _all_selected_match_rows(
    match_id: int, match_name: str, match_date: str,
    scores: list[dict[str, Any]],
    rankings: dict[tuple[str, str, str], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    results = []
    for row in scores:
        if _is_placeholder(row):
            continue
        key = (str(match_id), row.get("athlete_name", ""),
               row.get("discipline", ""))
        preferred = _preferred_ranking(rankings.get(key, []))
        score = _number(row.get("match_score_seconds"))
        results.append({
            "athlete_name": row.get("athlete_name", ""),
            "team_name": row.get("team_name", ""),
            "discipline": row.get("discipline", ""),
            "division": _division(row.get("class", "")),
            "class": row.get("class", ""),
            "gender": row.get("gender", ""),
            "match_score": _display(score),
            "place": preferred.get("place", ""),
            "rank_scope": preferred.get("rank_scope", ""),
            "field_size": preferred.get("field_size", ""),
            "notes": (
                f"Selected match {match_id}: {match_name} "
                f"({match_date})."
            ),
        })
    return sorted(results, key=lambda row: (
        row["discipline"], _integer(row["place"]) or 999999,
        row["athlete_name"]
    ))


def _selected_stage_rows(
    root: Path, match_id: int, match_name: str,
    scores: list[dict[str, Any]], profile: TeamProfile,
    rankings: dict[tuple[str, str, str], list[dict[str, Any]]],
    aliases: tuple[Any, ...],
) -> list[dict[str, Any]]:
    score_index = {
        (row["athlete_id"], row["discipline"]):
        row.get("match_score_seconds", "") for row in scores
    }
    rows = apply_athlete_aliases(_read_optional(
        root / str(match_id) / "tables" / "stage_scores.csv"
    ), aliases)
    results = []
    for row in rows:
        if not profile.matches_name(row.get("team_name")) \
                or _is_placeholder(row):
            continue
        key = (str(match_id), row["athlete_name"], row["discipline"])
        preferred = _preferred_ranking(rankings.get(key, []))
        stage_score = _number(row.get("stage_score_seconds"))
        results.append({
            "match_id": match_id, "match_name": match_name,
            "athlete_name": row["athlete_name"],
            "athlete_id": row["athlete_id"],
            "discipline": row["discipline"],
            "stage_name": row.get("stage_name", ""),
            "stage_score": _display(stage_score),
            "fastest_string": _display(
                _number(row.get("fastest_string_seconds"))
            ),
            "avg_string": _display(
                _number(row.get("scored_avg_string_seconds"))
            ),
            "string_count": 4 if stage_score is not None else 0,
            "match_total": score_index.get((
                row["athlete_id"], row["discipline"]
            ), ""),
            "current_rank": preferred.get("place", ""),
            "rank_scope": preferred.get("rank_scope", ""),
            "notes": (
                "Unscored stage; retained as entry context."
                if stage_score is None else
                "Capability detail only; not a formal stage record."
            ),
        })
    return sorted(results, key=lambda row: (
        row["athlete_name"], row["discipline"], row["stage_name"]
    ))


def _selected_ranking_detail(
    rankings: list[dict[str, Any]], match_id: int
) -> list[dict[str, Any]]:
    columns = _columns_for("published_ranking_detail.csv")
    return _project_rows(
        [row for row in rankings if _integer(row.get("match_id")) == match_id],
        columns,
    )


def _field_by_discipline_rows(
    scores: list[dict[str, Any]], profile: TeamProfile
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scores:
        if _is_placeholder(row):
            continue
        if _number(row.get("match_score_seconds")) is not None:
            grouped[row.get("discipline", "")].append(row)
    results = []
    for discipline, rows in grouped.items():
        values = [
            _number(row.get("match_score_seconds")) for row in rows
        ]
        values = [value for value in values if value is not None]
        wilco_values = [
            _number(row.get("match_score_seconds")) for row in rows
            if profile.matches_name(row.get("team_name"))
        ]
        wilco_values = [
            value for value in wilco_values if value is not None
        ]
        team_values: dict[str, list[float]] = defaultdict(list)
        athletes = set()
        for row in rows:
            score = _number(row.get("match_score_seconds"))
            if score is None:
                continue
            team_values[row.get("team_name", "")].append(score)
            athletes.add((row.get("athlete_id", ""), row.get("athlete_name", "")))
        field_avg = statistics.mean(values) if values else None
        wilco_avg = statistics.mean(wilco_values) if wilco_values else None
        results.append({
            "discipline": discipline,
            "teams_count": len([team for team in team_values if team]),
            "athletes_count": len(athletes),
            "entries_count": len(values),
            "field_avg": _display(field_avg),
            "field_best": _display(min(values) if values else None),
            "top_team_avg": _display(min(
                (statistics.mean(sorted(team_scores)[:4])
                 for team_scores in team_values.values() if team_scores),
                default=None,
            )),
            "Wilco Avg": _display(wilco_avg),
            "Wilco Best": _display(min(wilco_values) if wilco_values else None),
            "Wilco Gap To Field": _display(
                wilco_avg - field_avg
                if wilco_avg is not None and field_avg is not None else None
            ),
        })
    return sorted(results, key=lambda row: row["discipline"])


def _selected_match_best_rows(
    highlights: list[dict[str, Any]], scores: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    team_lookup = {
        (row.get("athlete_name", ""), row.get("discipline", "")):
        row.get("team_name", "") for row in scores
    }
    rows = []
    for row in highlights:
        rows.append({
            "highlight_type": row.get("highlight_type", ""),
            "athlete_name": row.get("athlete_name", ""),
            "team_name": team_lookup.get((
                row.get("athlete_name", ""), row.get("discipline", "")
            ), ""),
            "discipline": row.get("discipline", ""),
            "score": row.get("score", ""),
            "record_scope": row.get("record_scope", ""),
            "notes": row.get("notes", ""),
        })
    return rows


def _validation_rows(
    root: Path, match_id: int, match_name: str, include: bool
) -> list[dict[str, Any]]:
    if not include:
        return []
    findings = _read_optional(
        root / str(match_id) / "validation" / "validation_findings.csv"
    )
    results = []
    for row in findings:
        severity = row.get("severity", "")
        results.append({
            "match_id": match_id, "match_name": match_name,
            "severity": severity, "finding_type": row.get("finding_type", ""),
            "athlete_name": row.get("athlete_name", ""),
            "team_name": row.get("team_name", ""),
            "discipline": row.get("discipline", ""),
            "stage_name": row.get("stage_name", ""),
            "score": row.get("actual_value", ""),
            "expected_score": row.get("expected_value", ""),
            "difference": row.get("difference", ""),
            "finding_message": row.get("message", ""),
            "recommended_review": (
                "Correct source or parser mismatch."
                if severity == "ERROR" else
                "Coach or match staff should review context."
                if severity in {"WARNING", "REVIEW"} else
                "Context only."
            ),
            "notes": (
                "Review findings are not automatically invalid scores."
            ),
        })
    return results


def _records_rows(records: Path, seasons: tuple[str, ...]) -> list[dict[str, Any]]:
    results = []
    for row in _read_optional(records / "personal_records.csv"):
        if row.get("personal_record_season_label") not in seasons:
            continue
        results.append({
            "record_type": "personal_record",
            "athlete_name": row.get("athlete_name", ""),
            "discipline": row.get("discipline", ""),
            "score": row.get("personal_record_score", ""),
            "match_name": row.get("personal_record_match_name", ""),
            "match_date": row.get("personal_record_match_date", ""),
            "season_label": row.get("personal_record_season_label", ""),
            "previous_score": "", "improvement_seconds": "",
            "display_eligible": row.get("display_eligible", ""),
            "confidence_level": row.get("confidence_level", ""),
            "notes": row.get("notes", ""),
        })
    for filename, record_type in (
        ("wilco_all_time_records.csv", "wilco_all_time_record"),
        ("wilco_team_season_records.csv", "team_season_record"),
    ):
        for row in _read_optional(records / filename):
            if record_type == "team_season_record" \
                    and row.get("season_label") not in seasons:
                continue
            results.append({
                "record_type": record_type,
                "athlete_name": row.get("athlete_name", ""),
                "discipline": row.get("discipline", ""),
                "score": row.get("score", ""),
                "match_name": row.get("match_name", ""),
                "match_date": row.get("match_date", ""),
                "season_label": row.get("season_label", ""),
                "previous_score": "", "improvement_seconds": "",
                "display_eligible": "true",
                "confidence_level": row.get("confidence_level", ""),
                "notes": row.get("notes", ""),
            })
    for row in _read_optional(records / "recent_pr_highlights.csv"):
        if row.get("season_label") not in seasons:
            continue
        results.append({
            "record_type": "recent_pr_highlight",
            "athlete_name": row.get("athlete_name", ""),
            "discipline": row.get("discipline", ""),
            "score": row.get("new_pr_score", ""),
            "match_name": row.get("match_name", ""),
            "match_date": row.get("match_date", ""),
            "season_label": row.get("season_label", ""),
            "previous_score": row.get("previous_pr_score", ""),
            "improvement_seconds": row.get("improvement_seconds", ""),
            "display_eligible": row.get("display_eligible", ""),
            "confidence_level": row.get("confidence_level", ""),
            "notes": row.get("display_note", ""),
        })
    return sorted(results, key=lambda row: (
        row["record_type"], row["discipline"], row["athlete_name"]
    ))


def _highlight_rows(
    records: Path, match_id: int, match_name: str, match_date: str
) -> list[dict[str, Any]]:
    results = []
    for row in _read_optional(records / "new_personal_records_by_match.csv"):
        if _integer(row.get("match_id")) != match_id:
            continue
        results.append({
            "match_id": match_id, "match_name": match_name,
            "match_date": match_date, "athlete_name": row["athlete_name"],
            "discipline": row["discipline"],
            "highlight_type": row.get("pr_event_type", "new_personal_record"),
            "score": row.get("new_pr_score", ""),
            "previous_record_or_pr": row.get("previous_pr_score", ""),
            "improvement_seconds": row.get("improvement_seconds", ""),
            "record_scope": "personal_record",
            "display_eligible": "true",
            "notes": row.get("notes", ""),
        })
    for row in _read_optional(records / "match_bests.csv"):
        if _integer(row.get("match_id")) != match_id:
            continue
        results.append({
            "match_id": match_id, "match_name": match_name,
            "match_date": match_date, "athlete_name": row["athlete_name"],
            "discipline": row["discipline"], "highlight_type": "match_best",
            "score": row.get("score", ""), "previous_record_or_pr": "",
            "improvement_seconds": "",
            "record_scope": row.get("match_best_scope", ""),
            "display_eligible": "true", "notes": row.get("notes", ""),
        })
    return sorted(results, key=lambda row: (
        row["discipline"], row["highlight_type"], row["athlete_name"]
    ))


def _quality_rows(
    *, source_rows: list[dict[str, str]], selected_match_id: int,
    selected_no_score: bool, selected_scores: list[dict[str, Any]],
    profile: TeamProfile,
    participation: list[dict[str, Any]], aliases_applied: int,
    validation_rows: list[dict[str, Any]], readiness_dir: Path,
) -> list[dict[str, Any]]:
    partial = sum(row.get("data_status") == "partial" for row in source_rows)
    selected_team_scores = [
        row for row in selected_scores
        if profile.matches_name(row.get("team_name"))
    ]
    placeholders = sum(_is_placeholder(row) for row in selected_team_scores)
    ranking_gaps = sum(
        "No preferred individual ranking" in row.get("notes", "")
        for row in participation if row.get("match_id") == str(selected_match_id)
    )
    schedule_missing = sum(
        not (readiness_dir.parent / row["match_id"] / "raw"
             / f"{row['match_id']}_schedule.json").exists()
        for row in source_rows
    )
    rows = [
        {
            "area": "selected_match", "issue_type": "no_scores",
            "affected_rows": sum(
                _number(row.get("match_score_seconds")) is None
                for row in selected_team_scores
            ) if selected_no_score else 0,
            "severity": "INFO",
            "notes": (
                "Selected match has entries but no current Wilco scores; "
                "score-dependent cells remain blank."
                if selected_no_score else "Selected match has scored results."
            ),
        },
        {
            "area": "history", "issue_type": "partial_matches",
            "affected_rows": partial, "severity": "REVIEW",
            "notes": "Partial matches remain visible with data-status context.",
        },
        {
            "area": "raw_metadata", "issue_type": "missing_schedules",
            "affected_rows": schedule_missing, "severity": "INFO",
            "notes": "Schedule metadata is optional and does not block analysis.",
        },
        {
            "area": "raw_data", "issue_type": "empty_slot_matches_excluded",
            "affected_rows": sum(
                row.get("core_complete") != "true" for row in source_rows
            ), "severity": "INFO",
            "notes": "Matches without useful core JSON are not analyzed.",
        },
        {
            "area": "identity", "issue_type": "placeholder_rows_excluded",
            "affected_rows": placeholders, "severity": "WARNING",
            "notes": "Blank athlete names and athlete ID 9999 are excluded.",
        },
        {
            "area": "identity", "issue_type": "alias_mappings_applied",
            "affected_rows": aliases_applied, "severity": "INFO",
            "notes": "Configured athlete aliases were canonicalized.",
        },
        {
            "area": "rankings", "issue_type": "ranking_gaps",
            "affected_rows": ranking_gaps, "severity": "INFO",
            "notes": "Some entries do not have a matching published ranking.",
        },
        {
            "area": "validation", "issue_type": "validation_limitations",
            "affected_rows": len(validation_rows), "severity": "INFO",
            "notes": (
                "Validation findings are review aids; explainable findings "
                "are not automatically scoring errors."
            ),
        },
    ]
    return rows


def _load_rankings(
    root: Path, match_ids: set[str],
    sources: dict[str, dict[str, str]], aliases: tuple[Any, ...],
) -> list[dict[str, Any]]:
    results = []
    for match_id in sorted(match_ids, key=_integer):
        path = root / match_id / "tables" / "rankings.csv"
        rows = apply_athlete_aliases(_read_optional(path), aliases)
        source = sources.get(match_id, {})
        grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            if not row.get("athlete_name") or _is_placeholder(row):
                continue
            grouped[(row.get("discipline", ""), row.get("leaderboard_name", ""),
                     row.get("rank_scope", ""))].append(row)
        for row in rows:
            if not row.get("athlete_name") or _is_placeholder(row):
                continue
            group = grouped[(row.get("discipline", ""),
                             row.get("leaderboard_name", ""),
                             row.get("rank_scope", ""))]
            podium = next(
                (_number(item.get("score_seconds")) for item in group
                 if _integer(item.get("place")) == 3), None
            )
            score = _number(row.get("score_seconds"))
            field_size = _integer(row.get("field_size"))
            place = _integer(row.get("place"))
            results.append({
                "match_id": match_id,
                "match_name": source.get("match_name")
                or row.get("match_name", ""),
                "match_date": source.get("match_date", ""),
                "discipline": row.get("discipline", ""),
                "athlete_name": row.get("athlete_name", ""),
                "athlete_id": row.get("athlete_id", ""),
                "score": _display(score),
                "rank_type": row.get("rank_type", ""),
                "rank_scope": row.get("rank_scope", ""),
                "place": place or "",
                "field_size": field_size or "",
                "percentile": _display(
                    100 * (field_size - place + 1) / field_size
                    if field_size and place else None
                ),
                "margin_to_leader": row.get("margin_to_leader", ""),
                "margin_to_previous_place":
                    row.get("margin_to_previous_place", ""),
                "margin_to_podium_cutoff": _display(
                    score - podium if score is not None and podium is not None
                    else None
                ),
                "team_name": row.get("team_name", ""),
                "athlete_class": row.get("class", ""),
                "gender": row.get("gender_context", ""),
                "squad_name": "",
                "source_file": str(path),
                "_award_scope": row.get("award_scope", ""),
            })
    return sorted(results, key=lambda row: (
        row["match_date"], row["match_id"], row["discipline"],
        row["rank_scope"], _integer(row["place"])
    ))


def _ranking_index(
    rankings: list[dict[str, Any]],
) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    result: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rankings:
        row["award_scope"] = row.pop("_award_scope", "")
        result[(str(row["match_id"]), row["athlete_name"],
                row["discipline"])].append(row)
    return result


def _load_squad_places(
    root: Path, match_ids: set[str]
) -> dict[tuple[str, str, str], str]:
    result = {}
    for match_id in match_ids:
        for row in _read_optional(
            root / match_id / "tables" / "squad_results.csv"
        ):
            for number in range(1, 5):
                name = row.get(f"athlete_{number}_name", "")
                if name:
                    result[(match_id, name, row.get("discipline", ""))] = (
                        row.get("squad_place", "")
                    )
    return result


def _load_all_scores(
    root: Path, match_ids: set[str], aliases: tuple[Any, ...]
) -> list[dict[str, Any]]:
    results = []
    for match_id in sorted(match_ids, key=_integer):
        rows = apply_athlete_aliases(_read_optional(
            root / match_id / "tables" / "match_scores.csv"
        ), aliases)
        results.extend(row for row in rows if not _is_placeholder(row))
    return results


def _selected_seasons(
    source_rows: list[dict[str, str]], count: int, cutoff_date: str
) -> tuple[str, ...]:
    if count < 1:
        raise AnalysisWorkbookError("--past-seasons must be at least 1.")
    pairs = sorted({
        (row.get("match_date", ""), row.get("season_label", ""))
        for row in source_rows
        if row.get("match_date") and row.get("season_label")
        and row.get("match_date", "") <= cutoff_date
        and row.get("season_label") != "Unknown Season"
    }, reverse=True)
    seasons = []
    for _, season in pairs:
        if season not in seasons:
            seasons.append(season)
        if len(seasons) == count:
            break
    if not seasons:
        raise AnalysisWorkbookError("No dated seasons are available.")
    return tuple(seasons)


def _latest_scored_match(participation: list[dict[str, Any]]) -> int:
    scored = [row for row in participation if _number(row.get("score")) is not None]
    if not scored:
        raise AnalysisWorkbookError("No scored Wilco match is available.")
    return _integer(max(scored, key=lambda row: (
        row.get("match_date", ""), _integer(row.get("match_id"))
    ))["match_id"])


def _preferred_ranking(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return min(rows, key=lambda row: (
        0 if row.get("award_scope") not in {"", "Comparison"} else 1,
        0 if "class" in row.get("rank_scope", "").casefold() else 1,
        _integer(row.get("place")) or 999999,
    ), default={})


def _place_for(rows: list[dict[str, Any]], scope: str) -> Any:
    for row in sorted(rows, key=lambda item: _integer(item.get("place")) or 999999):
        rank_scope = row.get("rank_scope", "").casefold()
        rank_type = row.get("rank_type", "").casefold()
        if scope == "all" and (
            rank_scope == "all" or "overall" in rank_type
        ):
            return row.get("place", "")
        if scope in rank_scope or scope in rank_type:
            return row.get("place", "")
    return ""


def _write_workbook(
    path: Path, cover_values: dict[str, Any],
    rows_by_filename: dict[str, list[dict[str, Any]]],
    columns_by_filename: dict[str, tuple[str, ...]],
    sheets: tuple[tuple[str, str, tuple[str, ...]], ...],
) -> bool:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False
    cover["A1"] = cover_values["workbook_name"]
    cover["A1"].font = Font(size=20, bold=True, color="1F4E78")
    cover.merge_cells("A1:D1")
    for row_number, (label, value) in enumerate(cover_values.items(), 3):
        cover.cell(row_number, 1, label.replace("_", " ").title())
        cover.cell(row_number, 2, value)
        cover.cell(row_number, 1).fill = LABEL_FILL
        cover.cell(row_number, 1).font = Font(bold=True)
        cover.cell(row_number, 2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
    cover.column_dimensions["A"].width = 25
    cover.column_dimensions["B"].width = 90
    cover.freeze_panes = "A3"
    chart_included = False
    for number, (sheet_name, filename, _) in enumerate(sheets, 1):
        sheet = workbook.create_sheet(sheet_name)
        rows = rows_by_filename[filename]
        columns = columns_by_filename[filename]
        _write_table_sheet(sheet, rows, columns, number)
        if sheet_name == "Athlete Capability Matrix":
            sheet.freeze_panes = "D2"
        if sheet_name == "Wilco vs Field by Discipline" and rows:
            _add_comparison_chart(
                sheet, columns, min(len(rows), EXCEL_MAX_ROWS - 1)
            )
            chart_included = True
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(path)
    except OSError as exc:
        raise AnalysisWorkbookError(f"Could not save workbook {path}: {exc}") \
            from exc
    return chart_included


def _write_table_sheet(
    sheet: Any, rows: list[dict[str, Any]],
    columns: tuple[str, ...], number: int,
) -> None:
    sheet.sheet_view.showGridLines = False
    visible_rows = rows[:EXCEL_MAX_ROWS - 1]
    header_row = 1
    if sheet.title == "Competitive Analysis":
        sheet["A1"] = "Competitive Analysis"
        sheet["A1"].font = Font(size=18, bold=True, color="1F4E78")
        sheet["A2"] = (
            "Best historical match score by discipline for selected match "
            "competitors."
        )
        sheet["A2"].font = Font(italic=True, color="666666")
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1,
            end_column=max(len(columns), 1)
        )
        sheet.merge_cells(
            start_row=2, start_column=1, end_row=2,
            end_column=max(len(columns), 1)
        )
        header_row = 3
    sheet.append([column.replace("_", " ").title() for column in columns])
    for cell in sheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in visible_rows:
        sheet.append([_workbook_value(column, row.get(column, ""))
                      for column in columns])
    sheet.freeze_panes = "A4" if header_row == 3 else "A2"
    last_col = get_column_letter(max(len(columns), 1))
    table_end_row = header_row + max(len(visible_rows), 1)
    table_ref = f"A{header_row}:{last_col}{table_end_row}"
    sheet.auto_filter.ref = table_ref
    if visible_rows:
        table = Table(displayName=f"AnalysisTable{number}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        note_cell = sheet.cell(header_row + 1, 1)
        note_cell.value = "No rows are currently available for this view."
        note_cell.font = Font(italic=True, color="666666")
    if len(rows) > len(visible_rows):
        note_row = sheet.max_row + 2
        sheet.cell(note_row, 1, (
            f"Excel row limit reached. This sheet shows {len(visible_rows):,} "
            f"of {len(rows):,} source rows; the complete table is available "
            "as CSV under output/analysis/tables/."
        ))
        sheet.cell(note_row, 1).font = Font(italic=True, color="9C6500")
    for column_number, column in enumerate(columns, 1):
        letter = get_column_letter(column_number)
        values = [sheet.cell(row, column_number).value
                  for row in range(header_row + 1, sheet.max_row + 1)]
        sheet.column_dimensions[letter].width = _column_width(column, values)
        if column in SECONDS_COLUMNS or any(
            marker in column for marker in (
                " Avg String", " Fastest String",
                " Avg Stage", " Fastest Stage",
            )
        ):
            for cell in sheet[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = (
                        "0.00" if column in {
                            "Best Score", "Most Recent Score", "Avg Score",
                            "Median Score",
                        } else "0.000"
                    )
        if column == "percentile":
            for cell in sheet[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0"
        if column in {"notes", "finding_message", "recommended_review",
                      "trend_note"}:
            for cell in sheet[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _add_comparison_chart(
    sheet: Any, columns: tuple[str, ...], row_count: int
) -> None:
    category_col = columns.index("discipline") + 1
    wilco_col = columns.index("Wilco Avg") + 1
    field_col = columns.index("Field Avg") + 1
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Average Match Time: Wilco vs Field"
    chart.x_axis.title = "Seconds (lower is better)"
    chart.y_axis.title = "Discipline"
    chart.height = 10
    chart.width = 22
    chart.add_data(
        Reference(sheet, min_col=wilco_col, max_col=field_col,
                  min_row=1, max_row=row_count + 1),
        titles_from_data=True,
    )
    chart.set_categories(Reference(
        sheet, min_col=category_col, min_row=2, max_row=row_count + 1
    ))
    chart.legend.position = "r"
    sheet.add_chart(chart, f"M2")


def _write_csv(
    path: Path, columns: tuple[str, ...], rows: list[dict[str, Any]]
) -> None:
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns,
                                    extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise AnalysisWorkbookError(f"Could not write {path}: {exc}") from exc


def _read_required(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise AnalysisWorkbookError(f"Required local input is missing: {path}")
    return _read_optional(path)


def _read_optional(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except (OSError, csv.Error) as exc:
        raise AnalysisWorkbookError(f"Could not read {path}: {exc}") from exc


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _display(value: float | None) -> float | str:
    return round(value, 3) if value is not None else ""


def _boolean(value: bool) -> str:
    return "true" if value else "false"


def _is_placeholder(row: dict[str, Any]) -> bool:
    return (
        not str(row.get("athlete_name", "")).strip()
        or str(row.get("athlete_id", "")).strip() == PLACEHOLDER_ID
    )


def _division(value: str) -> str:
    return value.split("/", 1)[0].strip() if value else ""


def _first(rows: list[dict[str, Any]], column: str) -> str:
    return next((str(row.get(column, "")) for row in rows
                 if row.get(column)), "")


def _normalized_name(value: str | None) -> str:
    return " ".join((value or "").casefold().split())


def _confidence(count: int) -> str:
    return "high" if count >= 4 else "medium" if count >= 2 else "low"


def _trend_note(rows: list[dict[str, Any]]) -> str:
    ordered = sorted(rows, key=lambda row: (
        row.get("match_date", ""), _integer(row.get("match_id"))
    ))
    if len(ordered) < 2:
        return "Insufficient scored history."
    first = _number(ordered[0].get("score"))
    latest = _number(ordered[-1].get("score"))
    if first is None or latest is None:
        return "Insufficient scored history."
    change = latest - first
    if abs(change) < 0.01:
        return "Stable across available scored matches."
    return (
        f"Improved by {abs(change):.3f} seconds from first to latest."
        if change < 0 else
        f"Latest is {change:.3f} seconds slower than first."
    )


def _history_sort(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("match_date", ""), _integer(row.get("match_id")),
        row.get("athlete_name", ""), row.get("discipline", ""),
    )


def _workbook_value(column: str, value: Any) -> Any:
    if value in ("true", "false"):
        return value == "true"
    if value == "":
        return None
    if column.endswith("_id") or column in {
        "match_id", "athlete_id", "team_id"
    }:
        return value
    numeric = _number(value)
    if numeric is not None and (
        column in SECONDS_COLUMNS
        or column.endswith("_count")
        or column in {"place", "field_size", "current_rank",
                      "overall_place", "gender_place", "division_place",
                      "class_place", "squad_place", "athlete_count",
                      "entry_count", "number_of_matches",
                      "scored_matches_count", "string_count",
                      "wilco_entries", "field_entries", "affected_rows",
                      "percentile", "Wilco Entries", "Field Entries",
                      "teams_count", "athletes_count", "entries_count",
                      "Number Of Matches"}
        or " Avg String" in column or " Fastest String" in column
        or " Avg Stage" in column or " Fastest Stage" in column
    ):
        return numeric
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    if column in {"notes", "finding_message", "recommended_review",
                  "trend_note"}:
        return 55
    if column in {"match_name", "source_file"}:
        return 42
    if column in {"athlete_name", "team_name", "rank_scope",
                  "award_scope", "stage_name"}:
        return 24
    longest = max(
        [len(column.replace("_", " ").title())]
        + [len(str(value)) for value in values if value is not None],
    )
    return min(max(longest + 2, 10), 28)
