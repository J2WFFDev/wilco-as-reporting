"""Create a simple Excel report from report-ready match CSV tables."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SHEET_SOURCES = (
    ("Team Summary", "team_summary.csv"),
    ("Athlete Summary", "athlete_summary.csv"),
    ("Award Results", "award_results.csv"),
    ("Squad Summary", "squad_summary.csv"),
    ("Stage Performance", "stage_performance.csv"),
    ("Coach Review Queue", "coach_review_queue.csv"),
    ("Validation Rollup", "validation_rollup.csv"),
)

INTEGER_COLUMNS = {
    "match_id",
    "athlete_count",
    "entry_count",
    "discipline_count",
    "squad_count",
    "validation_error_count",
    "validation_warning_count",
    "validation_review_count",
    "best_individual_place",
    "best_squad_place",
    "athlete_id",
    "best_place",
    "total_review_findings",
    "total_warning_findings",
    "total_error_findings",
    "place",
    "field_size",
    "squad_place",
    "stage_number",
    "dropped_string_count",
    "penalty_count",
    "finding_count",
}

SECONDS_COLUMNS = {
    "best_score_seconds",
    "score_seconds",
    "margin_to_leader",
    "margin_to_previous_place",
    "squad_score_seconds",
    "athlete_1_score",
    "athlete_2_score",
    "athlete_3_score",
    "athlete_4_score",
    "stage_score_seconds",
    "fastest_string_seconds",
    "scored_avg_string_seconds",
    "expected_value",
    "actual_value",
    "difference",
}

BOOLEAN_COLUMNS = {"inside_award_places"}

TEXT_COLUMN_WIDTHS = {
    "match_name": 42,
    "team_name": 48,
    "athlete_name": 24,
    "disciplines": 60,
    "leaderboard_name": 38,
    "rank_scope": 24,
    "squad_name": 36,
    "stage_name": 22,
    "check_name": 30,
    "finding_type": 30,
    "entity_type": 22,
    "notes": 60,
    "message": 60,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")


class MatchWorkbookError(RuntimeError):
    """Raised when a match workbook cannot be created."""


@dataclass(frozen=True)
class WorkbookResult:
    path: Path
    sheet_names: tuple[str, ...]


def build_match_workbook(
    match_id: int,
    output_dir: Path | str,
) -> WorkbookResult:
    """Create an Excel workbook from report-ready CSV files."""
    output_path = Path(output_dir)
    report_dir = output_path / "report_tables"
    workbook_dir = output_path / "workbooks"

    tables = {
        sheet_name: _read_csv(report_dir / filename)
        for sheet_name, filename in SHEET_SOURCES
    }
    team_rows = tables["Team Summary"]
    match_name = next(
        (
            row.get("match_name", "")
            for row in team_rows
            if row.get("match_name", "")
        ),
        f"Match {match_id}",
    )
    review_counts = Counter(
        row.get("severity", "")
        for row in tables["Coach Review Queue"]
    )

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    _write_cover(
        cover,
        match_id,
        match_name,
        review_counts,
    )

    for sheet_name, _ in SHEET_SOURCES:
        sheet = workbook.create_sheet(sheet_name)
        _write_table_sheet(sheet, tables[sheet_name], sheet_name)

    try:
        workbook_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = workbook_dir / f"match_{match_id}_report.xlsx"
        workbook.save(workbook_path)
    except OSError as exc:
        raise MatchWorkbookError(
            f"Could not save match workbook: {exc}"
        ) from exc

    return WorkbookResult(
        path=workbook_path,
        sheet_names=tuple(workbook.sheetnames),
    )


def _write_cover(
    sheet: Any,
    match_id: int,
    match_name: str,
    review_counts: Counter[str],
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "SASP Match Report"
    sheet["A1"].font = Font(size=18, bold=True, color="1F4E78")
    sheet.merge_cells("A1:D1")

    rows = (
        ("Match ID", match_id),
        ("Match Name", match_name),
        (
            "Generated At",
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        ),
        (
            "Pipeline Steps",
            "Fetch → Parse → Validate → Report → Workbook",
        ),
        ("Validation Errors", review_counts.get("ERROR", 0)),
        ("Validation Warnings", review_counts.get("WARNING", 0)),
        ("Validation Reviews", review_counts.get("REVIEW", 0)),
        (
            "Artifact Note",
            "Generated workbook artifact; raw JSON and CSV tables remain "
            "the source pipeline data.",
        ),
    )
    for row_number, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value)
        sheet.cell(row=row_number, column=1).font = Font(bold=True)
        sheet.cell(row=row_number, column=1).fill = TITLE_FILL
        sheet.cell(row=row_number, column=2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 70
    sheet.freeze_panes = "A3"


def _write_table_sheet(
    sheet: Any,
    rows: list[dict[str, str]],
    sheet_name: str,
) -> None:
    if not rows:
        sheet["A1"] = "No rows available"
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in rows:
        sheet.append(
            [
                _cell_value(header, row.get(header, ""))
                for header in headers
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    table = Table(
        displayName=_table_name(sheet_name),
        ref=sheet.dimensions,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for column_number, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_number)
        if header in SECONDS_COLUMNS:
            for cell in sheet[column_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"
        width = _column_width(
            header,
            (
                sheet.cell(row=row_number, column=column_number).value
                for row_number in range(2, sheet.max_row + 1)
            ),
        )
        sheet.column_dimensions[column_letter].width = width


def _cell_value(header: str, value: str) -> Any:
    if value == "":
        return None
    if header in INTEGER_COLUMNS:
        try:
            return int(value)
        except ValueError:
            return value
    if header in SECONDS_COLUMNS:
        try:
            return float(value)
        except ValueError:
            return value
    if header in BOOLEAN_COLUMNS:
        normalized = value.casefold()
        if normalized in {"true", "false"}:
            return normalized == "true"
    return value


def _column_width(
    header: str,
    values: Any,
) -> float:
    if header in TEXT_COLUMN_WIDTHS:
        return TEXT_COLUMN_WIDTHS[header]
    longest = len(header)
    for value in values:
        if value is not None:
            longest = max(longest, len(str(value)))
    return min(max(longest + 2, 10), 32)


def _table_name(sheet_name: str) -> str:
    return "".join(part for part in sheet_name.split()) + "Table"


def _read_csv(path: Path) -> list[dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as source:
            return list(csv.DictReader(source))
    except FileNotFoundError as exc:
        raise MatchWorkbookError(
            f"Missing report table: {path}"
        ) from exc
    except (OSError, csv.Error) as exc:
        raise MatchWorkbookError(
            f"Could not read report table {path}: {exc}"
        ) from exc
