"""Create a coach-readable workbook from team report tables."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from wilco_as_reporting.reports.team_report import (
    ATHLETE_SUMMARY_COLUMNS,
    AWARD_COLUMNS,
    COMPARISON_COLUMNS,
    REVIEW_COLUMNS,
    SQUAD_COLUMNS,
    STAGE_COLUMNS,
    TEAM_SUMMARY_COLUMNS,
)
from wilco_as_reporting.team_profiles import TeamProfile
from wilco_as_reporting.workbooks.excel_writer import (
    BOOLEAN_COLUMNS,
    HEADER_FILL,
    HEADER_FONT,
    INTEGER_COLUMNS,
    SECONDS_COLUMNS,
    TITLE_FILL,
)

TEAM_SHEETS = (
    (
        "Wilco Summary",
        "wilco_team_summary.csv",
        TEAM_SUMMARY_COLUMNS,
    ),
    (
        "Wilco Athletes",
        "wilco_athlete_summary.csv",
        ATHLETE_SUMMARY_COLUMNS,
    ),
    (
        "Award Highlights",
        "wilco_award_highlights.csv",
        AWARD_COLUMNS,
    ),
    (
        "Comparison Results",
        "wilco_comparison_results.csv",
        COMPARISON_COLUMNS,
    ),
    (
        "Squad Summary",
        "wilco_squad_summary.csv",
        SQUAD_COLUMNS,
    ),
    (
        "Stage Coach View",
        "wilco_stage_coach_view.csv",
        STAGE_COLUMNS,
    ),
    (
        "Coach Review Queue",
        "wilco_coach_review_queue.csv",
        REVIEW_COLUMNS,
    ),
)

TEXT_WIDTHS = {
    "match_name": 42,
    "team_name": 32,
    "athlete_name": 24,
    "disciplines": 58,
    "leaderboard_name": 38,
    "rank_scope": 24,
    "squad_name": 32,
    "stage_name": 22,
    "coach_flag": 42,
    "coach_note": 62,
    "message": 55,
    "notes": 62,
    "check_name": 30,
    "finding_type": 30,
    "entity_type": 22,
}

ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
REVIEW_FILL = PatternFill("solid", fgColor="FCE5CD")
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")


class TeamWorkbookError(RuntimeError):
    """Raised when a team workbook cannot be created."""


@dataclass(frozen=True)
class TeamWorkbookResult:
    path: Path
    sheet_names: tuple[str, ...]


def build_team_workbook(
    match_id: int,
    output_dir: Path | str,
    profile: TeamProfile,
) -> TeamWorkbookResult:
    """Create a coach-readable team workbook."""
    output_path = Path(output_dir)
    team_dir = (
        output_path
        / "team_report_tables"
        / profile.team_key
    )
    workbook_dir = output_path / "workbooks"
    tables = {
        sheet_name: _read_csv(team_dir / filename)
        for sheet_name, filename, _ in TEAM_SHEETS
    }
    summary = (
        tables["Wilco Summary"][0]
        if tables["Wilco Summary"]
        else {}
    )
    match_name = summary.get("match_name") or f"Match {match_id}"
    counts = Counter(
        row.get("severity", "")
        for row in tables["Coach Review Queue"]
    )

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    _write_cover(
        cover,
        match_id,
        match_name,
        profile,
        counts,
    )
    for sheet_name, _, columns in TEAM_SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        _write_team_sheet(
            sheet,
            tables[sheet_name],
            columns,
            sheet_name,
        )

    try:
        workbook_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = (
            workbook_dir
            / f"match_{match_id}_{profile.team_key}_report.xlsx"
        )
        workbook.save(workbook_path)
    except OSError as exc:
        raise TeamWorkbookError(
            f"Could not save team workbook: {exc}"
        ) from exc
    return TeamWorkbookResult(
        path=workbook_path,
        sheet_names=tuple(workbook.sheetnames),
    )


def _write_cover(
    sheet: Any,
    match_id: int,
    match_name: str,
    profile: TeamProfile,
    counts: Counter[str],
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = f"{profile.team_name} Coach Report"
    sheet["A1"].font = Font(
        size=18,
        bold=True,
        color="1F4E78",
    )
    sheet.merge_cells("A1:D1")
    rows = (
        ("Match ID", match_id),
        ("Match Name", match_name),
        ("Team Key", profile.team_key),
        ("Team Name", profile.team_name),
        (
            "Generated At",
            datetime.now(timezone.utc)
            .replace(microsecond=0)
            .isoformat(),
        ),
        (
            "Pipeline Steps",
            "Fetch → Parse → Validate → Full Report → "
            "Team Report → Team Workbook",
        ),
        ("Validation Errors", counts.get("ERROR", 0)),
        ("Validation Warnings", counts.get("WARNING", 0)),
        ("Validation Reviews", counts.get("REVIEW", 0)),
        (
            "Artifact Note",
            "Generated coaching artifact; do not treat the workbook "
            "as source data.",
        ),
        (
            "Source Data Note",
            "Raw JSON, parsed CSVs, validation outputs, and report "
            "tables remain the auditable source pipeline.",
        ),
    )
    for row_number, (label, value) in enumerate(rows, start=3):
        label_cell = sheet.cell(
            row=row_number,
            column=1,
            value=label,
        )
        value_cell = sheet.cell(
            row=row_number,
            column=2,
            value=value,
        )
        label_cell.font = Font(bold=True)
        label_cell.fill = TITLE_FILL
        value_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 76
    sheet.freeze_panes = "A3"


def _write_team_sheet(
    sheet: Any,
    rows: list[dict[str, str]],
    columns: Iterable[str],
    sheet_name: str,
) -> None:
    headers = list(columns)
    sheet.append([_human_label(header) for header in headers])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row in rows:
        sheet.append(
            [
                _cell_value(header, row.get(header, ""))
                for header in headers
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    if rows:
        table = Table(
            displayName=_table_name(sheet_name),
            ref=sheet.dimensions,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet["A2"] = "No report rows are currently available."
        sheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=max(len(headers), 2),
        )
        sheet["A2"].font = Font(italic=True, color="666666")

    for column_number, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_number)
        if header in SECONDS_COLUMNS:
            for cell in sheet[column_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"
        if header in {"coach_note", "notes", "message"}:
            for cell in sheet[column_letter][1:]:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )
        sheet.column_dimensions[column_letter].width = (
            TEXT_WIDTHS.get(
                header,
                _column_width(
                    header,
                    (
                        sheet.cell(
                            row=row_number,
                            column=column_number,
                        ).value
                        for row_number in range(
                            2,
                            sheet.max_row + 1,
                        )
                    ),
                ),
            )
        )
    _apply_coach_highlights(sheet, headers)


def _apply_coach_highlights(
    sheet: Any,
    headers: list[str],
) -> None:
    header_positions = {
        header: index + 1
        for index, header in enumerate(headers)
    }
    severity_column = header_positions.get("severity")
    flag_column = header_positions.get("coach_flag")
    for row_number in range(2, sheet.max_row + 1):
        if severity_column:
            severity = sheet.cell(
                row=row_number,
                column=severity_column,
            ).value
            fill = {
                "ERROR": ERROR_FILL,
                "WARNING": WARNING_FILL,
                "REVIEW": REVIEW_FILL,
            }.get(severity)
            if fill:
                sheet.cell(
                    row=row_number,
                    column=severity_column,
                ).fill = fill
        if flag_column:
            flag_cell = sheet.cell(
                row=row_number,
                column=flag_column,
            )
            if flag_cell.value:
                flag_cell.fill = FLAG_FILL


def _human_label(header: str) -> str:
    labels = {
        "match_id": "Match ID",
        "team_key": "Team Key",
        "athlete_id": "Athlete ID",
        "coach_flag": "Coach Flag",
        "coach_note": "Coach Note",
        "inside_award_places": "Inside Award Places",
    }
    return labels.get(
        header,
        header.replace("_", " ").title(),
    )


def _cell_value(header: str, value: str) -> Any:
    if value == "":
        return None
    if header in INTEGER_COLUMNS or header.endswith("_count"):
        try:
            return int(value)
        except ValueError:
            return value
    if header in SECONDS_COLUMNS:
        try:
            return float(value)
        except ValueError:
            return value
    if header in BOOLEAN_COLUMNS:
        normalized = value.casefold()
        if normalized in {"true", "false"}:
            return normalized == "true"
    return value


def _column_width(
    header: str,
    values: Iterable[Any],
) -> float:
    longest = len(_human_label(header))
    for value in values:
        if value is not None:
            longest = max(longest, len(str(value)))
    return min(max(longest + 2, 10), 32)


def _table_name(sheet_name: str) -> str:
    return "".join(sheet_name.split()) + "TeamTable"


def _read_csv(path: Path) -> list[dict[str, str]]:
    try:
        with path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as source:
            return list(csv.DictReader(source))
    except FileNotFoundError as exc:
        raise TeamWorkbookError(
            f"Missing team report table: {path}"
        ) from exc
    except (OSError, csv.Error) as exc:
        raise TeamWorkbookError(
            f"Could not read team report table {path}: {exc}"
        ) from exc
