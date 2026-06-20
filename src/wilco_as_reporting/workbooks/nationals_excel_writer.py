"""Create the coach-readable Nationals operations workbook."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from wilco_as_reporting.nationals_ops import (
    ATHLETE_CHANGE_COLUMNS,
    AWARD_CHANGE_COLUMNS,
    CHANGE_SUMMARY_COLUMNS,
    DAILY_BRIEF_COLUMNS,
    REVIEW_CHANGE_COLUMNS,
    SQUAD_CHANGE_COLUMNS,
    STAGE_CHANGE_COLUMNS,
)
from wilco_as_reporting.reports.team_report import TEAM_SUMMARY_COLUMNS
from wilco_as_reporting.team_profiles import TeamProfile
from wilco_as_reporting.workbooks.excel_writer import (
    HEADER_FILL,
    HEADER_FONT,
    TITLE_FILL,
)

SHEETS = (
    (
        "Daily Brief",
        "wilco_daily_brief.csv",
        DAILY_BRIEF_COLUMNS,
    ),
    (
        "Change Summary",
        "wilco_change_summary.csv",
        CHANGE_SUMMARY_COLUMNS,
    ),
    (
        "Athlete Changes",
        "wilco_athlete_changes.csv",
        ATHLETE_CHANGE_COLUMNS,
    ),
    (
        "Award Changes",
        "wilco_award_changes.csv",
        AWARD_CHANGE_COLUMNS,
    ),
    (
        "Squad Changes",
        "wilco_squad_changes.csv",
        SQUAD_CHANGE_COLUMNS,
    ),
    (
        "Stage Changes",
        "wilco_stage_changes.csv",
        STAGE_CHANGE_COLUMNS,
    ),
    (
        "Review Changes",
        "wilco_review_changes.csv",
        REVIEW_CHANGE_COLUMNS,
    ),
)

SECONDS_COLUMNS = {
    "previous_score_seconds",
    "current_score_seconds",
    "score_change_seconds",
    "previous_squad_score_seconds",
    "current_squad_score_seconds",
    "previous_stage_score_seconds",
    "current_stage_score_seconds",
    "stage_score_change_seconds",
}

INTEGER_COLUMNS = {
    "match_id",
    "athlete_count_current",
    "athlete_count_previous",
    "changed_athlete_count",
    "changed_award_count",
    "changed_squad_count",
    "new_review_count",
    "resolved_review_count",
    "validation_error_count",
    "validation_warning_count",
    "validation_review_count",
    "previous_best_place",
    "current_best_place",
    "place_change",
    "previous_place",
    "current_place",
    "previous_squad_place",
    "current_squad_place",
    "stage_number",
}

TEXT_WIDTHS = {
    "current_snapshot": 55,
    "previous_snapshot": 55,
    "notes": 58,
    "detail": 72,
    "title": 42,
    "message": 65,
    "athlete_name": 24,
    "related_athlete": 24,
    "discipline": 26,
    "related_discipline": 26,
    "squad_name": 34,
    "related_squad": 34,
    "leaderboard_name": 38,
    "rank_scope": 26,
    "award_scope": 24,
    "check_name": 30,
    "finding_type": 30,
    "stage_name": 22,
}

PRIORITY_FILLS = {
    "HIGH": PatternFill("solid", fgColor="F4CCCC"),
    "MEDIUM": PatternFill("solid", fgColor="FFF2CC"),
    "LOW": PatternFill("solid", fgColor="D9EAD3"),
    "INFO": PatternFill("solid", fgColor="D9EAF7"),
}


class NationalsWorkbookError(RuntimeError):
    """Raised when the Nationals operations workbook cannot be created."""


@dataclass(frozen=True)
class NationalsWorkbookResult:
    path: Path
    sheet_names: tuple[str, ...]


def build_nationals_workbook(
    match_id: int,
    output_dir: Path | str,
    profile: TeamProfile,
    *,
    generated_at: str,
    snapshot_label: str,
    current_snapshot: Path,
    previous_snapshot: Path | None,
    comparison_status: str,
    data_status: str,
    validation_counts: dict[str, int],
    notes: tuple[str, ...],
) -> NationalsWorkbookResult:
    """Build the current Nationals operations workbook."""
    output_path = Path(output_dir)
    ops_dir = output_path / "nationals_ops" / profile.team_key
    team_dir = (
        output_path / "team_report_tables" / profile.team_key
    )
    tables = {
        sheet_name: _read_csv(ops_dir / filename)
        for sheet_name, filename, _ in SHEETS
    }
    current_summary = _read_csv(
        team_dir / "wilco_team_summary.csv"
    )
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    _write_cover(
        cover,
        match_id,
        profile,
        generated_at,
        snapshot_label,
        current_snapshot,
        previous_snapshot,
        comparison_status,
        data_status,
        validation_counts,
        notes,
    )
    for sheet_name, _, columns in SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        _write_sheet(sheet, tables[sheet_name], columns, sheet_name)
    summary_sheet = workbook.create_sheet("Current Wilco Summary")
    _write_sheet(
        summary_sheet,
        current_summary,
        TEAM_SUMMARY_COLUMNS,
        "Current Wilco Summary",
    )

    workbook_dir = output_path / "workbooks"
    workbook_path = (
        workbook_dir
        / f"match_{match_id}_{profile.team_key}_nationals_ops.xlsx"
    )
    try:
        workbook_dir.mkdir(parents=True, exist_ok=True)
        workbook.save(workbook_path)
    except OSError as exc:
        raise NationalsWorkbookError(
            f"Could not save Nationals operations workbook: {exc}"
        ) from exc
    return NationalsWorkbookResult(
        path=workbook_path,
        sheet_names=tuple(workbook.sheetnames),
    )


def _write_cover(
    sheet: Any,
    match_id: int,
    profile: TeamProfile,
    generated_at: str,
    snapshot_label: str,
    current_snapshot: Path,
    previous_snapshot: Path | None,
    comparison_status: str,
    data_status: str,
    validation_counts: dict[str, int],
    notes: tuple[str, ...],
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = f"{profile.team_name} Nationals Operations"
    sheet["A1"].font = Font(size=18, bold=True, color="1F4E78")
    sheet.merge_cells("A1:D1")
    incomplete_note = (
        " ".join(notes)
        if data_status != "complete"
        else "Current team data appears complete."
    )
    rows = (
        ("Match ID", match_id),
        ("Team Key", profile.team_key),
        ("Generated At", generated_at),
        ("Snapshot Label", snapshot_label),
        ("Current Snapshot", str(current_snapshot)),
        (
            "Previous Snapshot",
            str(previous_snapshot) if previous_snapshot else "Baseline",
        ),
        ("Comparison Status", comparison_status),
        ("Data Status", data_status),
        ("Validation Errors", validation_counts.get("ERROR", 0)),
        (
            "Validation Warnings",
            validation_counts.get("WARNING", 0),
        ),
        ("Validation Reviews", validation_counts.get("REVIEW", 0)),
        (
            "Artifact Note",
            "Generated operations artifact. Raw JSON, CSV tables, "
            "manifest history, and snapshots remain the audit trail.",
        ),
        ("Incomplete Data Note", incomplete_note),
    )
    for row_number, (label, value) in enumerate(rows, start=3):
        label_cell = sheet.cell(row=row_number, column=1, value=label)
        value_cell = sheet.cell(row=row_number, column=2, value=value)
        label_cell.font = Font(bold=True)
        label_cell.fill = TITLE_FILL
        value_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 82
    sheet.freeze_panes = "A3"


def _write_sheet(
    sheet: Any,
    rows: list[dict[str, str]],
    columns: Iterable[str],
    sheet_name: str,
) -> None:
    headers = list(columns)
    sheet.append([_human_label(header) for header in headers])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row in rows:
        sheet.append(
            [_cell_value(header, row.get(header, "")) for header in headers]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    if rows:
        table = Table(
            displayName=_table_name(sheet_name),
            ref=sheet.dimensions,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet["A2"] = "No report rows are currently available."
        sheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=max(len(headers), 2),
        )
        sheet["A2"].font = Font(italic=True, color="666666")
    for column_number, header in enumerate(headers, start=1):
        letter = get_column_letter(column_number)
        if header in SECONDS_COLUMNS:
            for cell in sheet[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"
        if header in {"detail", "message", "notes"}:
            for cell in sheet[letter][1:]:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )
        sheet.column_dimensions[letter].width = TEXT_WIDTHS.get(
            header,
            _column_width(
                header,
                (
                    sheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                    for row_number in range(2, sheet.max_row + 1)
                ),
            ),
        )
    _apply_priority_fills(sheet, headers)


def _apply_priority_fills(sheet: Any, headers: list[str]) -> None:
    if "priority" not in headers:
        return
    column = headers.index("priority") + 1
    for row_number in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_number, column=column)
        fill = PRIORITY_FILLS.get(str(cell.value))
        if fill:
            cell.fill = fill


def _cell_value(header: str, value: str) -> Any:
    if value == "":
        return None
    if header in INTEGER_COLUMNS or header.endswith("_count"):
        try:
            return int(float(value))
        except ValueError:
            return value
    if header in SECONDS_COLUMNS:
        try:
            return float(value)
        except ValueError:
            return value
    return value


def _human_label(header: str) -> str:
    return {
        "match_id": "Match ID",
        "team_key": "Team Key",
        "athlete_id": "Athlete ID",
    }.get(header, header.replace("_", " ").title())


def _column_width(header: str, values: Iterable[Any]) -> float:
    longest = len(_human_label(header))
    for value in values:
        if value is not None:
            longest = max(longest, len(str(value)))
    return min(max(longest + 2, 10), 32)


def _table_name(sheet_name: str) -> str:
    return "".join(sheet_name.split()) + "OpsTable"


def _read_csv(path: Path) -> list[dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as source:
            return list(csv.DictReader(source))
    except FileNotFoundError as exc:
        raise NationalsWorkbookError(
            f"Missing Nationals workbook input: {path}"
        ) from exc
    except (OSError, csv.Error) as exc:
        raise NationalsWorkbookError(
            f"Could not read Nationals workbook input {path}: {exc}"
        ) from exc
