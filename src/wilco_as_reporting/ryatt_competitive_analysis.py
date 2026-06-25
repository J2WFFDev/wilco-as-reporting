"""Build Ryatt West's competitive scouting report from local score tables."""

from __future__ import annotations

import csv
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PLACEHOLDER_ID = "9999"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LABEL_FILL = PatternFill("solid", fgColor="D9EAF7")
NEGATIVE_GAP_FILL = PatternFill("solid", fgColor="C6EFCE")
EXCEL_MAX_ROWS = 1_048_576

ANALYSIS_COLUMNS = (
    "Discipline",
    "Ryatt Best Score",
    "Athlete Name",
    "Athlete ID",
    "Team",
    "Gender",
    "Class",
    "Division",
    "Best Score",
    "Gap To Ryatt",
    "Comparison Group",
    "Match With Best Score",
    "Best Score Date",
    "Most Recent Score",
    "Most Recent Match",
    "Most Recent Date",
    "Avg Score",
    "Median Score",
    "Number Of Matches",
    "Notes",
)

SUMMARY_COLUMNS = ("metric", "value", "notes")

DISCIPLINE_COLUMNS = (
    "Discipline",
    "Ryatt Best Score",
    "Faster Athletes Count",
    "Within Gap Athletes Count",
    "Closest Faster Athlete",
    "Closest Faster Gap",
    "Closest Slower Athlete",
    "Closest Slower Gap",
    "Competitive Note",
)

QUALITY_COLUMNS = ("issue_type", "affected_rows", "severity", "notes")

SECONDS_COLUMNS = {
    "Ryatt Best Score",
    "Best Score",
    "Gap To Ryatt",
    "Most Recent Score",
    "Avg Score",
    "Median Score",
    "Closest Faster Gap",
    "Closest Slower Gap",
}


class RyattCompetitiveAnalysisError(RuntimeError):
    """Raised when the Ryatt competitive analysis cannot be built."""


@dataclass(frozen=True)
class RyattCompetitiveAnalysisResult:
    output_dir: Path
    analysis_csv: Path
    summary_csv: Path
    discipline_summary_csv: Path
    data_quality_csv: Path
    workbook_path: Path | None
    target_athlete_name: str
    target_athlete_id: str
    gap_seconds: float
    historical_score_rows_scanned: int
    target_disciplines_found: int
    comparison_rows_generated: int
    faster_than_ryatt_count: int
    within_gap_slower_count: int
    placeholder_rows_excluded: int
    missing_score_rows_skipped: int
    non_wilco_rows_included: int


def build_ryatt_competitive_analysis(
    *,
    output_root: Path | str = Path("output"),
    target_athlete_id: int | str = "76179",
    target_athlete_name: str = "Ryatt West",
    gap_seconds: float = 2.0,
    match_id: int | None = 671,
    history_dir: Path | str | None = None,
    records_dir: Path | str | None = None,
    workbook: bool = True,
) -> RyattCompetitiveAnalysisResult:
    """Build Ryatt's competitive scouting report from local CSV outputs."""
    del records_dir  # Reserved for future local-only context; no record logic.
    if gap_seconds < 0:
        raise RyattCompetitiveAnalysisError("--gap-seconds must be non-negative.")

    root = Path(output_root)
    history = Path(history_dir) if history_dir else root / "history"
    output_dir = root / "special_reports" / "ryatt_competitive_analysis"
    output_dir.mkdir(parents=True, exist_ok=True)

    match_dates = _load_match_dates(history)
    scan = _load_local_scores(root, match_dates)
    target_id = str(target_athlete_id).strip()
    target_name = target_athlete_name.strip()

    target_rows = [
        row for row in scan.rows
        if _same_athlete(row, target_id, target_name)
    ]
    target_best_by_discipline = _best_by_discipline(target_rows)
    analysis_rows = _analysis_rows(
        rows=scan.rows,
        target_best_by_discipline=target_best_by_discipline,
        target_athlete_id=target_id,
        target_athlete_name=target_name,
        gap_seconds=gap_seconds,
    )
    discipline_summary = _discipline_summary_rows(
        analysis_rows, gap_seconds
    )
    faster_count = sum(
        row["Comparison Group"] == "Faster Than Ryatt"
        for row in analysis_rows
    )
    within_count = sum(
        row["Comparison Group"] in {
            "Within 2 Seconds Slower",
            "Within Gap Slower",
        }
        for row in analysis_rows
    )
    non_wilco_rows = sum(
        bool(row.get("Team"))
        and row.get("Team", "").casefold() != "wilco shooting sports"
        for row in analysis_rows
    )
    summary_rows = _summary_rows(
        target_athlete_name=target_name,
        target_athlete_id=target_id,
        gap_seconds=gap_seconds,
        historical_score_rows_scanned=len(scan.rows),
        target_disciplines_found=len(target_best_by_discipline),
        comparison_rows_generated=len(analysis_rows),
        faster_than_ryatt_count=faster_count,
        within_gap_slower_count=within_count,
    )
    quality_rows = _quality_rows(
        target_found=bool(target_rows),
        target_disciplines_found=len(target_best_by_discipline),
        historical_rows_scanned=len(scan.rows),
        placeholder_rows_excluded=scan.placeholder_rows_excluded,
        missing_score_rows_skipped=scan.missing_score_rows_skipped,
        unresolved_disciplines=[],
        match_id=match_id,
    )

    analysis_csv = output_dir / "ryatt_competitive_analysis.csv"
    summary_csv = output_dir / "ryatt_competitive_summary.csv"
    discipline_csv = output_dir / "ryatt_discipline_summary.csv"
    quality_csv = output_dir / "data_quality_notes.csv"
    _write_csv(analysis_csv, ANALYSIS_COLUMNS, analysis_rows)
    _write_csv(summary_csv, SUMMARY_COLUMNS, summary_rows)
    _write_csv(discipline_csv, DISCIPLINE_COLUMNS, discipline_summary)
    _write_csv(quality_csv, QUALITY_COLUMNS, quality_rows)

    workbook_path = output_dir / "ryatt_competitive_analysis.xlsx" \
        if workbook else None
    if workbook_path:
        _write_workbook(
            workbook_path,
            summary_rows,
            analysis_rows,
            discipline_summary,
            quality_rows,
            target_name,
            gap_seconds,
        )

    return RyattCompetitiveAnalysisResult(
        output_dir=output_dir,
        analysis_csv=analysis_csv,
        summary_csv=summary_csv,
        discipline_summary_csv=discipline_csv,
        data_quality_csv=quality_csv,
        workbook_path=workbook_path,
        target_athlete_name=target_name,
        target_athlete_id=target_id,
        gap_seconds=gap_seconds,
        historical_score_rows_scanned=len(scan.rows),
        target_disciplines_found=len(target_best_by_discipline),
        comparison_rows_generated=len(analysis_rows),
        faster_than_ryatt_count=faster_count,
        within_gap_slower_count=within_count,
        placeholder_rows_excluded=scan.placeholder_rows_excluded,
        missing_score_rows_skipped=scan.missing_score_rows_skipped,
        non_wilco_rows_included=non_wilco_rows,
    )


@dataclass(frozen=True)
class _ScoreScan:
    rows: list[dict[str, Any]]
    placeholder_rows_excluded: int
    missing_score_rows_skipped: int


def _load_local_scores(
    root: Path,
    match_dates: dict[str, str],
) -> _ScoreScan:
    rows = []
    placeholders = 0
    missing_scores = 0
    for path in sorted(root.rglob("tables/match_scores.csv")):
        match_id = path.parent.parent.name
        if not match_id.isdigit():
            continue
        for row in _read_optional(path):
            if _is_placeholder(row):
                placeholders += 1
                continue
            score = _number(row.get("match_score_seconds"))
            if score is None:
                missing_scores += 1
                continue
            rows.append({
                "match_id": row.get("match_id") or match_id,
                "match_name": row.get("match_name", ""),
                "match_date": match_dates.get(match_id, ""),
                "athlete_id": row.get("athlete_id", "").strip(),
                "athlete_name": row.get("athlete_name", "").strip(),
                "team_name": row.get("team_name", "").strip(),
                "gender": row.get("gender", "").strip(),
                "class": row.get("class", "").strip(),
                "division": _division(row.get("class", "")),
                "discipline": row.get("discipline", "").strip(),
                "score": score,
                "name_key": _normalized_name(row.get("athlete_name", "")),
            })
    return _ScoreScan(rows, placeholders, missing_scores)


def _load_match_dates(history_dir: Path) -> dict[str, str]:
    path = history_dir / "history_source_matches.csv"
    result = {}
    for row in _read_optional(path):
        match_id = row.get("match_id", "").strip()
        if match_id:
            result[match_id] = row.get("match_date", "")
    return result


def _best_by_discipline(
    rows: list[dict[str, Any]]
) -> dict[str, dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("discipline"):
            grouped[row["discipline"]].append(row)
    return {
        discipline: min(values, key=lambda row: row["score"])
        for discipline, values in grouped.items()
    }


def _analysis_rows(
    *,
    rows: list[dict[str, Any]],
    target_best_by_discipline: dict[str, dict[str, Any]],
    target_athlete_id: str,
    target_athlete_name: str,
    gap_seconds: float,
) -> list[dict[str, Any]]:
    results = []
    group_rows: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        identity = row.get("athlete_id") or row.get("name_key")
        if identity and row.get("discipline"):
            group_rows[(row["discipline"], identity)].append(row)

    for discipline, target_best in target_best_by_discipline.items():
        ryatt_best = target_best["score"]
        for (_, identity), athlete_rows in group_rows.items():
            if athlete_rows[0].get("discipline") != discipline:
                continue
            best = min(athlete_rows, key=lambda row: row["score"])
            gap = best["score"] - ryatt_best
            is_target = _same_athlete(
                best, target_athlete_id, target_athlete_name
            )
            if is_target:
                group = "Ryatt West"
            elif gap < 0:
                group = "Faster Than Ryatt"
            elif gap <= gap_seconds:
                group = (
                    "Within 2 Seconds Slower"
                    if abs(gap_seconds - 2.0) < 0.000001
                    else "Within Gap Slower"
                )
            else:
                continue
            results.append(_result_row(
                discipline, ryatt_best, athlete_rows, best, gap, group
            ))

    return sorted(results, key=lambda row: (
        row["Discipline"],
        _number(row["Gap To Ryatt"]) or 0,
        _number(row["Best Score"]) or 999999,
        row["Athlete Name"],
    ))


def _result_row(
    discipline: str,
    ryatt_best: float,
    athlete_rows: list[dict[str, Any]],
    best: dict[str, Any],
    gap: float,
    comparison_group: str,
) -> dict[str, Any]:
    scores = [row["score"] for row in athlete_rows]
    recent = max(
        athlete_rows,
        key=lambda row: (row.get("match_date", ""), _integer(row["match_id"])),
    )
    return {
        "Discipline": discipline,
        "Ryatt Best Score": _display(ryatt_best),
        "Athlete Name": best.get("athlete_name", ""),
        "Athlete ID": best.get("athlete_id", ""),
        "Team": best.get("team_name", ""),
        "Gender": best.get("gender", ""),
        "Class": best.get("class", ""),
        "Division": best.get("division", ""),
        "Best Score": _display(best["score"]),
        "Gap To Ryatt": _display(gap),
        "Comparison Group": comparison_group,
        "Match With Best Score": best.get("match_name", ""),
        "Best Score Date": best.get("match_date", ""),
        "Most Recent Score": _display(recent["score"]),
        "Most Recent Match": recent.get("match_name", ""),
        "Most Recent Date": recent.get("match_date", ""),
        "Avg Score": _display(statistics.mean(scores)),
        "Median Score": _display(statistics.median(scores)),
        "Number Of Matches": len({row["match_id"] for row in athlete_rows}),
        "Notes": (
            "Negative gap means faster than Ryatt; positive gap means slower."
        ),
    }


def _discipline_summary_rows(
    analysis_rows: list[dict[str, Any]],
    gap_seconds: float,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in analysis_rows:
        grouped[row["Discipline"]].append(row)
    results = []
    for discipline, rows in sorted(grouped.items()):
        faster = [
            row for row in rows
            if row["Comparison Group"] == "Faster Than Ryatt"
        ]
        slower = [
            row for row in rows
            if row["Comparison Group"] in {
                "Within 2 Seconds Slower",
                "Within Gap Slower",
            }
        ]
        ryatt = next(
            (row for row in rows if row["Comparison Group"] == "Ryatt West"),
            {},
        )
        closest_faster = max(
            faster, key=lambda row: _number(row["Gap To Ryatt"]) or -999999,
            default={},
        )
        closest_slower = min(
            slower, key=lambda row: _number(row["Gap To Ryatt"]) or 999999,
            default={},
        )
        results.append({
            "Discipline": discipline,
            "Ryatt Best Score": ryatt.get("Ryatt Best Score", ""),
            "Faster Athletes Count": len(faster),
            "Within Gap Athletes Count": len(slower),
            "Closest Faster Athlete": closest_faster.get("Athlete Name", ""),
            "Closest Faster Gap": closest_faster.get("Gap To Ryatt", ""),
            "Closest Slower Athlete": closest_slower.get("Athlete Name", ""),
            "Closest Slower Gap": closest_slower.get("Gap To Ryatt", ""),
            "Competitive Note": (
                f"Includes faster athletes and athletes within "
                f"{gap_seconds:.2f} seconds slower than Ryatt."
            ),
        })
    return results


def _summary_rows(
    *,
    target_athlete_name: str,
    target_athlete_id: str,
    gap_seconds: float,
    historical_score_rows_scanned: int,
    target_disciplines_found: int,
    comparison_rows_generated: int,
    faster_than_ryatt_count: int,
    within_gap_slower_count: int,
) -> list[dict[str, Any]]:
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    return [
        {
            "metric": "target_athlete_name",
            "value": target_athlete_name,
            "notes": "Scouting report target.",
        },
        {
            "metric": "target_athlete_id",
            "value": target_athlete_id,
            "notes": "Matched by athlete ID first.",
        },
        {
            "metric": "gap_seconds",
            "value": _display(gap_seconds),
            "notes": "Slower-than-Ryatt inclusion window.",
        },
        {
            "metric": "historical_score_rows_scanned",
            "value": historical_score_rows_scanned,
            "notes": "Rows from local processed match_scores.csv files.",
        },
        {
            "metric": "target_disciplines_found",
            "value": target_disciplines_found,
            "notes": "Disciplines where Ryatt has local historical scores.",
        },
        {
            "metric": "comparison_rows_generated",
            "value": comparison_rows_generated,
            "notes": "Rows in ryatt_competitive_analysis.csv.",
        },
        {
            "metric": "faster_than_ryatt_count",
            "value": faster_than_ryatt_count,
            "notes": "Competitor rows with negative gap.",
        },
        {
            "metric": "within_gap_slower_count",
            "value": within_gap_slower_count,
            "notes": "Competitor rows slower than Ryatt but inside gap.",
        },
        {
            "metric": "generated_at",
            "value": generated_at,
            "notes": "UTC timestamp.",
        },
    ]


def _quality_rows(
    *,
    target_found: bool,
    target_disciplines_found: int,
    historical_rows_scanned: int,
    placeholder_rows_excluded: int,
    missing_score_rows_skipped: int,
    unresolved_disciplines: list[str],
    match_id: int | None,
) -> list[dict[str, Any]]:
    return [
        {
            "issue_type": "target_athlete_found",
            "affected_rows": 1 if target_found else 0,
            "severity": "INFO" if target_found else "WARNING",
            "notes": (
                "Target athlete was found in local score history."
                if target_found else
                "Target athlete was not found in local score history."
            ),
        },
        {
            "issue_type": "target_disciplines_found",
            "affected_rows": target_disciplines_found,
            "severity": "INFO",
            "notes": "Disciplines with a local Ryatt score.",
        },
        {
            "issue_type": "historical_rows_scanned",
            "affected_rows": historical_rows_scanned,
            "severity": "INFO",
            "notes": "Local processed match score rows scanned.",
        },
        {
            "issue_type": "placeholder_rows_excluded",
            "affected_rows": placeholder_rows_excluded,
            "severity": "INFO",
            "notes": "Blank athlete names and athlete ID 9999 are excluded.",
        },
        {
            "issue_type": "missing_score_data_skipped",
            "affected_rows": missing_score_rows_skipped,
            "severity": "INFO",
            "notes": "Rows without numeric match scores are skipped.",
        },
        {
            "issue_type": "selected_match_context",
            "affected_rows": match_id or "",
            "severity": "INFO",
            "notes": (
                "Match ID is retained as report context; analysis uses all "
                "available local historical score rows."
            ),
        },
        {
            "issue_type": "unresolved_disciplines",
            "affected_rows": len(unresolved_disciplines),
            "severity": "INFO" if not unresolved_disciplines else "REVIEW",
            "notes": ", ".join(unresolved_disciplines)
            if unresolved_disciplines else "No unresolved disciplines.",
        },
    ]


def _write_workbook(
    path: Path,
    summary_rows: list[dict[str, Any]],
    analysis_rows: list[dict[str, Any]],
    discipline_rows: list[dict[str, Any]],
    quality_rows: list[dict[str, Any]],
    target_name: str,
    gap_seconds: float,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_title_sheet(
        summary,
        "Ryatt Competitive Analysis",
        "Athletes faster than Ryatt plus athletes within configured slower gap.",
        SUMMARY_COLUMNS,
        summary_rows,
        1,
    )
    _write_title_sheet(
        workbook.create_sheet("Ryatt Competitive Analysis"),
        "Ryatt Competitive Analysis",
        "Athletes faster than Ryatt plus athletes within configured slower gap.",
        ANALYSIS_COLUMNS,
        analysis_rows,
        2,
        negative_gap=True,
    )
    _write_title_sheet(
        workbook.create_sheet("Discipline Summary"),
        "Discipline Summary",
        f"Discipline-level scouting summary for {target_name}; gap "
        f"window {gap_seconds:.2f} seconds.",
        DISCIPLINE_COLUMNS,
        discipline_rows,
        3,
    )
    _write_title_sheet(
        workbook.create_sheet("Data Quality Notes"),
        "Data Quality Notes",
        "Local data coverage and caveats for this scouting report.",
        QUALITY_COLUMNS,
        quality_rows,
        4,
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(path)
    except OSError as exc:
        raise RyattCompetitiveAnalysisError(
            f"Could not save workbook {path}: {exc}"
        ) from exc


def _write_title_sheet(
    sheet: Any,
    title: str,
    subtitle: str,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
    table_number: int,
    *,
    negative_gap: bool = False,
) -> None:
    sheet.sheet_view.showGridLines = False
    visible_rows = rows[:EXCEL_MAX_ROWS - 3]
    sheet["A1"] = title
    sheet["A1"].font = Font(size=18, bold=True, color="1F4E78")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(italic=True, color="666666")
    sheet.merge_cells(
        start_row=1, start_column=1, end_row=1,
        end_column=max(len(columns), 1)
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2,
        end_column=max(len(columns), 1)
    )
    sheet.append(list(columns))
    for cell in sheet[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in visible_rows:
        sheet.append([_workbook_value(column, row.get(column, ""))
                      for column in columns])
    if not visible_rows:
        sheet["A4"] = "No rows are currently available for this view."
        sheet["A4"].font = Font(italic=True, color="666666")
    sheet.freeze_panes = "A4"
    last_col = get_column_letter(max(len(columns), 1))
    table_end_row = 3 + max(len(visible_rows), 1)
    table_ref = f"A3:{last_col}{table_end_row}"
    sheet.auto_filter.ref = table_ref
    if visible_rows:
        table = Table(displayName=f"RyattTable{table_number}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    if len(rows) > len(visible_rows):
        note_row = sheet.max_row + 2
        sheet.cell(note_row, 1, (
            f"Excel row limit reached. This sheet shows {len(visible_rows):,} "
            f"of {len(rows):,} source rows; the complete CSV is in the same "
            "output folder."
        ))
        sheet.cell(note_row, 1).font = Font(italic=True, color="9C6500")
    for column_number, column in enumerate(columns, 1):
        letter = get_column_letter(column_number)
        values = [
            sheet.cell(row, column_number).value
            for row in range(4, sheet.max_row + 1)
        ]
        sheet.column_dimensions[letter].width = _column_width(column, values)
        if column in SECONDS_COLUMNS:
            for cell in sheet[letter][3:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
        if column in {"Notes", "Competitive Note", "notes"}:
            for cell in sheet[letter][3:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    if negative_gap and rows and "Gap To Ryatt" in columns:
        gap_col = get_column_letter(columns.index("Gap To Ryatt") + 1)
        sheet.conditional_formatting.add(
            f"{gap_col}4:{gap_col}{table_end_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=NEGATIVE_GAP_FILL,
            ),
        )


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=columns, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)
    except OSError as exc:
        raise RyattCompetitiveAnalysisError(
            f"Could not write {path}: {exc}"
        ) from exc


def _read_optional(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except (OSError, csv.Error) as exc:
        raise RyattCompetitiveAnalysisError(
            f"Could not read {path}: {exc}"
        ) from exc


def _same_athlete(
    row: dict[str, Any],
    athlete_id: str,
    athlete_name: str,
) -> bool:
    row_id = str(row.get("athlete_id", "")).strip()
    if athlete_id and row_id:
        return row_id == athlete_id
    return _normalized_name(row.get("athlete_name", "")) == _normalized_name(
        athlete_name
    )


def _is_placeholder(row: dict[str, Any]) -> bool:
    return (
        not str(row.get("athlete_name", "")).strip()
        or str(row.get("athlete_id", "")).strip() == PLACEHOLDER_ID
    )


def _division(value: str) -> str:
    return value.split("/", 1)[0].strip() if value else ""


def _normalized_name(value: str | None) -> str:
    return " ".join((value or "").casefold().split())


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _display(value: float | None) -> float | str:
    return round(value, 3) if value is not None else ""


def _workbook_value(column: str, value: Any) -> Any:
    if value == "":
        return None
    if column in {"Athlete ID", "target_athlete_id"}:
        return value
    numeric = _number(value)
    if numeric is not None and (
        column in SECONDS_COLUMNS
        or column in {
            "Number Of Matches",
            "Faster Athletes Count",
            "Within Gap Athletes Count",
            "affected_rows",
            "value",
        }
    ):
        return numeric
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    if column in {"Notes", "Competitive Note", "notes"}:
        return 55
    if column in {"Match With Best Score", "Most Recent Match"}:
        return 38
    if column in {"Athlete Name", "Team", "Closest Faster Athlete",
                  "Closest Slower Athlete"}:
        return 26
    longest = max(
        [len(column)]
        + [len(str(value)) for value in values if value is not None],
    )
    return min(max(longest + 2, 10), 30)
