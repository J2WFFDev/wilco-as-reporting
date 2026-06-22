"""Wilco historical analytics foundation built from local match outputs."""

from __future__ import annotations

import csv
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from wilco_as_reporting.raw_content import inspect_json_file
from wilco_as_reporting.team_profiles import TeamProfile

SOURCE_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "data_status",
    "status",
    "core_complete",
    "has_wilco_entries",
    "included_in_participation",
    "included_in_performance",
    "notes",
)

PARTICIPATION_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "athlete_name",
    "athlete_id",
    "discipline",
    "division",
    "class",
    "category",
    "gender",
    "squad_name",
    "squad_number",
    "role",
    "score",
    "rank",
    "rank_scope",
    "award_scope",
    "data_status",
    "notes",
)

ATHLETE_DISCIPLINE_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "discipline",
    "season_label",
    "matches_count",
    "scored_matches_count",
    "best_score",
    "average_score",
    "median_score",
    "latest_score",
    "first_score",
    "score_change_first_to_latest",
    "improvement_seconds",
    "improvement_percent",
    "best_match_id",
    "best_match_name",
    "latest_match_id",
    "latest_match_name",
    "notes",
)

ATHLETE_OVERALL_COLUMNS = (
    "athlete_name",
    "athlete_id",
    "season_label",
    "disciplines_count",
    "matches_count",
    "scored_entries_count",
    "best_discipline",
    "most_recent_match_date",
    "trend_note",
    "notes",
)

DISCIPLINE_SUMMARY_COLUMNS = (
    "season_label",
    "discipline",
    "athletes_count",
    "entries_count",
    "scored_entries_count",
    "team_best_score",
    "team_average_score",
    "team_median_score",
    "top_athlete_name",
    "most_improved_athlete_name",
    "notes",
)

AWARD_GAP_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "athlete_name",
    "discipline",
    "division",
    "class",
    "score",
    "rank",
    "award_scope",
    "rank_scope",
    "award_position_cutoff",
    "cutoff_score",
    "seconds_from_award",
    "award_gap_note",
    "notes",
)

STAGE_COLUMNS = (
    "match_id",
    "match_name",
    "match_date",
    "season_label",
    "athlete_name",
    "discipline",
    "stage_name",
    "stage_score",
    "stage_rank_if_available",
    "team_stage_best",
    "team_stage_average",
    "gap_to_team_best",
    "notes",
)

QUALITY_COLUMNS = (
    "match_id",
    "match_name",
    "data_status",
    "has_wilco_entries",
    "included_in_participation",
    "included_in_performance",
    "validation_error_count",
    "validation_warning_count",
    "validation_review_count",
    "quality_note",
    "notes",
)

SUMMARY_COLUMNS = ("metric", "value", "notes")

SECONDS_COLUMNS = {
    "score",
    "best_score",
    "average_score",
    "median_score",
    "latest_score",
    "first_score",
    "score_change_first_to_latest",
    "improvement_seconds",
    "team_best_score",
    "team_average_score",
    "team_median_score",
    "cutoff_score",
    "seconds_from_award",
    "stage_score",
    "team_stage_best",
    "team_stage_average",
    "gap_to_team_best",
}

OUTPUTS = (
    ("history_source_matches.csv", SOURCE_COLUMNS),
    ("wilco_match_participation.csv", PARTICIPATION_COLUMNS),
    (
        "wilco_athlete_discipline_history.csv",
        ATHLETE_DISCIPLINE_COLUMNS,
    ),
    ("wilco_athlete_overall_history.csv", ATHLETE_OVERALL_COLUMNS),
    ("wilco_discipline_summary.csv", DISCIPLINE_SUMMARY_COLUMNS),
    ("wilco_award_gap_history.csv", AWARD_GAP_COLUMNS),
    ("wilco_stage_benchmark_history.csv", STAGE_COLUMNS),
    ("wilco_data_quality_summary.csv", QUALITY_COLUMNS),
    ("wilco_history_summary.csv", SUMMARY_COLUMNS),
)


class HistoryBuildError(RuntimeError):
    """Raised when local history inputs or outputs cannot be processed."""


@dataclass(frozen=True)
class HistoryBuildResult:
    output_dir: Path
    workbook_path: Path | None
    row_counts: dict[str, int]
    source_matches: int
    wilco_matches: int
    athletes: int
    disciplines: int
    scored_entries: int
    warnings: tuple[str, ...]


def build_history(
    *,
    output_root: Path | str,
    profile: TeamProfile,
    match_ids: tuple[int, ...] = (),
    match_ids_file: Path | str | None = None,
    match_index: Path | str | None = None,
    backfill_results: Path | str | None = None,
    include_partial: bool = True,
    include_no_scores: bool = True,
    exclude_no_scores_from_performance: bool = True,
    min_athlete_matches: int = 2,
    workbook: bool = True,
) -> HistoryBuildResult:
    """Build Wilco historical tables from existing local output only."""
    if min_athlete_matches < 1:
        raise HistoryBuildError("min_athlete_matches must be at least 1.")
    root = Path(output_root)
    index_path = _default_existing_path(
        match_index,
        root / "discovery" / "tables" / "effective_match_index.csv",
    )
    backfill_path = _default_existing_path(
        backfill_results,
        root / "backfill" / "backfill_results.csv",
    )
    selected_ids = _select_match_ids(
        root,
        match_ids,
        Path(match_ids_file) if match_ids_file else None,
        index_path,
        backfill_path,
    )
    metadata = _metadata_index(index_path)
    statuses = _status_index(backfill_path, root)
    warnings: list[str] = []
    source_rows: list[dict[str, Any]] = []
    participation: list[dict[str, Any]] = []
    award_gaps: list[dict[str, Any]] = []
    stage_rows: list[dict[str, Any]] = []
    quality_rows: list[dict[str, Any]] = []

    for match_id in selected_ids:
        match_dir = root / str(match_id)
        match_scores_path = match_dir / "tables" / "match_scores.csv"
        rankings_path = match_dir / "tables" / "rankings.csv"
        stages_path = match_dir / "tables" / "stage_scores.csv"
        squads_path = (
            match_dir
            / "team_report_tables"
            / profile.team_key
            / "wilco_squad_summary.csv"
        )
        source_meta = metadata.get(match_id, {})
        status_meta = statuses.get(match_id, {})
        match_scores = _read_optional_csv(match_scores_path, warnings)
        rankings = _read_optional_csv(rankings_path, warnings)
        stages = _read_optional_csv(stages_path, warnings)
        squads = _read_optional_csv(squads_path, warnings)
        team_scores = [
            row
            for row in match_scores
            if profile.matches_name(row.get("team_name"))
        ]
        team_stages = [
            row
            for row in stages
            if profile.matches_name(row.get("team_name"))
        ]
        match_name = (
            source_meta.get("name")
            or _first_value(team_scores, "match_name")
            or _first_value(match_scores, "match_name")
            or status_meta.get("match_name")
            or f"Match {match_id}"
        )
        match_date = (
            source_meta.get("start_date")
            or _match_date_from_rows(team_scores)
        )
        season = season_label(match_date)
        core_complete = _core_complete(match_dir, match_id)
        data_status = (
            status_meta.get("data_status")
            or _derive_data_status(team_scores)
        )
        status = status_meta.get("status") or (
            "LOCAL_OUTPUT" if match_scores_path.exists() else "MISSING_OUTPUT"
        )
        has_entries = bool(team_scores)
        include_participation = has_entries and _participation_allowed(
            data_status,
            include_partial,
            include_no_scores,
        )
        scored_team_rows = [
            row
            for row in team_scores
            if _number(row.get("match_score_seconds")) is not None
        ]
        include_performance = (
            include_participation
            and bool(scored_team_rows)
            and not (
                exclude_no_scores_from_performance
                and data_status == "no_scores"
            )
        )
        source_notes = _source_notes(
            match_scores_path,
            data_status,
            has_entries,
            include_performance,
        )
        source_rows.append(
            {
                "match_id": match_id,
                "match_name": match_name,
                "match_date": match_date,
                "season_label": season,
                "data_status": data_status,
                "status": status,
                "core_complete": _boolean(core_complete),
                "has_wilco_entries": _boolean(has_entries),
                "included_in_participation": _boolean(
                    include_participation
                ),
                "included_in_performance": _boolean(
                    include_performance
                ),
                "notes": source_notes,
            }
        )
        if include_participation:
            rank_index = _preferred_rank_index(rankings, profile)
            squad_index = _squad_membership_index(squads)
            for score_row in team_scores:
                participation.append(
                    _participation_row(
                        match_id,
                        match_name,
                        match_date,
                        season,
                        data_status,
                        score_row,
                        rank_index,
                        squad_index,
                        include_performance,
                    )
                )
        if include_performance:
            award_gaps.extend(
                _award_gap_rows(
                    match_id,
                    match_name,
                    match_date,
                    season,
                    rankings,
                    profile,
                )
            )
            stage_rows.extend(
                _stage_benchmark_rows(
                    match_id,
                    match_name,
                    match_date,
                    season,
                    team_stages,
                )
            )
        quality_rows.append(
            _quality_row(
                match_id,
                match_name,
                data_status,
                has_entries,
                include_participation,
                include_performance,
                status_meta,
                match_scores_path.exists(),
            )
        )

    athlete_discipline = _athlete_discipline_history(
        participation,
        min_athlete_matches,
    )
    athlete_overall = _athlete_overall_history(participation)
    discipline_summary = _discipline_summary(
        participation,
        athlete_discipline,
    )
    summary_rows = _history_summary(
        source_rows,
        participation,
    )
    all_rows = {
        "history_source_matches.csv": source_rows,
        "wilco_match_participation.csv": participation,
        "wilco_athlete_discipline_history.csv": athlete_discipline,
        "wilco_athlete_overall_history.csv": athlete_overall,
        "wilco_discipline_summary.csv": discipline_summary,
        "wilco_award_gap_history.csv": award_gaps,
        "wilco_stage_benchmark_history.csv": stage_rows,
        "wilco_data_quality_summary.csv": quality_rows,
        "wilco_history_summary.csv": summary_rows,
    }
    history_dir = root / "history"
    for filename, columns in OUTPUTS:
        _write_csv(history_dir / filename, columns, all_rows[filename])
    workbook_path = (
        _build_workbook(history_dir, all_rows)
        if workbook
        else None
    )
    athletes = {
        _athlete_key(row)
        for row in participation
        if row.get("athlete_name")
    }
    disciplines = {
        row["discipline"]
        for row in participation
        if row.get("discipline")
    }
    return HistoryBuildResult(
        output_dir=history_dir,
        workbook_path=workbook_path,
        row_counts={
            filename: len(all_rows[filename])
            for filename, _ in OUTPUTS
        },
        source_matches=len(source_rows),
        wilco_matches=sum(
            row["has_wilco_entries"] == "true"
            for row in source_rows
        ),
        athletes=len(athletes),
        disciplines=len(disciplines),
        scored_entries=sum(
            _number(row.get("score")) is not None
            for row in participation
        ),
        warnings=tuple(dict.fromkeys(warnings)),
    )


def season_label(value: str | date | datetime | None) -> str:
    """Return the September-through-August shooting season label."""
    parsed = _date_value(value)
    if parsed is None:
        return "Unknown Season"
    start_year = parsed.year if parsed.month >= 9 else parsed.year - 1
    return f"{start_year % 100:02d}-{(start_year + 1) % 100:02d} Season"


def _participation_row(
    match_id: int,
    match_name: str,
    match_date: str,
    season: str,
    data_status: str,
    row: dict[str, str],
    rank_index: dict[tuple[str, str], dict[str, str]],
    squad_index: dict[tuple[str, str], str],
    include_performance: bool,
) -> dict[str, Any]:
    athlete_name = row.get("athlete_name", "")
    discipline = row.get("discipline", "")
    ranking = rank_index.get((athlete_name, discipline), {})
    score = _number(row.get("match_score_seconds"))
    notes: list[str] = []
    if score is None:
        notes.append("No scored match result is currently available.")
    if not include_performance:
        notes.append("Entry is excluded from performance trends.")
    if not ranking:
        notes.append("No preferred individual ranking row was available.")
    return {
        "match_id": match_id,
        "match_name": match_name,
        "match_date": match_date,
        "season_label": season,
        "athlete_name": athlete_name,
        "athlete_id": row.get("athlete_id", ""),
        "discipline": discipline,
        "division": _division(row.get("class", "")),
        "class": row.get("class", ""),
        "category": "",
        "gender": row.get("gender", ""),
        "squad_name": squad_index.get((athlete_name, discipline), ""),
        "squad_number": "",
        "role": "",
        "score": _display(score),
        "rank": ranking.get("place", ""),
        "rank_scope": ranking.get("rank_scope", ""),
        "award_scope": ranking.get("award_scope", ""),
        "data_status": data_status,
        "notes": " ".join(notes),
    }


def _preferred_rank_index(
    rankings: list[dict[str, str]],
    profile: TeamProfile,
) -> dict[tuple[str, str], dict[str, str]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rankings:
        if (
            row.get("leaderboard_type") == "athlete"
            and profile.matches_name(row.get("team_name"))
        ):
            grouped[
                (
                    row.get("athlete_name", ""),
                    row.get("discipline", ""),
                )
            ].append(row)
    result: dict[tuple[str, str], dict[str, str]] = {}
    for key, rows in grouped.items():
        result[key] = min(
            rows,
            key=lambda row: (
                0 if row.get("award_scope") == "Class" else 1,
                0 if row.get("award_scope") != "Comparison" else 1,
                _integer(row.get("place")) or 10**9,
                row.get("leaderboard_name", ""),
            ),
        )
    return result


def _squad_membership_index(
    squads: list[dict[str, str]],
) -> dict[tuple[str, str], str]:
    result: dict[tuple[str, str], str] = {}
    for row in squads:
        discipline = row.get("discipline", "")
        squad_name = row.get("squad_name", "")
        for number in range(1, 5):
            athlete_name = row.get(f"athlete_{number}_name", "")
            if athlete_name:
                result.setdefault(
                    (athlete_name, discipline),
                    squad_name,
                )
    return result


def _athlete_discipline_history(
    participation: list[dict[str, Any]],
    min_matches: int,
) -> list[dict[str, Any]]:
    grouped: dict[
        tuple[str, str, str, str],
        list[dict[str, Any]],
    ] = defaultdict(list)
    for row in participation:
        grouped[
            (
                row.get("athlete_name", ""),
                row.get("athlete_id", ""),
                row.get("discipline", ""),
                row.get("season_label", ""),
            )
        ].append(row)
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        match_count = len({row["match_id"] for row in rows})
        if match_count < min_matches:
            continue
        scored = [
            row
            for row in rows
            if _number(row.get("score")) is not None
        ]
        ordered = sorted(
            scored,
            key=lambda row: (
                row.get("match_date") or "9999-12-31",
                int(row["match_id"]),
            ),
        )
        scores = [_number(row["score"]) for row in scored]
        numeric_scores = [score for score in scores if score is not None]
        first = _number(ordered[0]["score"]) if ordered else None
        latest = _number(ordered[-1]["score"]) if ordered else None
        best_row = (
            min(scored, key=lambda row: _number(row["score"]) or 10**9)
            if scored
            else {}
        )
        improvement = (
            first - latest
            if first is not None and latest is not None
            else None
        )
        improvement_percent = (
            improvement / first * 100
            if improvement is not None and first
            else None
        )
        notes = ""
        if not scored:
            notes = "No scored matches are available for this season."
        results.append(
            {
                "athlete_name": key[0],
                "athlete_id": key[1],
                "discipline": key[2],
                "season_label": key[3],
                "matches_count": match_count,
                "scored_matches_count": len(
                    {row["match_id"] for row in scored}
                ),
                "best_score": _display(
                    min(numeric_scores) if numeric_scores else None
                ),
                "average_score": _display(
                    statistics.fmean(numeric_scores)
                    if numeric_scores
                    else None
                ),
                "median_score": _display(
                    statistics.median(numeric_scores)
                    if numeric_scores
                    else None
                ),
                "latest_score": _display(latest),
                "first_score": _display(first),
                "score_change_first_to_latest": _display(
                    latest - first
                    if first is not None and latest is not None
                    else None
                ),
                "improvement_seconds": _display(improvement),
                "improvement_percent": _display(improvement_percent),
                "best_match_id": best_row.get("match_id", ""),
                "best_match_name": best_row.get("match_name", ""),
                "latest_match_id": (
                    ordered[-1].get("match_id", "") if ordered else ""
                ),
                "latest_match_name": (
                    ordered[-1].get("match_name", "") if ordered else ""
                ),
                "notes": notes,
            }
        )
    return sorted(
        results,
        key=lambda row: (
            row["season_label"],
            row["athlete_name"].casefold(),
            row["discipline"].casefold(),
        ),
    )


def _athlete_overall_history(
    participation: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[
        tuple[str, str, str],
        list[dict[str, Any]],
    ] = defaultdict(list)
    for row in participation:
        grouped[
            (
                row.get("athlete_name", ""),
                row.get("athlete_id", ""),
                row.get("season_label", ""),
            )
        ].append(row)
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        scored = [
            row
            for row in rows
            if _number(row.get("score")) is not None
        ]
        discipline_scores: dict[str, list[float]] = defaultdict(list)
        for row in scored:
            score = _number(row["score"])
            if score is not None:
                discipline_scores[row["discipline"]].append(score)
        best_discipline = (
            min(
                discipline_scores,
                key=lambda discipline: statistics.fmean(
                    discipline_scores[discipline]
                ),
            )
            if discipline_scores
            else ""
        )
        latest_date = max(
            (row.get("match_date", "") for row in rows),
            default="",
        )
        trend = _overall_trend_note(scored)
        results.append(
            {
                "athlete_name": key[0],
                "athlete_id": key[1],
                "season_label": key[2],
                "disciplines_count": len(
                    {row["discipline"] for row in rows if row["discipline"]}
                ),
                "matches_count": len({row["match_id"] for row in rows}),
                "scored_entries_count": len(scored),
                "best_discipline": best_discipline,
                "most_recent_match_date": latest_date,
                "trend_note": trend,
                "notes": (
                    "Cross-discipline scores are not combined into one "
                    "numeric performance metric."
                ),
            }
        )
    return sorted(
        results,
        key=lambda row: (
            row["season_label"],
            row["athlete_name"].casefold(),
        ),
    )


def _discipline_summary(
    participation: list[dict[str, Any]],
    athlete_discipline: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in participation:
        grouped[(row["season_label"], row["discipline"])].append(row)
    improvements: dict[tuple[str, str], dict[str, Any]] = {}
    for row in athlete_discipline:
        improvement = _number(row["improvement_seconds"])
        if improvement is None:
            continue
        key = (row["season_label"], row["discipline"])
        current = improvements.get(key)
        if current is None or improvement > (
            _number(current["improvement_seconds"]) or -10**9
        ):
            improvements[key] = row
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        scored = [
            row
            for row in rows
            if _number(row.get("score")) is not None
        ]
        scores = [
            _number(row["score"])
            for row in scored
            if _number(row["score"]) is not None
        ]
        top = (
            min(scored, key=lambda row: _number(row["score"]) or 10**9)
            if scored
            else {}
        )
        improved = improvements.get(key, {})
        results.append(
            {
                "season_label": key[0],
                "discipline": key[1],
                "athletes_count": len(
                    {_athlete_key(row) for row in rows}
                ),
                "entries_count": len(rows),
                "scored_entries_count": len(scored),
                "team_best_score": _display(min(scores) if scores else None),
                "team_average_score": _display(
                    statistics.fmean(scores) if scores else None
                ),
                "team_median_score": _display(
                    statistics.median(scores) if scores else None
                ),
                "top_athlete_name": top.get("athlete_name", ""),
                "most_improved_athlete_name": improved.get(
                    "athlete_name",
                    "",
                ),
                "notes": (
                    "Most improved requires the configured minimum match "
                    "count within the season."
                ),
            }
        )
    return sorted(results, key=lambda row: (row["season_label"], row["discipline"]))


def _award_gap_rows(
    match_id: int,
    match_name: str,
    match_date: str,
    season: str,
    rankings: list[dict[str, str]],
    profile: TeamProfile,
) -> list[dict[str, Any]]:
    boards: dict[
        tuple[str, str, str, str],
        list[dict[str, str]],
    ] = defaultdict(list)
    for row in rankings:
        if row.get("leaderboard_type") == "athlete":
            boards[_leaderboard_key(row)].append(row)
    results: list[dict[str, Any]] = []
    for row in rankings:
        if (
            row.get("leaderboard_type") != "athlete"
            or not profile.matches_name(row.get("team_name"))
            or row.get("award_scope") == "Comparison"
        ):
            continue
        cutoff_place = _integer(row.get("award_places"))
        score = _number(row.get("score_seconds"))
        cutoff_row = next(
            (
                candidate
                for candidate in boards[_leaderboard_key(row)]
                if _integer(candidate.get("place")) == cutoff_place
            ),
            None,
        )
        cutoff_score = (
            _number(cutoff_row.get("score_seconds"))
            if cutoff_row
            else None
        )
        rank = _integer(row.get("place"))
        gap = (
            max(score - cutoff_score, 0.0)
            if score is not None and cutoff_score is not None
            else None
        )
        if cutoff_place is None:
            note = "Award place count is unavailable for this board."
        elif cutoff_score is None:
            note = "The cutoff-ranked score is unavailable."
        elif rank is not None and rank <= cutoff_place:
            note = "Athlete was inside the listed award positions."
        else:
            note = "Positive seconds indicate time above the award cutoff."
        results.append(
            {
                "match_id": match_id,
                "match_name": match_name,
                "match_date": match_date,
                "season_label": season,
                "athlete_name": row.get("athlete_name", ""),
                "discipline": row.get("discipline", ""),
                "division": _division(row.get("class", "")),
                "class": row.get("class", ""),
                "score": _display(score),
                "rank": row.get("place", ""),
                "award_scope": row.get("award_scope", ""),
                "rank_scope": row.get("rank_scope", ""),
                "award_position_cutoff": _display(cutoff_place),
                "cutoff_score": _display(cutoff_score),
                "seconds_from_award": _display(gap),
                "award_gap_note": note,
                "notes": (
                    f"Leaderboard: {row.get('leaderboard_name', '')}"
                ),
            }
        )
    return results


def _stage_benchmark_rows(
    match_id: int,
    match_name: str,
    match_date: str,
    season: str,
    stages: list[dict[str, str]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in stages:
        grouped[
            (row.get("discipline", ""), row.get("stage_name", ""))
        ].append(row)
    results: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        scored = [
            (row, _number(row.get("stage_score_seconds")))
            for row in rows
            if _number(row.get("stage_score_seconds")) is not None
        ]
        values = [score for _, score in scored if score is not None]
        best = min(values) if values else None
        average = statistics.fmean(values) if values else None
        sorted_values = sorted(set(values))
        for row in rows:
            score = _number(row.get("stage_score_seconds"))
            rank = (
                sorted_values.index(score) + 1
                if score is not None and score in sorted_values
                else ""
            )
            results.append(
                {
                    "match_id": match_id,
                    "match_name": match_name,
                    "match_date": match_date,
                    "season_label": season,
                    "athlete_name": row.get("athlete_name", ""),
                    "discipline": key[0],
                    "stage_name": key[1],
                    "stage_score": _display(score),
                    "stage_rank_if_available": rank,
                    "team_stage_best": _display(best),
                    "team_stage_average": _display(average),
                    "gap_to_team_best": _display(
                        score - best
                        if score is not None and best is not None
                        else None
                    ),
                    "notes": (
                        "Stage rank is calculated within Wilco entries "
                        "for this match, discipline, and stage."
                    ),
                }
            )
    return results


def _quality_row(
    match_id: int,
    match_name: str,
    data_status: str,
    has_entries: bool,
    include_participation: bool,
    include_performance: bool,
    status_meta: dict[str, str],
    has_tables: bool,
) -> dict[str, Any]:
    notes: list[str] = []
    if not has_tables:
        notes.append("Parsed match score table is missing.")
    if data_status == "no_team_entries":
        notes.append("Wilco did not participate; this is not a failure.")
    elif data_status == "no_scores":
        notes.append(
            "Participation may be present, but no scores are available."
        )
    elif data_status == "partial":
        notes.append("Partial data is retained with performance safeguards.")
    error_count = _integer(status_meta.get("validation_error_count")) or 0
    warning_count = (
        _integer(status_meta.get("validation_warning_count")) or 0
    )
    review_count = (
        _integer(status_meta.get("validation_review_count")) or 0
    )
    quality_note = (
        "Historical validation counts are context, not coaching conclusions."
    )
    return {
        "match_id": match_id,
        "match_name": match_name,
        "data_status": data_status,
        "has_wilco_entries": _boolean(has_entries),
        "included_in_participation": _boolean(include_participation),
        "included_in_performance": _boolean(include_performance),
        "validation_error_count": error_count,
        "validation_warning_count": warning_count,
        "validation_review_count": review_count,
        "quality_note": quality_note,
        "notes": " ".join(notes),
    }


def _history_summary(
    sources: list[dict[str, Any]],
    participation: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    metrics = (
        ("total_source_matches", len(sources), ""),
        (
            "wilco_matches_found",
            sum(row["has_wilco_entries"] == "true" for row in sources),
            "",
        ),
        (
            "seasons_found",
            len(
                {
                    row["season_label"]
                    for row in sources
                    if row["season_label"] != "Unknown Season"
                }
            ),
            "",
        ),
        (
            "athletes_found",
            len(
                {
                    _athlete_key(row)
                    for row in participation
                    if row.get("athlete_name")
                }
            ),
            "",
        ),
        (
            "disciplines_found",
            len({row["discipline"] for row in participation}),
            "",
        ),
        (
            "scored_entries",
            sum(
                _number(row.get("score")) is not None
                for row in participation
            ),
            "No-score participation entries are excluded.",
        ),
        (
            "partial_matches",
            sum(row["data_status"] == "partial" for row in sources),
            "",
        ),
        (
            "no_score_matches",
            sum(row["data_status"] == "no_scores" for row in sources),
            "",
        ),
        (
            "excluded_from_performance",
            sum(
                row["included_in_performance"] == "false"
                for row in sources
            ),
            "",
        ),
        (
            "generated_at",
            datetime.now().astimezone().isoformat(timespec="seconds"),
            "Local generation time.",
        ),
    )
    return [
        {"metric": metric, "value": value, "notes": notes}
        for metric, value, notes in metrics
    ]


def _build_workbook(
    history_dir: Path,
    all_rows: dict[str, list[dict[str, Any]]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise HistoryBuildError(
            "openpyxl is required for --workbook."
        ) from exc
    sheets = (
        ("Summary", "wilco_history_summary.csv", SUMMARY_COLUMNS),
        ("Source Matches", "history_source_matches.csv", SOURCE_COLUMNS),
        (
            "Match Participation",
            "wilco_match_participation.csv",
            PARTICIPATION_COLUMNS,
        ),
        (
            "Athlete Discipline History",
            "wilco_athlete_discipline_history.csv",
            ATHLETE_DISCIPLINE_COLUMNS,
        ),
        (
            "Athlete Overall History",
            "wilco_athlete_overall_history.csv",
            ATHLETE_OVERALL_COLUMNS,
        ),
        (
            "Discipline Summary",
            "wilco_discipline_summary.csv",
            DISCIPLINE_SUMMARY_COLUMNS,
        ),
        ("Award Gaps", "wilco_award_gap_history.csv", AWARD_GAP_COLUMNS),
        (
            "Stage Benchmarks",
            "wilco_stage_benchmark_history.csv",
            STAGE_COLUMNS,
        ),
        (
            "Data Quality",
            "wilco_data_quality_summary.csv",
            QUALITY_COLUMNS,
        ),
    )
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_number, (name, filename, columns) in enumerate(sheets, 1):
        ws = wb.create_sheet(name)
        ws.append([column.replace("_", " ").title() for column in columns])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
        for row in all_rows[filename]:
            ws.append([_workbook_value(row.get(column, "")) for column in columns])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        if all_rows[filename]:
            table = Table(
                displayName=f"HistoryTable{sheet_number}",
                ref=ws.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        for column_number, column in enumerate(columns, 1):
            letter = get_column_letter(column_number)
            width = _column_width(
                column,
                (
                    ws.cell(row=row_number, column=column_number).value
                    for row_number in range(2, ws.max_row + 1)
                ),
            )
            ws.column_dimensions[letter].width = width
            if column in {"notes", "quality_note", "award_gap_note"}:
                for cell in ws[letter][1:]:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
            if column in SECONDS_COLUMNS:
                for cell in ws[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
            if column == "improvement_percent":
                for cell in ws[letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0"
    path = history_dir / "wilco_history_report.xlsx"
    try:
        wb.save(path)
    except OSError as exc:
        raise HistoryBuildError(
            f"Could not save history workbook {path}: {exc}"
        ) from exc
    return path


def _select_match_ids(
    root: Path,
    match_ids: tuple[int, ...],
    match_ids_file: Path | None,
    match_index: Path | None,
    backfill_results: Path | None,
) -> tuple[int, ...]:
    if match_ids:
        return tuple(sorted(set(match_ids)))
    if match_ids_file:
        try:
            values = [
                int(line.strip())
                for line in match_ids_file.read_text(
                    encoding="utf-8-sig"
                ).splitlines()
                if line.strip()
            ]
        except (OSError, ValueError) as exc:
            raise HistoryBuildError(
                f"Could not read match ID file {match_ids_file}: {exc}"
            ) from exc
        return tuple(sorted(set(values)))
    if backfill_results:
        rows = _read_csv(backfill_results)
        ids = {
            int(row["match_id"])
            for row in rows
            if row.get("match_id", "").isdigit()
            and row.get("status") in {"SUCCESS", "SKIPPED_UNCHANGED"}
        }
        if ids:
            return tuple(sorted(ids))
    if match_index:
        rows = _read_csv(match_index)
        return tuple(
            sorted(
                {
                    int(row["match_id"])
                    for row in rows
                    if row.get("match_id", "").isdigit()
                }
            )
        )
    return tuple(
        sorted(
            int(child.name)
            for child in root.iterdir()
            if child.is_dir() and child.name.isdigit()
        )
    )


def _metadata_index(path: Path | None) -> dict[int, dict[str, str]]:
    if path is None:
        return {}
    return {
        int(row["match_id"]): row
        for row in _read_csv(path)
        if row.get("match_id", "").isdigit()
    }


def _status_index(
    backfill_path: Path | None,
    root: Path,
) -> dict[int, dict[str, str]]:
    result: dict[int, dict[str, str]] = {}
    if backfill_path:
        for row in _read_csv(backfill_path):
            if not row.get("match_id", "").isdigit():
                continue
            result[int(row["match_id"])] = {
                "match_name": row.get("match_name", ""),
                "status": row.get("status", ""),
                "data_status": row.get("data_status", ""),
                "validation_error_count": row.get(
                    "validation_error_count",
                    "",
                ),
                "validation_warning_count": row.get(
                    "validation_warning_count",
                    "",
                ),
                "validation_review_count": row.get(
                    "validation_review_count",
                    "",
                ),
            }
    manifest = root / "state" / "match_refresh_manifest.csv"
    if manifest.exists():
        for row in _read_csv(manifest):
            if not row.get("match_id", "").isdigit():
                continue
            match_id = int(row["match_id"])
            current = result.setdefault(match_id, {})
            current.setdefault("match_name", row.get("match_name", ""))
            current.setdefault("status", row.get("last_status", ""))
            current.setdefault(
                "data_status",
                row.get("last_data_status", ""),
            )
            for column in (
                "validation_error_count",
                "validation_warning_count",
                "validation_review_count",
            ):
                current.setdefault(column, row.get(column, ""))
    return result


def _default_existing_path(
    supplied: Path | str | None,
    default: Path,
) -> Path | None:
    if supplied:
        path = Path(supplied)
        if not path.exists():
            raise HistoryBuildError(f"Input file does not exist: {path}")
        return path
    return default if default.exists() else None


def _core_complete(match_dir: Path, match_id: int) -> bool:
    raw = match_dir / "raw"
    return (
        inspect_json_file(raw / f"{match_id}_slots.json").useful_content
        and inspect_json_file(
            raw / f"{match_id}_leaderboard.json"
        ).useful_content
    )


def _derive_data_status(team_scores: list[dict[str, str]]) -> str:
    if not team_scores:
        return "no_team_entries"
    scored = sum(
        _number(row.get("match_score_seconds")) is not None
        for row in team_scores
    )
    if scored == 0:
        return "no_scores"
    if scored < len(team_scores):
        return "partial"
    return "complete"


def _participation_allowed(
    data_status: str,
    include_partial: bool,
    include_no_scores: bool,
) -> bool:
    if data_status == "partial" and not include_partial:
        return False
    if data_status == "no_scores" and not include_no_scores:
        return False
    return data_status != "no_team_entries"


def _source_notes(
    match_scores_path: Path,
    data_status: str,
    has_entries: bool,
    performance: bool,
) -> str:
    notes: list[str] = []
    if not match_scores_path.exists():
        notes.append("Parsed match score table is unavailable.")
    if not has_entries:
        notes.append("No Wilco entries were found; this is not a failure.")
    if data_status == "no_scores":
        notes.append("Participation retained without score trends.")
    if data_status == "partial":
        notes.append("Partial match data retained.")
    if has_entries and not performance:
        notes.append("Excluded from performance calculations.")
    return " ".join(notes)


def _overall_trend_note(rows: list[dict[str, Any]]) -> str:
    dates = sorted(
        {row.get("match_date", "") for row in rows if row.get("match_date")}
    )
    if len(dates) < 2:
        return "Insufficient scored match dates for a trend."
    return (
        "Review discipline-level trends; scores across disciplines are "
        "not directly comparable."
    )


def _division(class_value: str) -> str:
    return class_value.split("/", 1)[0] if class_value else ""


def _match_date_from_rows(rows: list[dict[str, str]]) -> str:
    return _first_value(rows, "match_date")


def _first_value(rows: list[dict[str, str]], column: str) -> str:
    return next(
        (row.get(column, "") for row in rows if row.get(column)),
        "",
    )


def _athlete_key(row: dict[str, Any]) -> tuple[str, str]:
    return (
        str(row.get("athlete_id", "")),
        str(row.get("athlete_name", "")),
    )


def _leaderboard_key(
    row: dict[str, str],
) -> tuple[str, str, str, str]:
    return (
        row.get("discipline_id") or row.get("discipline", ""),
        row.get("leaderboard_name", ""),
        row.get("rank_scope", ""),
        row.get("award_scope", ""),
    )


def _date_value(value: str | date | datetime | None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except ValueError:
        return None


def _number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _display(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        rounded = round(value, 3)
        return int(rounded) if rounded.is_integer() else rounded
    return value


def _boolean(value: bool) -> str:
    return str(bool(value)).lower()


def _workbook_value(value: Any) -> Any:
    if value == "":
        return None
    if isinstance(value, str):
        lowered = value.casefold()
        if lowered in {"true", "false"}:
            return lowered == "true"
    return value


def _column_width(column: str, values: Iterable[Any]) -> float:
    preferred = {
        "match_name": 42,
        "athlete_name": 24,
        "discipline": 26,
        "notes": 60,
        "quality_note": 55,
        "award_gap_note": 48,
    }
    if column in preferred:
        return preferred[column]
    longest = len(column.replace("_", " ").title())
    for value in values:
        if value is not None:
            longest = max(longest, len(str(value)))
    return min(max(longest + 2, 10), 32)


def _read_optional_csv(
    path: Path,
    warnings: list[str],
) -> list[dict[str, str]]:
    if not path.exists():
        warnings.append(f"Missing optional history input: {path}")
        return []
    return _read_csv(path)


def _read_csv(path: Path) -> list[dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as source:
            return list(csv.DictReader(source))
    except (OSError, csv.Error) as exc:
        raise HistoryBuildError(f"Could not read {path}: {exc}") from exc


def _write_csv(
    path: Path,
    columns: Iterable[str],
    rows: Iterable[dict[str, Any]],
) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as target:
            writer = csv.DictWriter(
                target,
                fieldnames=columns,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(rows)
    except (OSError, csv.Error) as exc:
        raise HistoryBuildError(f"Could not write {path}: {exc}") from exc
